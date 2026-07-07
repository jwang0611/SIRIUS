#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
iSDTaiM 端到端 Smoke Test
=========================

快速验证完整流程是否能跑通。步骤之间自动串联——Step1 的输出直接喂给 Step3，
Step3 的输出喂给 Step4。用户只需提供源数据即可。

用法:
    # 使用内置合成数据（自动生成，零配置）
    python tests/smoke_test.py

    # 指定自己的测试数据目录
    python tests/smoke_test.py --data-dir path/to/my_test

    # 只跑离线步骤（不需要 API Key / 不调 LLM）
    python tests/smoke_test.py --offline

    # 只跑 Web API 测试
    python tests/smoke_test.py --web-only

    # 跳过清理（保留测试产物供检查）
    python tests/smoke_test.py --keep

数据目录只需包含以下源文件（中间产物会自动串联生成）:
    <data-dir>/
    ├── ecrf_raw.xlsx          # [必需] eCRF 原始 Excel
    ├── *template*.xlsx        # [Step4 需要] SDTM Spec 模板
    └── als2sdtm_example.xlsx  # [可选] 历史映射样例，用于知识库构建

文件名匹配规则（不需要严格命名）:
    - eCRF 源文件:  ecrf_raw.xlsx (固定名)
    - 模板文件:     文件名含 "template" 或 "spec" 的 .xlsx
    - ALS 映射样例: als2sdtm_example.xlsx (固定名)
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from pathlib import Path

try:
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
PYTHON = sys.executable

PASS = "\033[92m[PASS]\033[0m"
FAIL = "\033[91m[FAIL]\033[0m"
SKIP = "\033[93m[SKIP]\033[0m"
INFO = "\033[94m[INFO]\033[0m"

ALS_REQUIRED_COLUMNS = {"SDTM_Domain", "SDTM_Variable"}
ALS_TABLE_COLUMNS = {"表名", "表名称", "annotation_table", "表"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


class TestResult:
    def __init__(self, name: str):
        self.name = name
        self.status = "pending"
        self.message = ""
        self.duration = 0.0
        self.outputs: list[Path] = []

    def passed(self, msg="", duration=0.0, outputs=None):
        self.status = "pass"
        self.message = msg
        self.duration = duration
        self.outputs = outputs or []

    def failed(self, msg="", duration=0.0):
        self.status = "fail"
        self.message = msg
        self.duration = duration

    def skipped(self, msg=""):
        self.status = "skip"
        self.message = msg


class PipelineContext:
    """Holds intermediate outputs for cross-step chaining."""

    def __init__(self):
        self.step1_json: Path | None = None
        self.step3_als_output: Path | None = None


def run_cmd(args: list, cwd=None, timeout=300) -> subprocess.CompletedProcess:
    """Run a command and capture output (Windows GBK-safe)."""
    result = subprocess.run(
        args,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=cwd or PROJECT_ROOT,
        timeout=timeout,
        env={**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONLEGACYWINDOWSSTDIO": "utf-8"},
    )
    result.stdout = result.stdout.decode("utf-8", errors="replace") if result.stdout else ""
    result.stderr = result.stderr.decode("utf-8", errors="replace") if result.stderr else ""
    return result


def print_result(result: TestResult):
    tag = {
        "pass": PASS,
        "fail": FAIL,
        "skip": SKIP,
    }.get(result.status, INFO)
    time_str = f" ({result.duration:.1f}s)" if result.duration > 0 else ""
    print(f"  {tag} {result.name}{time_str}")
    if result.message:
        for line in result.message.split("\n"):
            print(f"       {line}")
    if result.outputs:
        for p in result.outputs:
            print(f"       -> {p}")


def find_file(
    data_dir: Path, exact_name: str | None = None, patterns: list[str] | None = None, suffix: str = ".xlsx"
) -> Path | None:
    """Find a file by exact name or fuzzy pattern match."""
    if exact_name:
        p = data_dir / exact_name
        if p.exists():
            return p

    if patterns:
        for f in sorted(data_dir.iterdir()):
            if f.is_file() and f.suffix.lower() == suffix:
                name_lower = f.name.lower()
                if any(pat.lower() in name_lower for pat in patterns):
                    return f
    return None


def detect_als_sheet(excel_path: Path) -> str | None:
    """Auto-detect which sheet in the Excel has ALS-compatible columns."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = set()
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                if cell is not None:
                    headers.add(str(cell).strip())
            if headers & ALS_REQUIRED_COLUMNS and headers & ALS_TABLE_COLUMNS:
                wb.close()
                return sheet_name
        wb.close()
    except Exception:
        pass
    return None


def convert_llm_output_to_spec_mapper_format(src_xlsx: Path, dst_xlsx: Path) -> Path:
    """
    Convert Step3b LLM output (snake_case columns) to SpecMapper-compatible format
    (with Chinese columns).

    Source columns:  metadata_table, metadata_variable, annotation_table,
                     annotation_variable, SDTM_Domain, SDTM_Variable, ...
    Target columns:  表, 变量, 变量名, SDTM_Domain, SDTM_Variable, metadata_table, ...
    """
    import pandas as pd

    df = pd.read_excel(src_xlsx, sheet_name=0)
    rename_map = {
        "metadata_table": "表",
        "metadata_variable": "变量",
        "annotation_variable": "变量名",
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    if "metadata_table" not in df.columns and "表" in df.columns:
        df["metadata_table"] = df["表"]
    df.to_excel(dst_xlsx, sheet_name="Sheet1", index=False)
    return dst_xlsx


# ---------------------------------------------------------------------------
# Test: 0 - Module imports
# ---------------------------------------------------------------------------


def test_imports() -> TestResult:
    r = TestResult("模块导入检查")
    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                "-c",
                """
import sys; sys.path.insert(0, '.')
from src.processors.sdtm_processor import SDTMProcessor
from src.processors.postprocess import PostprocessMixin
from src.knowledge_base.llm_query_interface import LLMKnowledgeQueryInterface
from src.spec_mapper import SpecMapper
from src.web.session_manager import session_manager
from src.web.job_manager import job_manager
print("ALL_IMPORTS_OK")
""",
            ]
        )
        dt = time.time() - t0
        if "ALL_IMPORTS_OK" in proc.stdout:
            r.passed("所有核心模块导入成功", dt)
        else:
            r.failed(f"导入异常:\nstdout: {proc.stdout[:500]}\nstderr: {proc.stderr[:500]}", dt)
    except Exception as e:
        r.failed(str(e), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 1 - extract_ecrf_sheet  (eCRF Excel → JSON)
# ---------------------------------------------------------------------------


def test_extract_ecrf(data_dir: Path, work_dir: Path, ctx: PipelineContext) -> TestResult:
    r = TestResult("Step1: extract_ecrf_sheet (eCRF Excel → JSON)")
    ecrf_file = find_file(data_dir, exact_name="ecrf_raw.xlsx")
    if not ecrf_file:
        r.skipped("未找到 ecrf_raw.xlsx")
        return r

    output_dir = work_dir / "step1_output"
    output_dir.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                str(SCRIPTS_DIR / "extract_ecrf_sheet.py"),
                "-i",
                str(ecrf_file),
                "-o",
                str(output_dir),
            ]
        )
        dt = time.time() - t0

        json_files = list(output_dir.glob("*.json"))
        if proc.returncode == 0 and json_files:
            data = json.loads(json_files[0].read_text(encoding="utf-8"))
            count = len(data) if isinstance(data, list) else 0
            ctx.step1_json = json_files[0]
            r.passed(f"生成 {len(json_files)} 个 JSON, 共 {count} 条映射 → 自动串联到 Step3", dt, json_files)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"exit_code={proc.returncode}\n{err_msg}", dt)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 2 - convert_als2sdtm  (ALS Excel → KB)
# ---------------------------------------------------------------------------


def test_convert_als2sdtm(data_dir: Path, work_dir: Path) -> TestResult:
    r = TestResult("Step2: convert_als2sdtm (ALS Excel → KB)")
    als_file = find_file(data_dir, exact_name="als2sdtm_example.xlsx")
    if not als_file:
        r.skipped("未找到 als2sdtm_example.xlsx (可选步骤)")
        return r

    sheet_name = detect_als_sheet(als_file)
    if not sheet_name:
        r.skipped("als2sdtm_example.xlsx 中未找到含 SDTM_Domain/SDTM_Variable 列的 sheet，可能不是 ALS 映射格式")
        return r

    output_dir = work_dir / "step2_output"
    output_dir.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                str(SCRIPTS_DIR / "convert_als2sdtm.py"),
                "--input",
                str(als_file),
                "--output-dir",
                str(output_dir),
                "--sheet",
                sheet_name,
                "--format",
                "both",
            ]
        )
        dt = time.time() - t0

        outputs = list(output_dir.glob("*"))
        parquets = [f for f in outputs if f.suffix == ".parquet"]
        jsons = [f for f in outputs if f.suffix == ".json"]

        if proc.returncode == 0 and (parquets or jsons):
            r.passed(f"sheet='{sheet_name}', 生成 {len(parquets)} parquet + {len(jsons)} JSON", dt, outputs)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"exit_code={proc.returncode}, sheet='{sheet_name}'\n{err_msg}", dt)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 3 - generate_sdtm_recommendations  --dry-run (不调 LLM)
# ---------------------------------------------------------------------------


def test_recommendations_dryrun(data_dir: Path, work_dir: Path, ctx: PipelineContext) -> TestResult:
    r = TestResult("Step3: generate_sdtm_recommendations --dry-run (不调 LLM)")

    json_file = find_file(data_dir, exact_name="processed.json", suffix=".json")
    if not json_file and ctx.step1_json:
        json_file = ctx.step1_json
        r.name += " [串联自 Step1]"

    if not json_file:
        r.skipped("无 processed.json 且 Step1 未生成输出")
        return r

    output_base = work_dir / "step3_output" / "reco"
    output_base.parent.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                str(SCRIPTS_DIR / "generate_sdtm_recommendations.py"),
                "--json-file",
                str(json_file),
                "--output",
                str(output_base),
                "--dry-run",
                "--language",
                "cn",
            ]
        )
        dt = time.time() - t0

        if proc.returncode == 0:
            r.passed(f"dry-run 通过 (输入: {json_file.name})", dt)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"exit_code={proc.returncode}\n{err_msg}", dt)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 3b - generate_sdtm_recommendations  (live LLM, optional)
# ---------------------------------------------------------------------------


def test_recommendations_live(data_dir: Path, work_dir: Path, ctx: PipelineContext) -> TestResult:
    r = TestResult("Step3b: generate_sdtm_recommendations (调 LLM)")

    json_file = find_file(data_dir, exact_name="processed.json", suffix=".json")
    if not json_file and ctx.step1_json:
        json_file = ctx.step1_json
        r.name += " [串联自 Step1]"

    if not json_file:
        r.skipped("无 processed.json 且 Step1 未生成输出")
        return r

    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        r.skipped("OPENROUTER_API_KEY 未设置, 跳过在线测试")
        return r

    output_base = work_dir / "step3b_output" / "reco"
    output_base.parent.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                str(SCRIPTS_DIR / "generate_sdtm_recommendations.py"),
                "--json-file",
                str(json_file),
                "--output",
                str(output_base),
                "--language",
                "cn",
            ],
            timeout=600,
        )
        dt = time.time() - t0

        outputs = list(output_base.parent.glob("*"))
        xlsx = [f for f in outputs if f.suffix == ".xlsx"]
        if proc.returncode == 0 and xlsx:
            try:
                converted = output_base.parent / "reco_spec_mapper_compat.xlsx"
                convert_llm_output_to_spec_mapper_format(xlsx[0], converted)
                ctx.step3_als_output = converted
                r.passed(
                    f"LLM 推荐完成, 生成 {len(xlsx)} 个 Excel (已转换列名) → 自动串联到 Step4", dt, [*xlsx, converted]
                )
            except Exception as exc:
                ctx.step3_als_output = xlsx[0]
                r.passed(f"LLM 推荐完成, 生成 {len(xlsx)} 个 Excel (列名转换失败: {exc}) → 串联到 Step4", dt, xlsx)
        elif proc.returncode == 0:
            r.passed("LLM 推荐完成 (无 Excel 输出)", dt)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"exit_code={proc.returncode}\n{err_msg}", dt)
    except subprocess.TimeoutExpired:
        r.failed("超时 (>600s)", time.time() - t0)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 4 - generate_full_spec  (ALS → SDTM Spec)
# ---------------------------------------------------------------------------


def test_generate_spec(data_dir: Path, work_dir: Path, ctx: PipelineContext) -> TestResult:
    r = TestResult("Step4: generate_full_spec (ALS → SDTM Spec)")

    als_file = find_file(data_dir, exact_name="als_output.xlsx")
    if not als_file and ctx.step3_als_output:
        als_file = ctx.step3_als_output
        r.name += " [串联自 Step3]"

    template_file = find_file(data_dir, exact_name="template_spec.xlsx", patterns=["template", "spec"])

    if not als_file:
        r.skipped("无 als_output.xlsx 且 Step3 未生成 Excel 输出")
        return r
    if not template_file:
        r.skipped("未找到模板文件 (文件名需含 'template' 或 'spec')")
        return r

    output_file = work_dir / "step4_output" / "sdtm_spec_test.xlsx"
    output_file.parent.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                str(SCRIPTS_DIR / "generate_full_spec.py"),
                "--als-file",
                str(als_file),
                "--template-file",
                str(template_file),
                "--output",
                str(output_file),
            ]
        )
        dt = time.time() - t0

        if proc.returncode == 0 and output_file.exists():
            size_kb = output_file.stat().st_size / 1024
            r.passed(
                f"Spec 已生成 ({size_kb:.1f} KB), ALS={als_file.name}, 模板={template_file.name}", dt, [output_file]
            )
        elif proc.returncode == 0:
            r.passed("脚本执行成功", dt)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"exit_code={proc.returncode}\n{err_msg}", dt)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Test: 5 - Web API Smoke
# ---------------------------------------------------------------------------


def test_web_api(data_dir: Path, work_dir: Path) -> TestResult:
    r = TestResult("Step5: Web API 冒烟测试 (启动 → API → 上传)")

    try:
        import requests
    except ImportError:
        r.skipped("requests 库未安装")
        return r

    port = 18921
    base_url = f"http://127.0.0.1:{port}"

    t0 = time.time()
    server_proc = None
    try:
        server_proc = subprocess.Popen(
            [PYTHON, "-m", "uvicorn", "app:app", "--port", str(port), "--host", "127.0.0.1"],
            cwd=PROJECT_ROOT,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )

        ready = False
        for _ in range(30):
            time.sleep(1)
            try:
                resp = requests.get(base_url, timeout=3)
                if resp.status_code in (200, 304):
                    ready = True
                    break
            except requests.ConnectionError:
                continue
            except Exception:
                continue

        if not ready:
            stderr_out = ""
            try:
                server_proc.terminate()
                _, stderr_bytes = server_proc.communicate(timeout=5)
                stderr_out = (stderr_bytes.decode("utf-8", errors="replace")[:800]) if stderr_bytes else ""
            except Exception:
                pass
            r.failed(f"服务器 30s 内未就绪\n{stderr_out}", time.time() - t0)
            return r

        checks = []

        resp = requests.get(base_url, timeout=5)
        checks.append(("GET /", resp.status_code))

        resp = requests.get(f"{base_url}/api/session-stats", timeout=5)
        checks.append(("GET /api/session-stats", resp.status_code))

        resp = requests.get(f"{base_url}/api/processed-files", timeout=5)
        checks.append(("GET /api/processed-files", resp.status_code))

        ecrf_file = find_file(data_dir, exact_name="ecrf_raw.xlsx")
        if ecrf_file:
            with open(ecrf_file, "rb") as f:
                resp = requests.post(
                    f"{base_url}/api/upload/raw",
                    files={
                        "file": (
                            "ecrf_raw.xlsx",
                            f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                    headers={"X-Session-ID": "smoke-test-session"},
                    timeout=60,
                )
            checks.append(("POST /api/upload/raw", resp.status_code))

        dt = time.time() - t0

        all_ok = all(200 <= code < 400 for _, code in checks)
        check_str = ", ".join(f"{name}={code}" for name, code in checks)

        if all_ok:
            r.passed(f"API 全部通过: {check_str}", dt)
        else:
            r.failed(f"部分 API 失败: {check_str}", dt)

    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    finally:
        if server_proc and server_proc.poll() is None:
            server_proc.terminate()
            try:
                server_proc.wait(timeout=10)
            except subprocess.TimeoutExpired:
                server_proc.kill()

    return r


# ---------------------------------------------------------------------------
# Test: 6 - Spec Mapper Python API
# ---------------------------------------------------------------------------


def test_spec_mapper_api(data_dir: Path, work_dir: Path, ctx: PipelineContext) -> TestResult:
    r = TestResult("Step6: SpecMapper Python API 直接调用")

    als_file = find_file(data_dir, exact_name="als_output.xlsx")
    if not als_file and ctx.step3_als_output:
        als_file = ctx.step3_als_output

    template_file = find_file(data_dir, exact_name="template_spec.xlsx", patterns=["template", "spec"])

    if not als_file:
        r.skipped("无 als_output.xlsx 且 Step3 未生成 Excel")
        return r
    if not template_file:
        r.skipped("未找到模板文件")
        return r

    output_file = work_dir / "step6_output" / "spec_mapper_result.xlsx"
    output_file.parent.mkdir(exist_ok=True)

    t0 = time.time()
    try:
        proc = run_cmd(
            [
                PYTHON,
                "-c",
                f"""
import sys; sys.path.insert(0, '.')
from src.spec_mapper import SpecMapper
mapper = SpecMapper(
    als_file=r"{als_file}",
    template_file=r"{template_file}",
    als_sheet="Sheet1",
)
result = mapper.process(output_file=r"{output_file}", dry_run=True)
print("RESULT:", result)
print("SPEC_MAPPER_OK")
""",
            ]
        )
        dt = time.time() - t0
        if "SPEC_MAPPER_OK" in proc.stdout:
            r.passed(f"dry-run 成功 (ALS={als_file.name}, 模板={template_file.name})", dt)
        else:
            err_msg = proc.stderr[:800] if proc.stderr else proc.stdout[:800]
            r.failed(f"调用失败\n{err_msg}", dt)
    except Exception:
        r.failed(traceback.format_exc(), time.time() - t0)
    return r


# ---------------------------------------------------------------------------
# Main Runner
# ---------------------------------------------------------------------------


def generate_fixtures(out_dir: Path):
    """Use the built-in data generator."""
    gen_script = PROJECT_ROOT / "tests" / "generate_test_data.py"
    proc = run_cmd([PYTHON, str(gen_script), "--output-dir", str(out_dir)])
    if proc.returncode != 0:
        print(f"[ERROR] 测试数据生成失败:\n{proc.stderr}")
        sys.exit(1)
    print(f"{INFO} 已生成测试数据 -> {out_dir}")


def main():
    parser = argparse.ArgumentParser(
        description="iSDTaiM Smoke Test - 快速验证流程是否跑通",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python tests/smoke_test.py                          # 自动生成数据, 跑全部测试
  python tests/smoke_test.py --data-dir my_test/      # 用自己的数据
  python tests/smoke_test.py --offline                 # 只跑离线测试 (不调 LLM)
  python tests/smoke_test.py --web-only                # 只跑 Web API 测试
  python tests/smoke_test.py --keep                    # 保留中间产物

数据目录只需放:
  ecrf_raw.xlsx          eCRF 原始 Excel (必需)
  *template*.xlsx        SDTM Spec 模板 (Step4 需要)
  als2sdtm_example.xlsx  历史映射样例 (可选, 知识库)
""",
    )
    parser.add_argument(
        "--data-dir",
        default=None,
        help="测试数据目录, 留空则自动生成合成数据",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="只跑离线步骤 (不调用 LLM, 不启动 Web 服务)",
    )
    parser.add_argument(
        "--web-only",
        action="store_true",
        help="只跑 Web API 冒烟测试",
    )
    parser.add_argument(
        "--keep",
        action="store_true",
        help="测试结束后保留临时工作目录",
    )
    parser.add_argument(
        "--work-dir",
        default=None,
        help="指定工作目录 (存放测试产物), 默认使用临时目录",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  iSDTaiM Smoke Test Runner")
    print("=" * 60)
    print()

    # --- Resolve data directory ---
    if args.data_dir:
        data_dir = Path(args.data_dir).resolve()
        if not data_dir.exists():
            print(f"[ERROR] 指定的数据目录不存在: {data_dir}")
            sys.exit(1)
        print(f"{INFO} 使用用户数据: {data_dir}")
    else:
        data_dir = PROJECT_ROOT / "tests" / "fixtures"
        generate_fixtures(data_dir)

    # --- Resolve work directory ---
    if args.work_dir:
        work_dir = Path(args.work_dir).resolve()
        work_dir.mkdir(parents=True, exist_ok=True)
        tmp_dir = None
    else:
        tmp_dir = tempfile.mkdtemp(prefix="isdtaim_smoke_")
        work_dir = Path(tmp_dir)

    print(f"{INFO} 工作目录: {work_dir}")
    print(f"{INFO} 模式: {'仅离线' if args.offline else '仅Web' if args.web_only else '全部'}")
    print()

    # --- File discovery ---
    print(f"{INFO} 数据目录文件:")
    for f in sorted(data_dir.iterdir()):
        if f.is_file():
            size = f.stat().st_size / 1024
            print(f"       {f.name:35s} ({size:.1f} KB)")
    print()

    template_file = find_file(data_dir, exact_name="template_spec.xlsx", patterns=["template", "spec"])
    als_example = find_file(data_dir, exact_name="als2sdtm_example.xlsx")

    print(f"{INFO} 文件识别结果:")
    print(f"       eCRF 源文件:   {'ecrf_raw.xlsx' if (data_dir / 'ecrf_raw.xlsx').exists() else '(未找到)'}")
    print(f"       模板文件:      {template_file.name if template_file else '(未找到)'}")
    print(f"       ALS 映射样例:  {als_example.name if als_example else '(未找到, 可选)'}")
    print()

    # --- Pipeline context for chaining ---
    ctx = PipelineContext()

    # --- Run tests ---
    results: list[TestResult] = []

    if args.web_only:
        test_suite = [
            lambda: test_imports(),
            lambda: test_web_api(data_dir, work_dir),
        ]
    elif args.offline:
        test_suite = [
            lambda: test_imports(),
            lambda: test_extract_ecrf(data_dir, work_dir, ctx),
            lambda: test_convert_als2sdtm(data_dir, work_dir),
            lambda: test_recommendations_dryrun(data_dir, work_dir, ctx),
            lambda: test_spec_mapper_api(data_dir, work_dir, ctx),
            lambda: test_generate_spec(data_dir, work_dir, ctx),
        ]
    else:
        test_suite = [
            lambda: test_imports(),
            lambda: test_extract_ecrf(data_dir, work_dir, ctx),
            lambda: test_convert_als2sdtm(data_dir, work_dir),
            lambda: test_recommendations_dryrun(data_dir, work_dir, ctx),
            lambda: test_recommendations_live(data_dir, work_dir, ctx),
            lambda: test_spec_mapper_api(data_dir, work_dir, ctx),
            lambda: test_generate_spec(data_dir, work_dir, ctx),
            lambda: test_web_api(data_dir, work_dir),
        ]

    print("-" * 60)
    for test_fn in test_suite:
        result = test_fn()
        results.append(result)
        print_result(result)
    print("-" * 60)

    # --- Summary ---
    passed = sum(1 for r in results if r.status == "pass")
    failed = sum(1 for r in results if r.status == "fail")
    skipped = sum(1 for r in results if r.status == "skip")
    total_time = sum(r.duration for r in results)

    print()
    print(
        f"  总计: {len(results)} 个测试 | 通过: {passed} | 失败: {failed} | 跳过: {skipped} | 耗时: {total_time:.1f}s"
    )

    if failed > 0:
        print()
        print("  失败的测试:")
        for r in results:
            if r.status == "fail":
                print(f"    - {r.name}")
        print()

    # --- Cleanup ---
    if not args.keep and tmp_dir:
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
            print(f"{INFO} 已清理临时目录")
        except Exception:
            pass
    elif args.keep:
        print(f"{INFO} 保留工作目录: {work_dir}")

    print()
    if failed == 0:
        print("  ALL SMOKE TESTS PASSED")
    else:
        print(f"  {failed} TEST(S) FAILED")

    sys.exit(1 if failed > 0 else 0)


if __name__ == "__main__":
    main()
