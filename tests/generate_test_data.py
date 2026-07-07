"""
Generate minimal synthetic test data for SIRIUS smoke tests.

Usage:
    python tests/generate_test_data.py [--output-dir tests/fixtures]

Creates:
    tests/fixtures/
    ├── ecrf_raw.xlsx        # Simulated eCRF raw data (for extract_ecrf_sheet)
    ├── als2sdtm_example.xlsx # Simulated ALS2SDTM example (for convert_als2sdtm)
    ├── processed.json       # Simulated processed JSON (for generate_sdtm_recommendations)
    ├── als_output.xlsx      # Simulated ALS output (for spec mapper)
    └── template_spec.xlsx   # Minimal SDTM Spec template (for generate_full_spec)
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1. eCRF Raw Data  (simulates what extract_ecrf_sheet.py expects)
# ---------------------------------------------------------------------------

"""
extract_ecrf_sheet.py expects 3 sheets:
  - "eCRF界面": cols 类型/隐藏/表/变量  (filter 类型=变量, 隐藏=否)
  - "内嵌设计": cols 表名/表/内嵌表名/内嵌表
  - "eCRF":     cols 编号/表名/表/变量名/变量
"""

ECRF_SHEET_ECRF_UI = {
    "headers": ["类型", "隐藏", "表", "变量"],
    "rows": [
        ["变量", "否", "人口学", "SUBJID"],
        ["变量", "否", "人口学", "BRTHDTC"],
        ["变量", "否", "人口学", "SEX"],
        ["变量", "否", "生命体征", "SYSBP"],
        ["变量", "否", "生命体征", "DIABP"],
        ["变量", "否", "不良事件", "AETERM"],
        ["变量", "否", "不良事件", "AESTDTC"],
        ["标题", "否", "人口学", "HDR1"],
    ],
}

ECRF_SHEET_EMBEDDED = {
    "headers": ["表名", "表", "内嵌表名", "内嵌表"],
    "rows": [],
}

ECRF_SHEET_ECRF = {
    "headers": ["编号", "表名", "表", "变量名", "变量"],
    "rows": [
        [1, "人口学", "DM", "受试者编号", "SUBJID"],
        [2, "人口学", "DM", "出生日期", "BRTHDTC"],
        [3, "人口学", "DM", "性别", "SEX"],
        [4, "生命体征", "VS", "收缩压", "SYSBP"],
        [5, "生命体征", "VS", "舒张压", "DIABP"],
        [6, "不良事件", "AE", "不良事件名称", "AETERM"],
        [7, "不良事件", "AE", "开始日期", "AESTDTC"],
    ],
}


def create_ecrf_raw(out_dir: Path):
    """Create a minimal eCRF Raw Excel file with the 3 expected sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("eCRF界面")
    ws1.append(ECRF_SHEET_ECRF_UI["headers"])
    for row in ECRF_SHEET_ECRF_UI["rows"]:
        ws1.append(row)

    ws2 = wb.create_sheet("内嵌设计")
    ws2.append(ECRF_SHEET_EMBEDDED["headers"])
    for row in ECRF_SHEET_EMBEDDED["rows"]:
        ws2.append(row)

    ws3 = wb.create_sheet("eCRF")
    ws3.append(ECRF_SHEET_ECRF["headers"])
    for row in ECRF_SHEET_ECRF["rows"]:
        ws3.append(row)

    path = out_dir / "ecrf_raw.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 2. ALS Example  (simulates an existing ALS2SDTM mapping example)
# ---------------------------------------------------------------------------

ALS_EXAMPLE_HEADERS = [
    "表名",
    "表",
    "变量名",
    "变量",
    "SDTM_Domain",
    "SDTM_Variable",
]

ALS_EXAMPLE_ROWS = [
    ["人口学", "DM", "受试者编号", "SUBJID", "DM", "SUBJID"],
    ["人口学", "DM", "出生日期", "BRTHDTC", "DM", "BRTHDTC"],
    ["人口学", "DM", "性别", "SEX", "DM", "SEX"],
    ["生命体征", "VS", "收缩压", "SYSBP", "VS", "VSORRES"],
    ["生命体征", "VS", "舒张压", "DIABP", "VS", "VSORRES"],
    ["不良事件", "AE", "不良事件名称", "AETERM", "AE", "AETERM"],
    ["不良事件", "AE", "开始日期", "AESTDTC", "AE", "AESTDTC"],
]


def create_als_example(out_dir: Path):
    """Create a minimal ALS2SDTM example Excel (sheet='eCRF' for convert_als2sdtm default)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "eCRF"
    ws.append(ALS_EXAMPLE_HEADERS)
    for row in ALS_EXAMPLE_ROWS:
        ws.append(row)

    path = out_dir / "als2sdtm_example.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 3. Processed JSON  (the format consumed by generate_sdtm_recommendations)
# ---------------------------------------------------------------------------

PROCESSED_MAPPINGS = [
    {
        "annotation_table": "人口学",
        "metadata_table": "DM",
        "annotation_variable": "受试者编号",
        "metadata_variable": "SUBJID",
        "SDTM_Domain": None,
        "SDTM_Variable": None,
    },
    {
        "annotation_table": "人口学",
        "metadata_table": "DM",
        "annotation_variable": "性别",
        "metadata_variable": "SEX",
        "SDTM_Domain": None,
        "SDTM_Variable": None,
    },
    {
        "annotation_table": "生命体征",
        "metadata_table": "VS",
        "annotation_variable": "收缩压",
        "metadata_variable": "SYSBP",
        "SDTM_Domain": None,
        "SDTM_Variable": None,
    },
    {
        "annotation_table": "不良事件",
        "metadata_table": "AE",
        "annotation_variable": "不良事件名称",
        "metadata_variable": "AETERM",
        "SDTM_Domain": None,
        "SDTM_Variable": None,
    },
]


def create_processed_json(out_dir: Path):
    """Create a minimal processed JSON for AI recommendation."""
    path = out_dir / "processed.json"
    path.write_text(json.dumps(PROCESSED_MAPPINGS, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# 4. ALS Output  (simulates the AI-generated ALS output for spec mapper)
# ---------------------------------------------------------------------------

ALS_OUTPUT_HEADERS = [
    "表",
    "变量",
    "变量名",
    "metadata_table",
    "metadata_variable",
    "SDTM_Domain",
    "SDTM_Variable",
    "组件类型",
    "编码名称",
]

ALS_OUTPUT_ROWS = [
    ["人口学", "受试者编号", "Subject ID", "DM", "SUBJID", "DM", "SUBJID", "文本框", ""],
    ["人口学", "出生日期", "Birth Date", "DM", "BRTHDTC", "DM", "BRTHDTC", "日期", ""],
    ["人口学", "性别", "Sex", "DM", "SEX", "DM", "SEX", "单选", "SEX"],
    ["生命体征", "收缩压", "Systolic BP", "VS", "SYSBP", "VS", "VSORRES", "数值", ""],
    ["生命体征", "舒张压", "Diastolic BP", "VS", "DIABP", "VS", "VSORRES", "数值", ""],
    ["生命体征", "心率", "Heart Rate", "VS", "HR", "VS", "VSORRES", "数值", ""],
    ["不良事件", "不良事件名称", "AE Term", "AE", "AETERM", "AE", "AETERM", "文本框", ""],
    ["不良事件", "开始日期", "Start Date", "AE", "AESTDTC", "AE", "AESTDTC", "日期", ""],
]


def create_als_output(out_dir: Path):
    """Create a minimal ALS output Excel for spec mapper."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(ALS_OUTPUT_HEADERS)
    for row in ALS_OUTPUT_ROWS:
        ws.append(row)

    path = out_dir / "als_output.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 5. Template Spec  (minimal SDTM Spec template)
# ---------------------------------------------------------------------------

TEMPLATE_VARIABLE_HEADERS_ROW = 13  # 1-indexed row for variable headers (config: header_row=12, 0-based)

TEMPLATE_SHEETS = {
    "CONTENT": {"type": "info"},
    "DM": {
        "type": "domain",
        "headers": [
            "序号",
            "数据集",
            "变量名称",
            "变量标签",
            "类型",
            "长度",
            "受控术语/格式",
            "来源",
            "核心",
            "转换定义",
            "备注",
        ],
        "rows": [
            [1, "DM", "STUDYID", "Study Identifier", "Char", 20, "", "指定", "必需", "", ""],
            [2, "DM", "DOMAIN", "Domain Abbreviation", "Char", 2, "DM", "指定", "必需", "", ""],
            [3, "DM", "USUBJID", "Unique Subject ID", "Char", 40, "", "衍生", "必需", "", ""],
            [4, "DM", "SUBJID", "Subject ID", "Char", 20, "", "", "必需", "", ""],
            [5, "DM", "SEX", "Sex", "Char", 10, "SEX", "", "必需", "", ""],
            [6, "DM", "BRTHDTC", "Date of Birth", "Char", 19, "ISO 8601", "", "期望", "", ""],
        ],
    },
    "VS": {
        "type": "domain",
        "headers": [
            "序号",
            "数据集",
            "变量名称",
            "变量标签",
            "类型",
            "长度",
            "受控术语/格式",
            "来源",
            "核心",
            "转换定义",
            "备注",
        ],
        "rows": [
            [1, "VS", "STUDYID", "Study Identifier", "Char", 20, "", "指定", "必需", "", ""],
            [2, "VS", "DOMAIN", "Domain Abbreviation", "Char", 2, "VS", "指定", "必需", "", ""],
            [3, "VS", "USUBJID", "Unique Subject ID", "Char", 40, "", "指定", "必需", "", ""],
            [4, "VS", "VSSEQ", "Sequence Number", "Num", 8, "", "衍生", "必需", "", ""],
            [5, "VS", "VSTESTCD", "Vital Signs Test Code", "Char", 8, "VSTESTCD", "指定", "必需", "", ""],
            [6, "VS", "VSTEST", "Vital Signs Test Name", "Char", 40, "VSTEST", "指定", "必需", "", ""],
            [7, "VS", "VSORRES", "Result/Finding", "Char", 200, "", "", "期望", "", ""],
        ],
    },
    "AE": {
        "type": "domain",
        "headers": [
            "序号",
            "数据集",
            "变量名称",
            "变量标签",
            "类型",
            "长度",
            "受控术语/格式",
            "来源",
            "核心",
            "转换定义",
            "备注",
        ],
        "rows": [
            [1, "AE", "STUDYID", "Study Identifier", "Char", 20, "", "指定", "必需", "", ""],
            [2, "AE", "DOMAIN", "Domain Abbreviation", "Char", 2, "AE", "指定", "必需", "", ""],
            [3, "AE", "USUBJID", "Unique Subject ID", "Char", 40, "", "指定", "必需", "", ""],
            [4, "AE", "AESEQ", "Sequence Number", "Num", 8, "", "衍生", "必需", "", ""],
            [5, "AE", "AETERM", "Reported Term", "Char", 200, "", "", "必需", "", ""],
            [6, "AE", "AESTDTC", "Start Date/Time", "Char", 19, "ISO 8601", "", "期望", "", ""],
        ],
    },
}


def create_template_spec(out_dir: Path):
    """Create a minimal SDTM Spec template Excel."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, config in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        if config["type"] == "info":
            ws["A1"] = "SDTM Spec Template (Smoke Test)"
            continue

        for _ in range(TEMPLATE_VARIABLE_HEADERS_ROW - 1):
            ws.append([])
        ws.append(config["headers"])
        for row in config["rows"]:
            ws.append(row)

    path = out_dir / "template_spec.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Generate SIRIUS smoke-test fixtures")
    parser.add_argument(
        "--output-dir",
        default="tests/fixtures",
        help="Directory to write fixture files (default: tests/fixtures)",
    )
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    files = [
        ("eCRF Raw Excel", create_ecrf_raw(out_dir)),
        ("ALS2SDTM Example Excel", create_als_example(out_dir)),
        ("Processed JSON", create_processed_json(out_dir)),
        ("ALS Output Excel", create_als_output(out_dir)),
        ("Template Spec Excel", create_template_spec(out_dir)),
    ]

    print(f"[OK] Test fixtures generated in {out_dir}/")
    for label, path in files:
        print(f"     {label:25s} -> {path}")


if __name__ == "__main__":
    main()
