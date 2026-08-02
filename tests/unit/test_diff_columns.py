"""Unit tests for AI-vs-reference diff column logic."""

from __future__ import annotations

from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook

from src.processors.excel_styler import (
    DIFF_AI_ONLY,
    DIFF_DOMAIN_DIFF,
    DIFF_MATCH,
    DIFF_REF_ONLY,
    DIFF_VAR_DIFF,
    apply_diff_fill,
)
from src.processors.sdtm_processor import (
    SDTMProcessor,
    attach_reference_diff,
    compute_diff_status,
)


def _assert_fill_color(cell, expected_hex: str) -> None:
    assert cell.fill.start_color.rgb[-6:] == expected_hex


class TestComputeDiffStatus:
    @pytest.mark.parametrize(
        ("ai_domain", "ai_variable", "ref_domain", "ref_variable", "expected"),
        [
            ("DM", "BRTHDTC", "DM", "BRTHDTC", "match"),
            ("DM", "BRTHDTC", "DM", "BRTHYR", "var_diff"),
            ("DM", "BRTHDTC", "AE", "AETERM", "domain_diff"),
            ("DM", "BRTHDTC", None, None, "ai_only"),
            (None, None, "DM", "BRTHDTC", "ref_only"),
        ],
    )
    def test_compute_diff_status_matrix(
        self,
        ai_domain: str | None,
        ai_variable: str | None,
        ref_domain: str | None,
        ref_variable: str | None,
        expected: str,
    ) -> None:
        assert compute_diff_status(ai_domain, ai_variable, ref_domain, ref_variable) == expected


class TestApplyDiffFill:
    def test_apply_diff_fill_colors_each_status(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Diff_Status"])
        for status in ("match", "var_diff", "domain_diff", "ai_only", "ref_only"):
            ws.append([status])

        count = apply_diff_fill(ws, "A", header_row=1)
        assert count == 5

        _assert_fill_color(ws["A2"], DIFF_MATCH)
        _assert_fill_color(ws["A3"], DIFF_VAR_DIFF)
        _assert_fill_color(ws["A4"], DIFF_DOMAIN_DIFF)
        _assert_fill_color(ws["A5"], DIFF_AI_ONLY)
        _assert_fill_color(ws["A6"], DIFF_REF_ONLY)


class TestAttachReferenceDiff:
    def test_diff_columns_absent_when_no_reference_kb(self) -> None:
        rows = [
            {
                "annotation_table": "DM",
                "annotation_variable": "BRTHDTC",
                "SDTM_Domain": "DM",
                "SDTM_Variable": "BRTHDTC",
            }
        ]
        result = attach_reference_diff(rows, None)
        assert result == rows
        assert "Reference_Domain" not in result[0]
        assert "Reference_Variable" not in result[0]
        assert "Diff_Status" not in result[0]

    def test_diff_columns_present_when_reference_kb_has_overlap(self) -> None:
        rows = [
            {
                "annotation_table": "DM",
                "annotation_variable": "BRTHDTC",
                "SDTM_Domain": "DM",
                "SDTM_Variable": "BRTHDTC",
            },
            {"annotation_table": "DM", "annotation_variable": "SEX", "SDTM_Domain": "DM", "SDTM_Variable": "SEX"},
            {"annotation_table": "AE", "annotation_variable": "AETERM", "SDTM_Domain": "AE", "SDTM_Variable": "AETERM"},
        ]
        ref_kb = pd.DataFrame(
            {
                "annotation_table": ["DM", "DM"],
                "annotation_variable": ["BRTHDTC", "SEX"],
                "SDTM_Domain": ["DM", "DM"],
                "SDTM_Variable": ["BRTHYR", "SEX"],
            }
        )

        result = attach_reference_diff(rows, ref_kb)
        assert len(result) == 3

        # Row 0: var_diff (DM/BRTHDTC vs DM/BRTHYR)
        assert result[0]["Reference_Domain"] == "DM"
        assert result[0]["Reference_Variable"] == "BRTHYR"
        assert result[0]["Diff_Status"] == "var_diff"

        # Row 1: match
        assert result[1]["Reference_Domain"] == "DM"
        assert result[1]["Reference_Variable"] == "SEX"
        assert result[1]["Diff_Status"] == "match"

        # Row 2: ai_only (no reference match)
        assert result[2]["Reference_Domain"] == ""
        assert result[2]["Reference_Variable"] == ""
        assert result[2]["Diff_Status"] == "ai_only"


def test_save_uses_frozen_reference_kb_instead_of_live_session_state(tmp_path) -> None:
    snapshot = tmp_path / "project_frozen.parquet"
    snapshot.write_bytes(b"snapshot-marker")
    frozen_reference = pd.DataFrame(
        {
            "annotation_table": ["Demographics"],
            "annotation_variable": ["Birth date"],
            "SDTM_Domain": ["DM"],
            "SDTM_Variable": ["BRTHYR"],
        }
    )
    processor = object.__new__(SDTMProcessor)
    processor.client = SimpleNamespace(
        get_sanitized_model_name=lambda _model: "test_model",
    )
    processor.model_name = "test/model"
    processor.session_id = "session-with-newer-live-kb"
    processor.reference_kb_files = [str(snapshot)]
    processor._input_file = None
    processor._merge_with_ecrf_sheet = lambda _df: None
    recommendations = [
        {
            "table_name": "DM",
            "original_mappings": [
                {
                    "annotation_table": "Demographics",
                    "annotation_variable": "Birth date",
                    "metadata_variable": "BRTHDTC",
                }
            ],
            "domain_recommendations": [
                {
                    "domain": "DM",
                    "sdtm_variable": "BRTHDTC",
                    "variable_name": "BRTHDTC",
                    "score": 0.9,
                    "source": "LLM_reasoning",
                }
            ],
        }
    ]

    with (
        patch(
            "src.processors.project_ingest.load_reference_kb_files",
            return_value=frozen_reference,
        ) as frozen_loader,
        patch(
            "src.processors.project_ingest.load_session_reference_kb",
            side_effect=AssertionError("live session KB must not be read"),
        ),
    ):
        rows = processor.save_recommendations(
            recommendations,
            str(tmp_path / "result"),
        )

    frozen_loader.assert_called_once_with([str(snapshot)])
    assert rows[0]["Reference_Domain"] == "DM"
    assert rows[0]["Reference_Variable"] == "BRTHYR"
    assert rows[0]["Diff_Status"] == "var_diff"
