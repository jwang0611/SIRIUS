"""Session-scoped ALS upload/list/delete isolation."""

from __future__ import annotations

import os
import shutil
import threading
import time
import uuid
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from src.infrastructure.audit_logger import AuditLogger
from src.web.job_manager import JobManager
from src.web.session_manager import (
    SessionClosingError,
    SessionInfo,
    SessionManager,
    cleanup_orphaned_session_dirs,
    session_manager,
)


def _workbook_bytes(marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = marker
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def test_same_name_als_files_do_not_cross_sessions(tmp_path: Path, monkeypatch) -> None:
    from app import app

    monkeypatch.chdir(tmp_path)
    client = TestClient(app)
    session_a = "als-session-a"
    session_b = "als-session-b"
    headers_a = {"X-Session-ID": session_a}
    headers_b = {"X-Session-ID": session_b}
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    try:
        response_a = client.post(
            "/api/upload/als_output",
            headers=headers_a,
            files={"file": ("same.xlsx", _workbook_bytes("A"), mime)},
        )
        response_b = client.post(
            "/api/upload/als_output",
            headers=headers_b,
            files={"file": ("same.xlsx", _workbook_bytes("B"), mime)},
        )
        assert response_a.status_code == response_b.status_code == 200
        assert response_a.json()["stored_to"] == "same.xlsx"
        assert str(tmp_path) not in response_a.text

        path_a = session_manager.get_session_als_dir(session_a) / "same.xlsx"
        path_b = session_manager.get_session_als_dir(session_b) / "same.xlsx"
        assert path_a.resolve() != path_b.resolve()
        wb_a = load_workbook(path_a, read_only=True)
        wb_b = load_workbook(path_b, read_only=True)
        try:
            assert wb_a.active["A1"].value == "A"
            assert wb_b.active["A1"].value == "B"
        finally:
            wb_a.close()
            wb_b.close()

        assert [item["file_name"] for item in client.get("/api/als-files", headers=headers_a).json()["files"]] == [
            "same.xlsx"
        ]
        assert [item["file_name"] for item in client.get("/api/als-files", headers=headers_b).json()["files"]] == [
            "same.xlsx"
        ]

        assert (
            client.request(
                "DELETE",
                "/api/als-files",
                headers=headers_a,
                json={"file_id": "same.xlsx"},
            ).status_code
            == 200
        )
        assert client.delete("/api/als-files/same.xlsx", headers=headers_a).status_code == 404
        assert not path_a.exists()
        assert path_b.exists(), "deleting session A must not remove session B's same-name file"
    finally:
        session_manager.cleanup_session(session_a)
        session_manager.cleanup_session(session_b)


def test_web_session_cannot_overwrite_global_standards(tmp_path: Path, monkeypatch) -> None:
    from app import app

    monkeypatch.chdir(tmp_path)
    response = TestClient(app).post(
        "/api/upload/standards",
        headers={"X-Session-ID": "unprivileged-session"},
        files={
            "file": (
                "global-standard.xlsx",
                _workbook_bytes("must-not-write"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403
    assert response.json() == {"detail": "Web 端不允许更新全局标准库"}
    assert not (tmp_path / "data/knowledge_base/documents/standards/global-standard.xlsx").exists()


def test_cleanup_after_restart_removes_session_owned_output_dirs(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "not-present-in-memory"
    als_dir = manager.get_session_als_dir(session_id)
    spec_dir = manager.get_session_spec_job_dir(session_id, "job-1")
    (als_dir / "input.xlsx").write_bytes(b"input")
    (spec_dir / "output.xlsx").write_bytes(b"output")

    result = manager.cleanup_session(session_id)

    assert result["errors"] == []
    assert not als_dir.exists()
    assert not (Path("data/spec_output/sessions") / manager.session_dir_key(session_id)).exists()


def test_cleanup_retires_unknown_bearer_against_late_retry(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "body-not-yet-parsed"

    result = manager.cleanup_session(session_id)

    assert result["cleanup_pending"] is False
    with pytest.raises(SessionClosingError):
        with manager.operation(session_id):
            pass


def test_raw_and_processed_files_are_session_scoped(tmp_path: Path, monkeypatch) -> None:
    from app import app

    monkeypatch.chdir(tmp_path)
    client = TestClient(app)
    session_a = "raw-session-a"
    session_b = "raw-session-b"
    session_manager.get_or_create(session_b)
    headers_a = {"X-Session-ID": session_a}
    headers_b = {"X-Session-ID": session_b}
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def fake_extract(command: list[str]) -> str:
        output_dir = Path(command[command.index("--output-dir") + 1])
        input_path = Path(command[command.index("--input") + 1])
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / f"{input_path.stem}.json").write_text("[]", encoding="utf-8")
        (output_dir / f"{input_path.stem}.xlsx").write_bytes(b"derived")
        return "raw clinical stdout must not be returned"

    try:
        with patch("src.web.routers.upload.run_command", side_effect=fake_extract):
            response_a = client.post(
                "/api/upload/raw",
                headers=headers_a,
                files={"file": ("same.xlsx", _workbook_bytes("A"), mime)},
            )
        assert response_a.status_code == 200
        assert response_a.json()["script_output"] == "completed"
        assert "raw clinical stdout" not in response_a.text

        raw_a = session_manager.get_session_raw_dir(session_a) / "same.xlsx"
        processed_a = session_manager.get_session_processed_dir(session_a) / "same.json"
        assert raw_a.exists() and processed_a.exists()
        assert client.get("/api/processed-files", headers=headers_a).json()["files"] == ["same.json"]
        assert client.get("/api/processed-files", headers=headers_b).json()["files"] == []

        # Session B cannot launch a recommendation job against session A's
        # same-name processed input.
        with patch("src.web.routers.jobs.start_recommendations_job") as start_job:
            forbidden = client.post(
                "/api/recommendations",
                headers=headers_b,
                json={"json_file": "same.json"},
            )
        assert forbidden.status_code == 404
        start_job.assert_not_called()

        with patch("src.web.routers.upload.run_command", side_effect=fake_extract):
            response_b = client.post(
                "/api/upload/raw",
                headers=headers_b,
                files={"file": ("same.xlsx", _workbook_bytes("B"), mime)},
            )
        assert response_b.status_code == 200
        processed_b = session_manager.get_session_processed_dir(session_b) / "same.json"
        assert processed_a.resolve() != processed_b.resolve()
        assert processed_b.exists()

        with (
            patch("src.web.routers.jobs.start_recommendations_job") as start_job,
            patch("src.web.routers.jobs.job_manager") as manager,
        ):
            allowed = client.post(
                "/api/recommendations",
                headers=headers_b,
                json={"json_file": "same.json"},
            )
        assert allowed.status_code == 200
        start_job.assert_called_once()
        snapshot = Path(start_job.call_args.kwargs["json_file"])
        assert snapshot != processed_b.resolve()
        assert (
            snapshot.parent
            == (
                session_manager.get_session_recommendation_job_dir(session_b, manager.create_job.call_args.args[0])
                / "processed"
            ).resolve()
        )
        assert snapshot.read_bytes() == processed_b.read_bytes()
        assert manager.create_job.call_args.kwargs["owner_session_id"] == session_b
    finally:
        session_manager.cleanup_session(session_a)
        session_manager.cleanup_session(session_b)


def test_orphan_cleanup_scans_all_session_data_roots(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "orphan-session"
    roots = [
        manager.get_session_raw_dir(session_id),
        manager.get_session_processed_dir(session_id),
        manager.get_session_als_dir(session_id),
        manager.get_session_spec_job_dir(session_id, "job"),
    ]
    session_roots = [roots[0], roots[1], roots[2], roots[3].parent]
    old = time.time() - 3 * 3600
    for path in session_roots:
        for descendant in path.rglob("*"):
            os.utime(descendant, (old, old))
        os.utime(path, (old, old))

    result = cleanup_orphaned_session_dirs(max_age_hours=1)

    assert result["cleaned_dirs"] == 4
    assert all(not path.exists() for path in session_roots)


def test_session_cleanup_preserves_default_audit_log_directory(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "audit-cleanup-session"
    manager.get_or_create(session_id)
    auditor = AuditLogger(session_id=session_id)
    auditor.log_mapping(
        variable_data={"metadata_variable": "AETERM"},
        result={"domain": "AE", "sdtm_variable": "AETERM"},
    )
    audit_dir = tmp_path / "data/audit_logs/sessions" / manager.session_dir_key(session_id)
    assert list(audit_dir.glob("audit_*.jsonl"))

    result = manager.cleanup_session(session_id)

    assert result["cleanup_pending"] is False
    assert list(audit_dir.glob("audit_*.jsonl"))


def test_invalid_in_memory_session_is_discarded_without_poisoning_cleanup(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager(expire_hours=0)
    manager._sessions[""] = SessionInfo(session_id="")

    assert manager.active_dir_keys() == set()
    assert "" not in manager._sessions
    assert manager.cleanup_expired() == {"expired_sessions": 0, "cleaned_files": 0, "cleaned_jobs": 0}
    assert manager.cleanup_session("")["cleanup_pending"] is False


def test_orphan_cleanup_never_removes_active_session(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    session_id = f"active-{uuid.uuid4().hex}"
    session_manager.get_or_create(session_id)
    roots = [
        session_manager.get_session_raw_dir(session_id),
        session_manager.get_session_processed_dir(session_id),
        session_manager.get_session_als_dir(session_id),
        session_manager.get_session_spec_job_dir(session_id, "job").parent,
    ]
    old = time.time() - 24 * 3600
    for path in roots:
        os.utime(path, (old, old))

    try:
        result = cleanup_orphaned_session_dirs(max_age_hours=1)
        assert result["cleaned_dirs"] == 0
        assert result["skipped_dirs"] == 4
        assert all(path.exists() for path in roots)
    finally:
        session_manager.cleanup_session(session_id)


def test_orphan_cleanup_rechecks_activity_immediately_before_detach(tmp_path: Path, monkeypatch) -> None:
    """A session activated after the initial scan must not be removed."""
    monkeypatch.chdir(tmp_path)
    session_id = f"late-active-{uuid.uuid4().hex}"
    orphan_dir = session_manager.get_session_raw_dir(session_id)
    old = time.time() - 24 * 3600
    os.utime(orphan_dir, (old, old))
    real_delete = session_manager.delete_orphan_dir_if_inactive

    def activate_then_check(candidate: Path) -> bool:
        session_manager.get_or_create(session_id)
        return real_delete(candidate)

    monkeypatch.setattr(session_manager, "delete_orphan_dir_if_inactive", activate_then_check)
    try:
        result = cleanup_orphaned_session_dirs(max_age_hours=1)
        assert result == {"cleaned_dirs": 0, "skipped_dirs": 1}
        assert orphan_dir.exists()
    finally:
        monkeypatch.setattr(session_manager, "delete_orphan_dir_if_inactive", real_delete)
        session_manager.cleanup_session(session_id)


def test_cleanup_cancels_and_waits_for_live_worker(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    jobs = JobManager()
    manager.set_job_manager(jobs)
    session_id = "live-worker-session"
    job_id = "live-worker-job"
    manager.get_or_create(session_id)
    tracked = manager.get_session_als_dir(session_id) / "input.xlsx"
    tracked.write_bytes(b"input")
    manager.add_file(session_id, str(tracked))
    manager.add_job(session_id, job_id)
    jobs.create_job(job_id, owner_session_id=session_id)

    started = threading.Event()
    release = threading.Event()

    def work() -> None:
        jobs.update_job(job_id, state="running")
        started.set()
        release.wait(timeout=5)

    worker = threading.Thread(target=work)
    assert jobs.attach_worker(job_id, worker)
    worker.start()
    assert started.wait(timeout=1)

    result = manager.cleanup_session(session_id)
    assert result["deferred_jobs"] == 1
    assert jobs.get_job(job_id).state == "cancelling"
    assert tracked.exists()
    with pytest.raises(SessionClosingError):
        manager.get_or_create(session_id)
    assert manager.cancel_cleanup(session_id) is False

    release.set()
    worker.join(timeout=1)
    deadline = time.time() + 2
    while (tracked.exists() or jobs.get_job(job_id) is not None) and time.time() < deadline:
        time.sleep(0.02)
    assert not tracked.exists()
    assert jobs.get_job(job_id) is None


def test_cleanup_retries_detached_tombstone_after_delete_failure(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "delete-retry-session"
    manager.get_or_create(session_id)
    session_dir = manager.get_session_als_dir(session_id)
    (session_dir / "clinical.xlsx").write_bytes(b"clinical")

    real_rmtree = shutil.rmtree
    failed_once = False
    scheduled: list[str] = []

    def fail_first_tombstone(path, *args, **kwargs):
        nonlocal failed_once
        if not failed_once and Path(path).name.startswith(".cleanup-"):
            failed_once = True
            raise OSError("synthetic lock")
        return real_rmtree(path, *args, **kwargs)

    monkeypatch.setattr("src.web.session_manager.shutil.rmtree", fail_first_tombstone)
    monkeypatch.setattr(manager, "_schedule_drain_retry", scheduled.append)

    first = manager.cleanup_session(session_id)
    assert first["cleanup_pending"] is True
    assert first["errors"] == ["session_output_dirs_delete_failed"]
    assert scheduled == [session_id]
    assert manager.get_session_info(session_id) is not None
    with pytest.raises(SessionClosingError):
        with manager.operation(session_id):
            pass

    monkeypatch.setattr("src.web.session_manager.shutil.rmtree", real_rmtree)
    second = manager.cleanup_session(session_id)
    assert second["cleanup_pending"] is False
    assert second["errors"] == []
    assert manager.get_session_info(session_id) is None
    als_parent = Path("data/output/sessions")
    assert not list(als_parent.glob(f".cleanup-{manager.session_dir_key(session_id)}-*"))


def test_cleanup_retries_tracked_external_file_after_unlink_failure(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "external-delete-retry"
    manager.get_or_create(session_id)
    external = tmp_path / "legacy-output.xlsx"
    external.write_bytes(b"clinical")
    manager.add_file(session_id, str(external))

    real_unlink = Path.unlink
    failed_once = False
    scheduled: list[str] = []

    def fail_first_unlink(path: Path, *args, **kwargs):
        nonlocal failed_once
        if not failed_once and path.resolve() == external.resolve():
            failed_once = True
            raise OSError("synthetic lock")
        return real_unlink(path, *args, **kwargs)

    monkeypatch.setattr(Path, "unlink", fail_first_unlink)
    monkeypatch.setattr(manager, "_schedule_drain_retry", scheduled.append)
    first = manager.cleanup_session(session_id)
    assert first["cleanup_pending"] is True
    assert first["errors"] == ["tracked_file_delete_failed:OSError"]
    assert external.exists()
    assert manager.get_session_info(session_id) is not None
    assert scheduled == [session_id]

    monkeypatch.setattr(Path, "unlink", real_unlink)
    second = manager.cleanup_session(session_id)
    assert second["cleanup_pending"] is False
    assert second["errors"] == []
    assert not external.exists()
    assert manager.get_session_info(session_id) is None


def test_cleaned_session_capability_cannot_recreate_generation(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "recreate-session"
    manager.get_or_create(session_id)
    old_file = manager.get_session_als_dir(session_id) / "old.xlsx"
    old_file.write_bytes(b"old")
    manager.add_file(session_id, str(old_file))

    entered_delete = threading.Event()
    allow_delete = threading.Event()
    real_clean_dirs = manager._clean_session_output_dirs

    def slow_clean_dirs(target_session_id: str) -> bool:
        entered_delete.set()
        assert allow_delete.wait(timeout=2)
        return real_clean_dirs(target_session_id)

    monkeypatch.setattr(manager, "_clean_session_output_dirs", slow_clean_dirs)
    cleanup_thread = threading.Thread(target=manager.cleanup_session, args=(session_id,))
    cleanup_thread.start()
    assert entered_delete.wait(timeout=1)

    with pytest.raises(SessionClosingError):
        manager.get_or_create(session_id)

    allow_delete.set()
    cleanup_thread.join(timeout=1)
    assert not cleanup_thread.is_alive()
    assert not old_file.exists()
    with pytest.raises(SessionClosingError):
        manager.get_or_create(session_id)
    with pytest.raises(SessionClosingError):
        with manager.operation(session_id):
            pass


def test_active_session_registry_has_a_hard_capacity_limit() -> None:
    from src.web.session_manager import SessionCapacityError

    manager = SessionManager(max_active_sessions=1)
    manager.get_or_create("first-session")

    with pytest.raises(SessionCapacityError):
        manager.get_or_create("second-session")


def test_cleanup_waits_for_request_lease_before_recursive_delete(tmp_path: Path, monkeypatch) -> None:
    """Cleanup waits for a leased writer, then removes everything it published."""
    monkeypatch.chdir(tmp_path)
    manager = SessionManager()
    session_id = "detach-generation-session"
    writer_entered = threading.Event()
    release_writer = threading.Event()
    writer_finished = threading.Event()

    def write_during_request() -> None:
        with manager.operation(session_id):
            target = manager.get_session_als_dir(session_id) / "fresh.xlsx"
            writer_entered.set()
            assert release_writer.wait(timeout=2)
            target.write_bytes(b"fresh")
            assert manager.add_file(session_id, str(target))
        writer_finished.set()

    writer_thread = threading.Thread(target=write_during_request)
    writer_thread.start()
    assert writer_entered.wait(timeout=1)

    cleanup_thread = threading.Thread(target=manager.cleanup_session, args=(session_id,))
    cleanup_thread.start()
    time.sleep(0.05)
    assert cleanup_thread.is_alive(), "cleanup must wait for the active request lease"

    release_writer.set()
    writer_thread.join(timeout=2)
    cleanup_thread.join(timeout=2)
    assert writer_finished.is_set()
    assert not cleanup_thread.is_alive()
    assert not (Path("data/output/sessions") / manager.session_dir_key(session_id)).exists()
    assert manager.get_session_info(session_id) is None


def test_cleanup_lease_wait_is_bounded_and_retried(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    manager = SessionManager(cleanup_wait_timeout=0.02)
    session_id = "bounded-lease-wait"
    lease_entered = threading.Event()
    release_lease = threading.Event()

    def hold_lease() -> None:
        with manager.operation(session_id):
            lease_entered.set()
            assert release_lease.wait(timeout=2)

    lease_thread = threading.Thread(target=hold_lease)
    lease_thread.start()
    assert lease_entered.wait(timeout=1)

    started = time.monotonic()
    result = manager.cleanup_session(session_id)
    elapsed = time.monotonic() - started

    assert elapsed < 0.5
    assert result["cleanup_pending"] is True
    assert result["errors"] == ["request_lease_drain_pending"]

    release_lease.set()
    lease_thread.join(timeout=1)
    deadline = time.monotonic() + 2
    while manager.get_session_info(session_id) is not None and time.monotonic() < deadline:
        time.sleep(0.01)
    assert manager.get_session_info(session_id) is None
