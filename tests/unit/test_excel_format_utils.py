"""Unit tests for :mod:`src.spec_mapper.core.excel.format_utils`."""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.spec_mapper.core.excel import format_utils as fmt


@pytest.fixture
def workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb


@pytest.fixture
def populated_ws(workbook):
    ws = workbook.active
    ws.cell(row=1, column=1, value="A1")
    ws.cell(row=1, column=2, value="B1")
    ws.cell(row=2, column=1, value="A2")
    ws.cell(row=2, column=2, value="B2")
    return ws


class TestCopyAndApplyRowFormat:
    def test_roundtrip_copies_font_fill_border(self, populated_ws):
        cell = populated_ws.cell(row=1, column=1)
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.number_format = "0.00"

        snapshot = fmt.copy_row_format(populated_ws, 1)
        assert 1 in snapshot and 2 in snapshot

        fmt.apply_row_format(populated_ws, 2, snapshot)

        applied = populated_ws.cell(row=2, column=1)
        assert applied.font.bold is True
        assert applied.font.size == 14
        assert applied.alignment.horizontal == "center"
        assert applied.alignment.wrap_text is True
        assert applied.number_format == "0.00"

    def test_missing_values_tolerated(self, populated_ws):
        snapshot = {1: {"font": None, "fill": None, "border": None, "alignment": None, "number_format": None}}
        fmt.apply_row_format(populated_ws, 1, snapshot)

    def test_empty_row_returns_mapping(self, workbook):
        ws = workbook.active
        snapshot = fmt.copy_row_format(ws, 1)
        assert isinstance(snapshot, dict)


class TestUnmergeRowsInRange:
    def test_no_merges_returns_empty(self, populated_ws):
        assert fmt.unmerge_rows_in_range(populated_ws, 1, 2) == []

    def test_unmerge_single_range(self, populated_ws):
        populated_ws.merge_cells("A1:B1")
        assert len(populated_ws.merged_cells.ranges) == 1
        unmerged = fmt.unmerge_rows_in_range(populated_ws, 1, 1)
        assert unmerged == ["A1:B1"]
        assert len(populated_ws.merged_cells.ranges) == 0

    def test_skips_non_overlapping(self, populated_ws):
        populated_ws.merge_cells("A5:B5")
        unmerged = fmt.unmerge_rows_in_range(populated_ws, 1, 2)
        assert unmerged == []
        assert len(populated_ws.merged_cells.ranges) == 1

    def test_zero_amount_no_op(self, populated_ws):
        populated_ws.merge_cells("A1:B1")
        assert fmt.unmerge_rows_in_range(populated_ws, 1, 0) == []
        assert len(populated_ws.merged_cells.ranges) == 1


class TestFindLastRowWithValue:
    def test_finds_last_matching_row(self, populated_ws):
        populated_ws.cell(row=3, column=1, value="TARGET")
        populated_ws.cell(row=5, column=1, value="TARGET")
        populated_ws.cell(row=7, column=1, value="OTHER")
        row = fmt.find_last_row_with_value(populated_ws, column=1, value="TARGET")
        assert row == 5

    def test_case_insensitive_match(self, populated_ws):
        populated_ws.cell(row=3, column=1, value="supp")
        row = fmt.find_last_row_with_value(populated_ws, column=1, value="SUPP")
        assert row == 3

    def test_case_sensitive(self, populated_ws):
        populated_ws.cell(row=3, column=1, value="supp")
        row = fmt.find_last_row_with_value(populated_ws, column=1, value="SUPP", case_insensitive=False)
        assert row is None

    def test_none_returned_when_no_match(self, populated_ws):
        assert fmt.find_last_row_with_value(populated_ws, column=1, value="NOPE") is None

    def test_respects_start_row(self, populated_ws):
        populated_ws.cell(row=2, column=1, value="x")
        populated_ws.cell(row=5, column=1, value="x")
        row = fmt.find_last_row_with_value(populated_ws, column=1, value="x", start_row=3)
        assert row == 5

    def test_strips_whitespace(self, populated_ws):
        populated_ws.cell(row=3, column=1, value="  TARGET  ")
        row = fmt.find_last_row_with_value(populated_ws, column=1, value="TARGET")
        assert row == 3


class TestSetColumnWrapText:
    def test_applies_wrap_text(self, populated_ws):
        updated = fmt.set_column_wrap_text(populated_ws, column=1, start_row=1)
        assert updated == 2
        for row_idx in (1, 2):
            assert populated_ws.cell(row=row_idx, column=1).alignment.wrap_text is True

    def test_preserves_existing_alignment(self, populated_ws):
        populated_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", indent=2)
        fmt.set_column_wrap_text(populated_ws, column=1, start_row=1)
        cell = populated_ws.cell(row=1, column=1)
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.indent == 2
        assert cell.alignment.wrap_text is True


class TestHighlightSheetTabs:
    def test_highlights_only_listed_sheets(self, workbook):
        workbook.create_sheet("A")
        workbook.create_sheet("B")
        workbook.create_sheet("C")

        highlighted = fmt.highlight_sheet_tabs(workbook, modified_sheets=["A", "C"], color="FF0000")
        assert highlighted == 2
        assert workbook["A"].sheet_properties.tabColor.rgb.endswith("FF0000")
        assert workbook["C"].sheet_properties.tabColor.rgb.endswith("FF0000")
        assert workbook["B"].sheet_properties.tabColor is None

    def test_clears_existing_colors(self, workbook):
        workbook.create_sheet("A")
        workbook.create_sheet("B")
        workbook["A"].sheet_properties.tabColor = "ABCDEF"

        fmt.highlight_sheet_tabs(workbook, modified_sheets=["B"], color="FFFF00", clear_existing=True)
        assert workbook["A"].sheet_properties.tabColor is None

    def test_skips_unknown_sheets(self, workbook):
        highlighted = fmt.highlight_sheet_tabs(workbook, modified_sheets=["nonexistent"], color="FFFF00")
        assert highlighted == 0

    def test_empty_modified_returns_zero(self, workbook):
        assert fmt.highlight_sheet_tabs(workbook, modified_sheets=[], color="FFFF00") == 0


class TestFindMergedRangesOverlapping:
    def test_finds_overlaps(self, populated_ws):
        populated_ws.merge_cells("A1:B2")
        populated_ws.merge_cells("A5:B6")
        overlaps = fmt.find_merged_ranges_overlapping(populated_ws, 1, 3)
        assert len(overlaps) == 1
        assert str(overlaps[0]) == "A1:B2"

    def test_no_overlap(self, populated_ws):
        populated_ws.merge_cells("A10:B10")
        assert fmt.find_merged_ranges_overlapping(populated_ws, 1, 3) == []
