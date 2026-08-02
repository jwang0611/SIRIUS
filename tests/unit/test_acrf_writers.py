"""End-to-end extraction → writers → als2sdtm converter round-trip (integration)."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.processors.acrf import (
    extract_acrf,
    write_als2sdtm_xlsx,
    write_processed_json,
    write_processed_xlsx,
)
from src.processors.acrf.models import AcrfConfig
from src.processors.acrf.records import assemble_records
from src.processors.acrf.writers import ALS2SDTM_SHEET_NAME
from src.processors.als_converter import convert_als2sdtm

pytestmark = pytest.mark.integration


def test_extract_writes_all_outputs_and_roundtrips(sample_acrf_pdf: Path, tmp_path: Path):
    result = extract_acrf(str(sample_acrf_pdf))
    assert result.records  # Cover skipped, two forms extracted

    json_path = tmp_path / "sample.json"
    xlsx_path = tmp_path / "sample.xlsx"
    als_path = tmp_path / "sample_ALS2SDTM.xlsx"
    write_processed_json(json_path, result.records)
    write_processed_xlsx(xlsx_path, result.records)
    write_als2sdtm_xlsx(als_path, result.records)

    # Processed JSON: exact 4-key schema, non-empty metadata_table.
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    assert all(
        set(r) == {"metadata_table", "metadata_variable", "annotation_table", "annotation_variable"} for r in rows
    )
    assert all(r["metadata_table"] for r in rows)

    # Sibling xlsx keeps the num order column first (pipeline coverage pass).
    df_sib = pd.read_excel(xlsx_path)
    assert next(iter(df_sib.columns)) == "num"

    # Portable als2sdtm workbook: sheet named eCRF + converter-recognised headers.
    xl = pd.ExcelFile(als_path)
    assert xl.sheet_names == [ALS2SDTM_SHEET_NAME]
    df_als = pd.read_excel(als_path, sheet_name=ALS2SDTM_SHEET_NAME)
    assert {"表名", "表", "变量名", "变量", "SDTM_Domain", "SDTM_Variable"}.issubset(df_als.columns)

    # Round-trip: convert_als2sdtm re-ingests it with zero extra flags.
    outputs = convert_als2sdtm(str(als_path), str(tmp_path), output_name="roundtrip", output_format="json")
    rt = json.loads(Path(outputs["json"]).read_text(encoding="utf-8"))
    assert rt
    assert {
        "annotation_table",
        "metadata_table",
        "annotation_variable",
        "metadata_variable",
        "SDTM_Domain",
        "SDTM_Variable",
    } <= set(rt[0])
    assert rt[0]["annotation_table"] == "Demographics"


def test_repeated_body_field_labels_are_not_dropped(acrf_pdf_builder, tmp_path: Path):
    pdf = tmp_path / "repeat.pdf"
    acrf_pdf_builder(
        pdf,
        [
            ("V1", "Visit 1", ["Assessment Date", "Weight"]),
            ("V2", "Visit 2", ["Assessment Date", "Height"]),
            ("V3", "Visit 3", ["Assessment Date", "Temperature"]),
        ],
    )
    result = extract_acrf(str(pdf))
    labels = [r.annotation_variable for r in result.records]

    # "Assessment Date" recurs mid-page across all three forms — a real field,
    # not header/footer boilerplate — so all three occurrences must survive.
    assert labels.count("Assessment Date") == 3
    assert len(result.records) == 6


def test_llm_merges_and_never_overwrites_deterministic(acrf_pdf_builder, tmp_path: Path):
    pdf = tmp_path / "merge.pdf"
    acrf_pdf_builder(pdf, [("Demographics", "Demographics", ["Subject ID", "Date of Birth", "Sex", "Age"])])

    class _Fake:
        def generate_content(self, prompt, **_kw):
            return '["Subject ID"]'  # model returns only a subset of the fields

    # llm_min_fields high enough that the LLM runs for this form.
    cfg = AcrfConfig(use_llm=True, llm_min_fields=99)
    result = extract_acrf(str(pdf), cfg=cfg, use_llm=True, client=_Fake())

    fields = {r.annotation_variable for r in result.records}
    assert fields == {"Subject ID", "Date of Birth", "Sex", "Age"}  # nothing dropped


def test_extraction_stats_count_forms_without_fields(acrf_pdf_builder, tmp_path: Path):
    pdf = tmp_path / "partial.pdf"
    acrf_pdf_builder(
        pdf,
        [
            ("Empty Form", "Empty Form", []),
            ("Demographics", "Demographics", ["Subject ID"]),
        ],
    )

    result = extract_acrf(str(pdf))

    assert result.stats["forms_total"] == 2
    assert result.stats["forms_with_fields"] == 1
    assert result.stats["forms_without_fields"] == 1
    assert "form 'Empty Form': no fields extracted" in result.warnings


def test_writers_neutralize_formula_injection_without_mutating_content(tmp_path: Path):
    labels = ['=HYPERLINK("http://evil","x")', "+1", "-2", "@cmd", "normal"]
    recs = assemble_records([("Form", labels)], AcrfConfig())
    xp = tmp_path / "p.xlsx"
    ap = tmp_path / "a.xlsx"
    write_processed_xlsx(xp, recs)
    write_als2sdtm_xlsx(ap, recs)

    # No cell is a formula (injection surface removed)...
    for path in (xp, ap):
        ws = load_workbook(path).active
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", f"{path.name}!{cell.coordinate} became a formula"

    # ...and the exact original strings survive (no quote-prefix mutation), so
    # read-back and the converter round-trip stay faithful.
    assert pd.read_excel(xp)["annotation_variable"].astype(str).tolist() == labels
    assert pd.read_excel(ap, sheet_name=ALS2SDTM_SHEET_NAME)["变量名"].astype(str).tolist() == labels
