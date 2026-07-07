"""Tests for SDTM IG 3.4 template compatibility in the spec mapper.

The IG 3.4 template differs from the v3.2 (v1.1) template in two ways the
code cares about:
  * it has **no VISIT sheet** (only SV), so VISIT/VISITNUM hyperlinks must
    degrade gracefully instead of pointing at a missing sheet;
  * its CONTENT "class" column is **English** ("Findings") rather than the
    Chinese "发现", so findings-domain detection must accept both.

These tests exercise those two behaviours directly, plus a lightweight check
that the English-header template loads via :class:`ExcelReader`.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.spec_mapper.core.excel_reader import ExcelReader
from src.spec_mapper.core.excel_writer import ExcelWriter
from src.spec_mapper.utils.config_loader import ConfigLoader

V34_TEMPLATE = Path("data/knowledge_base/template_spec/SDTM_template_IG3.4.xlsx")
V32_TEMPLATE = Path("data/knowledge_base/template_spec/SDTM_template_IG3.2.xlsx")

# Mirrors src/spec_mapper/config/config.yaml fixed_variables.VISIT / VISITNUM
VISIT_RULES = {
    "VISIT": {
        "exclude_domains": ["TA", "TD", "TE", "TI", "TS", "TV"],
        "all": {"source": None, "transformation": '=HYPERLINK("#VISIT!A1","转换定义参考VISIT表")'},
    },
    "VISITNUM": {
        "exclude_domains": ["TA", "TD", "TE", "TI", "TS", "TV"],
        "all": {"source": None, "transformation": '=HYPERLINK("#VISIT!A1","转换定义参考VISIT表")'},
    },
}


def _writer_for(tmp_path: Path, build) -> ExcelWriter:
    """Build a workbook via *build(wb)*, save it, return an ExcelWriter on it."""
    wb = Workbook()
    build(wb)
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return ExcelWriter(path)


def _domain_sheet_with_visit(wb: Workbook, *, add_visit_sheet: bool) -> None:
    ws = wb.active
    ws.title = "VS"
    ws.cell(row=14, column=1, value="VISIT")
    ws.cell(row=15, column=1, value="VISITNUM")
    if add_visit_sheet:
        wb.create_sheet("VISIT")


class TestVisitHyperlinkGuard:
    def test_visit_hyperlink_skipped_when_no_visit_sheet(self, tmp_path: Path) -> None:
        writer = _writer_for(tmp_path, lambda wb: _domain_sheet_with_visit(wb, add_visit_sheet=False))
        writer.apply_fixed_variable_rules("VS", VISIT_RULES)

        ws = writer.workbook["VS"]
        for row in (14, 15):  # VISIT, VISITNUM
            value = ws.cell(row=row, column=8).value  # column H = transformation
            assert "#VISIT!" not in str(value or ""), f"row {row} kept a dead VISIT hyperlink"

    def test_visit_hyperlink_written_when_visit_sheet_exists(self, tmp_path: Path) -> None:
        writer = _writer_for(tmp_path, lambda wb: _domain_sheet_with_visit(wb, add_visit_sheet=True))
        writer.apply_fixed_variable_rules("VS", VISIT_RULES)

        ws = writer.workbook["VS"]
        value = str(ws.cell(row=14, column=8).value or "")
        assert "#VISIT!" in value and "HYPERLINK" in value.upper()


class TestFindingsClassLabel:
    def _content_writer(self, tmp_path: Path, class_label: str) -> ExcelWriter:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "CONTENT"
            ws.cell(row=7, column=1, value="VS")  # domain (column A)
            ws.cell(row=7, column=5, value=class_label)  # class (column E)

        return _writer_for(tmp_path, build)

    @pytest.mark.parametrize("class_label", ["Findings", "发现"])
    def test_findings_class_adds_sdtm_sv(self, tmp_path: Path, class_label: str) -> None:
        writer = self._content_writer(tmp_path, class_label)
        writer.update_content_f_column(domain_sources={"VS": {"VSRAW"}})

        f_value = str(writer.workbook["CONTENT"].cell(row=7, column=6).value or "")
        assert "SDTM.SV" in f_value

    def test_non_findings_class_omits_sdtm_sv(self, tmp_path: Path) -> None:
        writer = self._content_writer(tmp_path, "Events")
        writer.update_content_f_column(domain_sources={"VS": {"VSRAW"}})

        f_value = str(writer.workbook["CONTENT"].cell(row=7, column=6).value or "")
        assert "SDTM.SV" not in f_value


@pytest.mark.skipif(not V34_TEMPLATE.exists(), reason="IG 3.4 template not present")
def test_reader_parses_english_v34_headers() -> None:
    """The IG 3.4 template uses English headers; the reader's *_en fallbacks
    must still locate the variable-name column and read domain rows."""
    reader = ExcelReader(V34_TEMPLATE)
    records = reader.read_template_records(["AE"])

    variable_names = {r.variable_name.upper() for r in records}
    assert {"STUDYID", "DOMAIN"} <= variable_names


# Minimal fixed-variable rule matching config.yaml's STUDYID (non-DM) source="指定"
STUDYID_RULE = {"STUDYID": {"other": {"source": "指定", "transformation": "DM.STUDYID"}}}


def _studyid_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "VS"  # non-DM domain
    ws.cell(row=14, column=1, value="STUDYID")


class TestSourceLabelTranslation:
    def test_config_ig34_extends_base_and_adds_overrides(self) -> None:
        import src.spec_mapper as sm_pkg

        cfg = ConfigLoader(Path(sm_pkg.__file__).parent / "config" / "config_ig34.yaml")
        assert cfg.get("fixed_variables"), "should inherit fixed_variables from base config"
        assert cfg.get("source_label_overrides") == {"指定": "Assigned", "衍生": "Derived"}

    def test_source_translated_when_override_set(self, tmp_path: Path) -> None:
        writer = _writer_for(tmp_path, _studyid_sheet)
        writer.source_label_overrides = {"指定": "Assigned", "衍生": "Derived"}
        writer.apply_fixed_variable_rules("VS", STUDYID_RULE)
        assert writer.workbook["VS"].cell(row=14, column=6).value == "Assigned"

    def test_source_stays_chinese_without_override(self, tmp_path: Path) -> None:
        writer = _writer_for(tmp_path, _studyid_sheet)  # no overrides (v3.2 path)
        writer.apply_fixed_variable_rules("VS", STUDYID_RULE)
        assert writer.workbook["VS"].cell(row=14, column=6).value == "指定"

    def test_preserve_guard_keeps_english_assigned(self, tmp_path: Path) -> None:
        """With a translation active, an existing English 'Assigned' source on a
        mapped row must be preserved (not overwritten to 'CRF')."""

        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "VS"
            ws.cell(row=14, column=1, value="VSDRVFL")
            ws.cell(row=14, column=6, value="Assigned")  # template's English source
            ws.cell(row=14, column=8, value="some transformation")  # has a transformation

        writer = _writer_for(tmp_path, build)
        writer.source_label_overrides = {"指定": "Assigned", "衍生": "Derived"}
        writer.update_domain_source_column("VS")
        assert writer.workbook["VS"].cell(row=14, column=6).value == "Assigned"


@pytest.mark.skipif(not V34_TEMPLATE.exists(), reason="IG 3.4 template not present")
def test_select_config_path_picks_ig34_for_v34_template() -> None:
    from src.spec_mapper import SpecMapper

    assert SpecMapper._select_config_path(V34_TEMPLATE).name == "config_ig34.yaml"


@pytest.mark.skipif(not V32_TEMPLATE.exists(), reason="v3.2 template not present")
def test_select_config_path_picks_base_for_v32_template() -> None:
    from src.spec_mapper import SpecMapper

    assert SpecMapper._select_config_path(V32_TEMPLATE).name == "config.yaml"
