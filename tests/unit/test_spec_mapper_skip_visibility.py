"""Per-item skips inside batch writes must stay visible in the write result.

``SpecMapper._guard`` records ``record_written(n)`` from a writer method's
integer return value, so a batch method that silently passes over SOME of its
planned items used to report ``attempted == written`` and no issue at all.
These tests pin the fix: every skipped item is queued by the writer and drained
by the guard into a structured :class:`WriteIssue` (code / sheet / variable),
while the workbook produced by the run stays byte-for-cell identical.

All fixtures are synthetic, metadata-only workbooks written into ``tmp_path``;
no template, network, or LLM is involved.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.spec_mapper import SpecMapper
from src.spec_mapper.core.excel_writer import ExcelWriter
from src.spec_mapper.models.write_result import STAGE_EXTERNAL_CODING, WriteResult

# Two MedDRA-style variables, shaped like the ``external_coding_variables``
# config entries (metadata only — no clinical values).
AELLT = {
    "name": "AELLT",
    "label": "低位语",
    "type": "Char",
    "length": "200",
    "controlled_terms": "MedDRA",
    "source": "指定",
    "core": "期望",
    "transformation": "连接外部coding文件.",
}
AEDECOD = {
    "name": "AEDECOD",
    "label": "标准化名称",
    "type": "Char",
    "length": "200",
    "controlled_terms": "MedDRA",
    "source": "指定",
    "core": "必需",
    "transformation": "连接外部coding文件.",
}
AEDVER = {
    "name": "AEDVER",
    "label": "字典版本",
    "type": "Char",
    "length": "200",
    "source": "指定",
    "transformation": "连接外部coding文件.",
}


def _domain_workbook(path: Path, variables: tuple[str, ...]) -> Path:
    """A minimal AE domain sheet: column A variable names, column I type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AE"
    ws.cell(row=13, column=1).value = "Variable Name"
    for offset, name in enumerate(variables):
        row = 14 + offset
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=9).value = "SDTM"
    wb.save(path)
    return path


def _writer(tmp_path: Path, variables: tuple[str, ...], name: str = "ae.xlsx") -> ExcelWriter:
    return ExcelWriter(_domain_workbook(tmp_path / name, variables))


def _sheet_snapshot(writer: ExcelWriter, sheet_name: str) -> list[tuple]:
    """Value + fill + number-format snapshot of every cell in a sheet."""
    ws = writer.workbook[sheet_name]
    snapshot = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            snapshot.append((row, col, cell.value, cell.fill.start_color.rgb, cell.number_format))
    return snapshot


def _mapper(tmp_path: Path) -> SpecMapper:
    """A SpecMapper usable only for its ``_guard`` boundary (no processing)."""
    als = tmp_path / "als.xlsx"
    template = tmp_path / "template.xlsx"
    for path in (als, template):
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(path)
    return SpecMapper(als_file=als, template_file=template, als_sheet="Sheet1", log_level="CRITICAL")


# ---------------------------------------------------------------------------
# update_existing_variables: a configured variable missing from the sheet
# ---------------------------------------------------------------------------
class TestUpdateExistingVariablesSkips:
    def test_missing_variable_is_queued_with_code_sheet_and_name(self, tmp_path: Path) -> None:
        writer = _writer(tmp_path, ("AETERM", "AELLT"))

        updated = writer.update_existing_variables(sheet_name="AE", variables=[AELLT, AEDECOD])

        # Only the present variable was written; the missing one is reported.
        assert updated == 1
        issues = writer.drain_pending_skips(STAGE_EXTERNAL_CODING)
        assert [(i.code, i.sheet, i.variable, i.operation) for i in issues] == [
            ("variable_not_found", "AE", "AEDECOD", "update_existing_variables")
        ]
        assert issues[0].stage == STAGE_EXTERNAL_CODING
        writer.close()

    def test_one_issue_per_missing_item(self, tmp_path: Path) -> None:
        writer = _writer(tmp_path, ("AETERM",))

        updated = writer.update_existing_variables(sheet_name="AE", variables=[AELLT, AEDECOD])

        assert updated == 0
        assert [i.variable for i in writer.drain_pending_skips(STAGE_EXTERNAL_CODING)] == ["AELLT", "AEDECOD"]
        writer.close()

    def test_complete_batch_queues_nothing(self, tmp_path: Path) -> None:
        writer = _writer(tmp_path, ("AELLT", "AEDECOD"))

        updated = writer.update_existing_variables(sheet_name="AE", variables=[AELLT, AEDECOD])

        assert updated == 2
        assert writer.drain_pending_skips(STAGE_EXTERNAL_CODING) == []
        writer.close()

    def test_drain_clears_the_queue(self, tmp_path: Path) -> None:
        """A drained skip must not leak into the next guarded call's stage."""
        writer = _writer(tmp_path, ("AETERM",))

        writer.update_existing_variables(sheet_name="AE", variables=[AEDECOD])
        assert len(writer.drain_pending_skips(STAGE_EXTERNAL_CODING)) == 1
        assert writer.drain_pending_skips("styles") == []
        writer.close()


# ---------------------------------------------------------------------------
# add_external_coding_variables: a configured variable already in the sheet
# ---------------------------------------------------------------------------
class TestAddExternalCodingVariablesSkips:
    def test_already_present_variable_is_reported(self, tmp_path: Path) -> None:
        writer = _writer(tmp_path, ("AETERM", "AEDVER"))

        inserted = writer.add_external_coding_variables(
            sheet_name="AE", variables=[AEDVER, AELLT], variable_type="SUPP"
        )

        # AEDVER already exists (idempotency), AELLT is genuinely inserted.
        assert inserted == 1
        issues = writer.drain_pending_skips(STAGE_EXTERNAL_CODING)
        assert [(i.code, i.sheet, i.variable, i.operation) for i in issues] == [
            ("variable_already_present", "AE", "AEDVER", "add_external_coding_variables")
        ]
        writer.close()

    def test_fresh_sheet_inserts_all_and_queues_nothing(self, tmp_path: Path) -> None:
        writer = _writer(tmp_path, ("AETERM",))

        inserted = writer.add_external_coding_variables(
            sheet_name="AE", variables=[AEDVER, AELLT], variable_type="SUPP"
        )

        assert inserted == 2
        assert writer.drain_pending_skips(STAGE_EXTERNAL_CODING) == []
        writer.close()


# ---------------------------------------------------------------------------
# _guard accounting: skips reach the WriteResult the job / UI reads
# ---------------------------------------------------------------------------
class TestGuardRecordsPerItemSkips:
    def test_partial_batch_splits_written_and_skipped(self, tmp_path: Path) -> None:
        mapper = _mapper(tmp_path)
        writer = _writer(tmp_path, ("AETERM", "AELLT"))
        result = WriteResult()

        mapper._guard(
            result,
            STAGE_EXTERNAL_CODING,
            "update_existing_variables",
            writer.update_existing_variables,
            guard_sheet="AE",
            guard_code="external_coding_failed",
            sheet_name="AE",
            variables=[AELLT, AEDECOD],
        )

        stage = result.stages[STAGE_EXTERNAL_CODING]
        assert (stage.attempted, stage.written, stage.skipped) == (2, 1, 1)
        assert stage.attempted == stage.written + stage.skipped + len(stage.errors)
        assert [(w.code, w.variable) for w in stage.warnings] == [("variable_not_found", "AEDECOD")]
        writer.close()

    def test_fully_skipped_batch_reports_items_not_generic_no_op(self, tmp_path: Path) -> None:
        """Precise per-item codes replace the whole-call 'no_op' fallback."""
        mapper = _mapper(tmp_path)
        writer = _writer(tmp_path, ("AETERM",))
        result = WriteResult()

        mapper._guard(
            result,
            STAGE_EXTERNAL_CODING,
            "update_existing_variables",
            writer.update_existing_variables,
            guard_sheet="AE",
            guard_code="external_coding_failed",
            sheet_name="AE",
            variables=[AELLT, AEDECOD],
        )

        stage = result.stages[STAGE_EXTERNAL_CODING]
        assert (stage.attempted, stage.written, stage.skipped) == (2, 0, 2)
        assert {w.code for w in stage.warnings} == {"variable_not_found"}
        assert "no_op" not in {w.code for w in stage.warnings}
        writer.close()

    def test_no_op_fallback_still_applies_without_item_skips(self, tmp_path: Path) -> None:
        """A call that reports neither a mutation nor an item keeps 'no_op'."""
        mapper = _mapper(tmp_path)
        writer = _writer(tmp_path, ("AETERM",))
        result = WriteResult()

        mapper._guard(
            result,
            STAGE_EXTERNAL_CODING,
            "update_existing_variables",
            writer.update_existing_variables,
            guard_sheet="AE",
            guard_code="external_coding_failed",
            sheet_name="AE",
            variables=[],
        )

        stage = result.stages[STAGE_EXTERNAL_CODING]
        assert (stage.attempted, stage.written, stage.skipped) == (1, 0, 1)
        assert [w.code for w in stage.warnings] == ["no_op"]
        writer.close()

    def test_skips_surface_in_the_serialized_payload(self, tmp_path: Path) -> None:
        """The job/UI reads write_result.to_dict(); skips must be in there."""
        mapper = _mapper(tmp_path)
        writer = _writer(tmp_path, ("AETERM", "AEDVER"))
        result = WriteResult()

        mapper._guard(
            result,
            STAGE_EXTERNAL_CODING,
            "add_external_coding_variables",
            writer.add_external_coding_variables,
            guard_sheet="AE",
            guard_code="external_coding_failed",
            sheet_name="AE",
            variables=[AEDVER, AELLT],
            variable_type="SUPP",
        )

        payload = result.to_dict()
        assert payload["summary"] == {"attempted": 2, "written": 1, "skipped": 1, "warnings": 1, "errors": 0}
        assert payload["warnings"][0]["code"] == "variable_already_present"
        assert payload["warnings"][0]["sheet"] == "AE"
        assert payload["warnings"][0]["variable"] == "AEDVER"
        # No path, message, or clinical value may ride along.
        assert payload["warnings"][0]["detail"] is None
        assert result.has_write_problems is True
        writer.close()

    def test_issues_without_a_variable_keep_the_legacy_shape(self, tmp_path: Path) -> None:
        """``variable`` is additive: issue kinds that predate it are unchanged."""
        mapper = _mapper(tmp_path)
        writer = _writer(tmp_path, ("AETERM",))
        result = WriteResult()

        mapper._guard(
            result,
            STAGE_EXTERNAL_CODING,
            "update_existing_variables",
            writer.update_existing_variables,
            guard_sheet="AE",
            guard_code="external_coding_failed",
            sheet_name="AE",
            variables=[],
        )

        no_op = result.to_dict()["warnings"][0]
        assert set(no_op) == {"code", "stage", "operation", "sheet", "row", "column", "detail"}
        writer.close()


# ---------------------------------------------------------------------------
# Regression: reporting only — the workbook itself must be unchanged
# ---------------------------------------------------------------------------
class TestWorkbookUnchangedByReporting:
    @pytest.mark.parametrize(
        ("existing", "call"),
        [
            (("AETERM", "AELLT"), "update"),
            (("AETERM", "AEDVER"), "insert"),
        ],
    )
    def test_same_cells_with_and_without_skip_recording(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, existing: tuple[str, ...], call: str
    ) -> None:
        """Run the same batch twice on identical copies — once with skip
        recording disabled (pre-fix behaviour) — and compare every cell."""

        def run(name: str) -> tuple[int, list[tuple]]:
            writer = _writer(tmp_path, existing, name=name)
            if call == "update":
                count = writer.update_existing_variables(sheet_name="AE", variables=[AELLT, AEDECOD])
            else:
                count = writer.add_external_coding_variables(
                    sheet_name="AE", variables=[AEDVER, AELLT], variable_type="SUPP"
                )
            snapshot = _sheet_snapshot(writer, "AE")
            writer.close()
            return count, snapshot

        with_recording = run("with_recording.xlsx")

        monkeypatch.setattr(ExcelWriter, "_record_pending_skip", lambda *a, **k: None)
        without_recording = run("without_recording.xlsx")

        assert with_recording == without_recording
