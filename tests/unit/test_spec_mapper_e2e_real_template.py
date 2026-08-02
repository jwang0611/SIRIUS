"""End-to-end Spec Mapper tests against the **real** SDTM templates.

These tests copy the repo-tracked IG 3.2 / IG 3.4 templates into ``tmp_path``,
drive the full :meth:`SpecMapper.process` pipeline with a programmatically
generated *metadata-only* ALS workbook, then re-open the saved workbook and
assert on-disk behaviour (never in-memory mock calls).

Collectively across IG 3.2 and IG 3.4 they cover:
  1. normal cell update            5. CODELIST insert          9. generated-cell highlight
  2. SUPP row insertion            6. formula preserve/generate 10. merged cells intact
  3. QNAM / QVAL                   7. hyperlink preserve/gen    11. repeated run (no dupes)
  4. CODELIST merge (unit test)    8. style preservation        12. recoverable write failure
                                                                13. fatal save failure

Guarantees enforced here: the original template file is never modified, all
outputs stay in ``tmp_path``, no LLM / network is used, and no artifact is
committed.
"""

from __future__ import annotations

import hashlib
import logging
import shutil
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from src.spec_mapper import SpecMapper
from src.spec_mapper.core.excel_writer import ExcelWriter
from src.spec_mapper.models.write_result import RecoverableWriteError

REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = REPO_ROOT / "data/knowledge_base/template_spec"
TEMPLATES = {
    "IG3.2": TEMPLATE_DIR / "SDTM_template_IG3.2.xlsx",
    "IG3.4": TEMPLATE_DIR / "SDTM_template_IG3.4.xlsx",
}
# IG 3.4 has no per-domain TEST sheets (e.g. EGTEST); creating conditional TEST
# sheets there is a legitimate skip, so disable it for the clean-run fixture.
CREATE_TEST_SHEETS = {"IG3.2": True, "IG3.4": False}

# Domain present in both templates but absent from the synthetic ALS below;
# used as an "untouched" reference for preservation assertions.
UNTOUCHED_DOMAIN = "CE"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_synthetic_als(path: Path) -> None:
    """Write a metadata-only synthetic ALS workbook (no clinical values).

    Rows exercise a simple mapping (DM.STUDYID), a SUPP/QNAM-QVAL mapping
    (DM SUPPQUAL SITENAM), and a conditional TESTCD mapping (EG) that feeds the
    CODELIST sheet — all targeting domains present in both real templates.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["表", "变量", "变量名", "组件类型", "编码名称", "metadata_table", "SDTM_Domain", "SDTM_Variable"])
    for row in [
        ["SUBJECT", "STUDYID", "研究代码", "文本", None, "SUBJECT", "DM", "STUDYID"],
        ["SUBJECT", "SITENAME", "中心名称", "文本", None, "SUBJECT", "DM", "QVAL when QNAM=SITENAM"],
        ["EG_RAW", "QTCF", "QT间期", "单选", "EG_CT", "EG_RAW", "EG", "EGORRES when EGTESTCD=QT"],
    ]:
        ws.append(row)
    wb.save(path)


def _run_pipeline(base: Path, ig: str, *, create_test_sheets: bool | None = None) -> SimpleNamespace:
    src = TEMPLATES[ig]
    tpl = base / f"template_{ig}.xlsx"
    shutil.copy2(src, tpl)
    als = base / "als.xlsx"
    write_synthetic_als(als)
    out = base / "out.xlsx"
    if create_test_sheets is None:
        create_test_sheets = CREATE_TEST_SHEETS[ig]
    mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
    stats = mapper.process(output_file=out, highlight=True, create_test_sheets=create_test_sheets)
    return SimpleNamespace(ig=ig, src=src, tpl=tpl, als=als, out=out, stats=stats)


@pytest.fixture(scope="module", params=["IG3.2", "IG3.4"])
def e2e(request, tmp_path_factory):
    ig = request.param
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    run = None
    try:
        base = tmp_path_factory.mktemp(f"e2e_{ig.replace('.', '_')}")
        src_hash_before = _sha256(TEMPLATES[ig])
        run = _run_pipeline(base, ig)
        run.src_hash_before = src_hash_before
        # Load the saved output and the pristine copied template once each; the
        # template copy is the untouched reference for preservation assertions.
        run.wb = load_workbook(run.out)
        run.tpl_wb = load_workbook(run.tpl)
        yield run
    finally:
        if run is not None:
            if hasattr(run, "wb"):
                run.wb.close()
            if hasattr(run, "tpl_wb"):
                run.tpl_wb.close()
        logging.disable(logging.NOTSET)


def _find_row(ws, variable: str, start: int = 14) -> int | None:
    for r in range(start, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == variable.upper():
            return r
    return None


# ---------------------------------------------------------------------------
# stats structure + clean state
# ---------------------------------------------------------------------------
def test_stats_expose_planned_and_actual(e2e) -> None:
    stats = e2e.stats
    assert "planned" in stats and "actual" in stats and "write_result" in stats
    planned, actual = stats["planned"], stats["actual"]
    assert planned["cell_updates"] >= 1 and planned["supp_rows"] >= 1
    # actual is the real write outcome, not len(updates)
    assert set(actual) == {"attempted", "written", "skipped", "warnings", "errors"}
    assert actual["written"] >= 1
    # This fixture is configured for a fully-successful run.
    assert actual["written"] == actual["attempted"]
    assert actual["errors"] == 0


def test_original_template_not_modified(e2e) -> None:
    assert _sha256(e2e.src) == e2e.src_hash_before


# ---------------------------------------------------------------------------
# 1. normal cell update
# ---------------------------------------------------------------------------
def test_normal_cell_update(e2e) -> None:
    dm = e2e.wb["DM"]
    row = _find_row(dm, "STUDYID")
    assert row is not None
    assert dm.cell(row=row, column=8).value == "[RAW]SUBJECT.STUDYID"


# ---------------------------------------------------------------------------
# 2 + 3. SUPP row insertion + QNAM/QVAL
# ---------------------------------------------------------------------------
def test_supp_row_and_qnam_qval(e2e) -> None:
    dm = e2e.wb["DM"]
    supp_rows = [r for r in range(14, dm.max_row + 1) if str(dm.cell(row=r, column=9).value or "").upper() == "SUPP"]
    assert supp_rows, "expected at least one SUPP row in DM"
    qnam_row = _find_row(dm, "SITENAM")  # QNAM value becomes the SUPP variable name (A)
    assert qnam_row in supp_rows
    # QVAL source (H) references the raw variable.
    assert "[RAW]SUBJECT.SITENAME" in str(dm.cell(row=qnam_row, column=8).value or "")


# ---------------------------------------------------------------------------
# 5. CODELIST insert
# ---------------------------------------------------------------------------
def test_codelist_insert(e2e) -> None:
    cl = e2e.wb["CODELIST"]
    hits = [
        r
        for r in range(3, cl.max_row + 1)
        if str(cl.cell(row=r, column=1).value or "").upper() == "EG"
        and str(cl.cell(row=r, column=2).value or "").upper() == "EGTESTCD"
        and str(cl.cell(row=r, column=6).value or "").upper() == "QT"
    ]
    assert len(hits) == 1


# ---------------------------------------------------------------------------
# 6. formulas preserved + generated
# ---------------------------------------------------------------------------
def test_formulas_generated(e2e) -> None:
    eg = e2e.wb["EG"]
    testcd_row = _find_row(eg, "EGTESTCD")
    assert testcd_row is not None
    testcd_formula = str(eg.cell(row=testcd_row, column=8).value or "")
    assert testcd_formula.startswith("=HYPERLINK(") and "CODELIST" in testcd_formula
    # xxSEQ derived formula generated
    seq_row = _find_row(eg, "EGSEQ")
    assert seq_row is not None, "real template must contain EGSEQ"
    assert "=CONCATENATE(" in str(eg.cell(row=seq_row, column=8).value or "")
    # D10 sort-key lookup formula generated
    assert str(eg.cell(row=10, column=4).value or "").startswith("=IFERROR(INDEX(")


def test_untouched_domain_formulas_preserved(e2e) -> None:
    """A domain absent from the ALS must keep its template formulas verbatim."""
    assert UNTOUCHED_DOMAIN in e2e.tpl_wb.sheetnames
    ref = e2e.tpl_wb[UNTOUCHED_DOMAIN]
    got = e2e.wb[UNTOUCHED_DOMAIN]
    mismatches = 0
    for r in range(1, ref.max_row + 1):
        for c in range(1, ref.max_column + 1):
            if ref.cell(row=r, column=c).value != got.cell(row=r, column=c).value:
                mismatches += 1
    assert mismatches == 0


# ---------------------------------------------------------------------------
# 7. hyperlinks preserved + generated
# ---------------------------------------------------------------------------
def test_hyperlinks_generated_and_preserved(e2e) -> None:
    content = e2e.wb["CONTENT"]
    supp_row = _find_row(content, "SUPPDM", start=1)
    assert supp_row is not None
    hl = content.cell(row=supp_row, column=1).hyperlink
    assert hl is not None and "DM" in str(hl.location)

    # An existing domain hyperlink (untouched domain row) is preserved.
    ref_row = _find_row(content, UNTOUCHED_DOMAIN, start=1)
    assert ref_row is not None
    ref_link = e2e.tpl_wb["CONTENT"].cell(row=ref_row, column=1).hyperlink
    got_link = content.cell(row=ref_row, column=1).hyperlink
    assert ref_link is not None and got_link is not None
    assert got_link.target == ref_link.target
    assert got_link.location == ref_link.location
    assert got_link.tooltip == ref_link.tooltip
    assert got_link.display == ref_link.display


# ---------------------------------------------------------------------------
# 8. style preservation (untouched domain)
# ---------------------------------------------------------------------------
def test_untouched_domain_styles_preserved(e2e) -> None:
    assert UNTOUCHED_DOMAIN in e2e.tpl_wb.sheetnames
    ref = e2e.tpl_wb[UNTOUCHED_DOMAIN]
    got = e2e.wb[UNTOUCHED_DOMAIN]
    # Compare the complete immutable StyleArray plus the public components for
    # an untouched header row. This catches border/alignment/protection/number
    # format regressions that a bold/fill-only assertion would miss.
    for c in range(1, 11):
        ref_cell = ref.cell(row=13, column=c)
        got_cell = got.cell(row=13, column=c)
        assert got_cell._style == ref_cell._style
        assert got_cell.number_format == ref_cell.number_format
        assert str(got_cell.alignment) == str(ref_cell.alignment)
        assert str(got_cell.border) == str(ref_cell.border)
        assert str(got_cell.protection) == str(ref_cell.protection)


# ---------------------------------------------------------------------------
# 9. generated-cell highlight
# ---------------------------------------------------------------------------
def test_generated_cells_highlighted(e2e) -> None:
    dm = e2e.wb["DM"]
    qnam_row = _find_row(dm, "SITENAM")
    assert qnam_row is not None
    fill = dm.cell(row=qnam_row, column=1).fill
    assert fill.patternType == "solid" and str(fill.fgColor.rgb).endswith("FFFF00")


# ---------------------------------------------------------------------------
# 10. merged cells intact
# ---------------------------------------------------------------------------
def test_merged_cells_intact(e2e) -> None:
    dm = e2e.wb["DM"]
    merged = {str(m) for m in dm.merged_cells.ranges}
    assert "A1:H1" in merged  # header merge survived SUPP insertion

    # Untouched domain: merged ranges identical to the pristine template.
    assert UNTOUCHED_DOMAIN in e2e.tpl_wb.sheetnames
    ref = {str(m) for m in e2e.tpl_wb[UNTOUCHED_DOMAIN].merged_cells.ranges}
    got = {str(m) for m in e2e.wb[UNTOUCHED_DOMAIN].merged_cells.ranges}
    assert ref == got


# ---------------------------------------------------------------------------
# 11. repeated run — no duplicate SUPP / CODELIST rows
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("ig", ["IG3.2", "IG3.4"])
def test_repeated_run_no_duplicates(ig: str, tmp_path: Path) -> None:
    # Cover both template families: IG 3.2 (pre-populated CODELIST) and IG 3.4
    # (empty CODELIST) exercise the delete-existing-SUPP + CODELIST merge dedup
    # on re-run so a regression in either is caught by the gate. IG 3.2 keeps
    # its DEFAULT create_test_sheets=True so the conditional-mapping path (TEST
    # sheet CRF_* columns) is exercised too; IG 3.4 has no per-domain TEST
    # sheets, so that path stays disabled there.
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    create_test_sheets = CREATE_TEST_SHEETS[ig]
    logging.disable(logging.CRITICAL)
    try:
        # Run 1: real template -> out1
        run1 = _run_pipeline(tmp_path, ig, create_test_sheets=create_test_sheets)
        # Run 2: feed the generated workbook back in as the template -> out2
        als2 = tmp_path / "als2.xlsx"
        write_synthetic_als(als2)
        out2 = tmp_path / "out2.xlsx"
        mapper = SpecMapper(als_file=als2, template_file=run1.out, als_sheet="Sheet1", log_level="CRITICAL")
        mapper.process(output_file=out2, highlight=True, create_test_sheets=create_test_sheets)

        wb = load_workbook(out2)
        try:
            dm = wb["DM"]
            sitenam = [
                r for r in range(14, dm.max_row + 1) if str(dm.cell(row=r, column=1).value or "").upper() == "SITENAM"
            ]
            assert len(sitenam) == 1, "SUPP row must not duplicate on re-run"
            cl = wb["CODELIST"]
            qt = [
                r
                for r in range(3, cl.max_row + 1)
                if str(cl.cell(row=r, column=1).value or "").upper() == "EG"
                and str(cl.cell(row=r, column=2).value or "").upper() == "EGTESTCD"
                and str(cl.cell(row=r, column=6).value or "").upper() == "QT"
            ]
            assert len(qt) == 1, "CODELIST row must not duplicate on re-run"
            # The CONTENT sheet is a separate insert path (add_supp_to_content_sheet)
            # that must also be idempotent: exactly one SUPPDM row after re-run.
            content = wb["CONTENT"]
            supp_dm = [
                r
                for r in range(1, content.max_row + 1)
                if str(content.cell(row=r, column=1).value or "").strip().upper() == "SUPPDM"
            ]
            assert len(supp_dm) == 1, "CONTENT SUPPDM row must not duplicate on re-run"
            # Conditional-mapping path (IG 3.2): the CRF_* columns written into
            # a TEST sheet must be REUSED on re-run, never appended again.
            if create_test_sheets:
                crf_sheets = 0
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
                    if "CRF_TESTCD" in headers or "CRF_ORRES" in headers:
                        crf_sheets += 1
                        assert headers.count("CRF_TESTCD") <= 1, f"duplicate CRF_TESTCD columns in '{sheet}'"
                        assert headers.count("CRF_ORRES") <= 1, f"duplicate CRF_ORRES columns in '{sheet}'"
                assert crf_sheets >= 1, "expected the conditional-mapping path to have written CRF_* columns"
        finally:
            wb.close()
    finally:
        logging.disable(logging.NOTSET)


# ---------------------------------------------------------------------------
# 12. recoverable write failure (process level)
# ---------------------------------------------------------------------------
def test_recoverable_write_failure_still_saves(tmp_path: Path, monkeypatch) -> None:
    ig = "IG3.2"
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    try:
        src = TEMPLATES[ig]
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(src, tpl)
        als = tmp_path / "als.xlsx"
        write_synthetic_als(als)
        out = tmp_path / "out.xlsx"

        # Inject a *recognized* recoverable per-operation failure into a guarded
        # write step. It must be the dedicated RecoverableWriteError — a bare
        # ValueError is (intentionally) no longer treated as recoverable.
        def boom(self, *a, **k):
            raise RecoverableWriteError("injected recoverable failure")

        monkeypatch.setattr(ExcelWriter, "add_content_link_to_domain", boom)

        mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
        stats = mapper.process(output_file=out, highlight=True)

        actual = stats["actual"]
        assert actual["errors"] >= 1
        assert actual["written"] < actual["attempted"]
        # The workbook is still produced and openable for manual review.
        assert out.exists()
        wb = load_workbook(out)
        wb.close()
        # Error payload is structured + safe: only the exception *class* name,
        # never the raw message ("injected ...") or a traceback.
        errors = stats["write_result"]["errors"]
        assert errors and all(e["detail"] == "RecoverableWriteError" for e in errors)
        assert all("injected" not in str(e["detail"]) for e in errors)
    finally:
        logging.disable(logging.NOTSET)


def test_bare_valueerror_in_guarded_op_is_fatal(tmp_path: Path, monkeypatch) -> None:
    """A bare ValueError at the guard boundary must NOT be downgraded to a
    recoverable error — only the dedicated RecoverableWriteError is. This proves
    the classification is by recognized type, not by broad exception class."""
    ig = "IG3.2"
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    try:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(TEMPLATES[ig], tpl)
        als = tmp_path / "als.xlsx"
        write_synthetic_als(als)
        out = tmp_path / "out.xlsx"

        def boom(self, *a, **k):
            raise ValueError("looks recoverable but is not classified")

        monkeypatch.setattr(ExcelWriter, "add_content_link_to_domain", boom)

        mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
        with pytest.raises(ValueError):
            mapper.process(output_file=out, highlight=True)
    finally:
        logging.disable(logging.NOTSET)


def test_unknown_write_error_is_fatal(tmp_path: Path, monkeypatch) -> None:
    """An unknown (non-recoverable) exception in a guarded op must NOT be masked
    as completed_with_errors — it propagates so the job fails."""
    ig = "IG3.2"
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    try:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(TEMPLATES[ig], tpl)
        als = tmp_path / "als.xlsx"
        write_synthetic_als(als)
        out = tmp_path / "out.xlsx"

        def boom(self, *a, **k):
            raise RuntimeError("unexpected programming error")

        monkeypatch.setattr(ExcelWriter, "add_content_link_to_domain", boom)

        mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
        with pytest.raises(RuntimeError):
            mapper.process(output_file=out, highlight=True)
    finally:
        logging.disable(logging.NOTSET)


def test_guarded_no_op_is_skipped_not_written(tmp_path: Path, monkeypatch) -> None:
    """A guarded op that returns without mutating (count 0) is recorded as a
    skip, never as a phantom write, and drives completed_with_errors."""
    ig = "IG3.2"
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    try:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(TEMPLATES[ig], tpl)
        als = tmp_path / "als.xlsx"
        write_synthetic_als(als)
        out = tmp_path / "out.xlsx"

        # Exception-free call that performs no workbook mutation.
        monkeypatch.setattr(ExcelWriter, "add_content_link_to_domain", lambda self, *a, **k: 0)

        mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
        stats = mapper.process(output_file=out, highlight=True)

        actual = stats["actual"]
        assert actual["errors"] == 0
        assert actual["skipped"] >= 1
        assert actual["written"] < actual["attempted"]
        codes = {i["code"] for i in stats["write_result"]["warnings"]}
        assert "no_op" in codes
    finally:
        logging.disable(logging.NOTSET)


# ---------------------------------------------------------------------------
# 13. fatal save failure (process level)
# ---------------------------------------------------------------------------
def test_fatal_save_failure_propagates(tmp_path: Path, monkeypatch) -> None:
    ig = "IG3.2"
    assert TEMPLATES[ig].is_file(), f"required repo template missing: {TEMPLATES[ig]}"
    logging.disable(logging.CRITICAL)
    try:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(TEMPLATES[ig], tpl)
        als = tmp_path / "als.xlsx"
        write_synthetic_als(als)
        out = tmp_path / "out.xlsx"

        def boom(self, *a, **k):
            raise OSError("disk full")

        monkeypatch.setattr(ExcelWriter, "save", boom)

        mapper = SpecMapper(als_file=als, template_file=tpl, als_sheet="Sheet1", log_level="CRITICAL")
        with pytest.raises(Exception):
            mapper.process(output_file=out, highlight=True)
    finally:
        logging.disable(logging.NOTSET)
