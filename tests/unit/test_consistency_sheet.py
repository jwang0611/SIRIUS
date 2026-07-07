"""Unit tests for consistency sheet, critic flags, and tab highlighting."""

from __future__ import annotations

from openpyxl import Workbook

from src.processors.excel_styler import (
    CRITIC_TAB_ERROR,
    CRITIC_TAB_WARNING,
    add_critic_summary_comment,
    apply_critic_flags,
    highlight_issues_tab,
    write_consistency_sheet,
)


class TestWriteConsistencySheet:
    def test_write_consistency_sheet_returns_false_on_empty(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        result = write_consistency_sheet(wb, [])
        assert result is False
        assert "Consistency_Issues" not in wb.sheetnames

    def test_write_consistency_sheet_creates_5_column_sheet(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        issues = [
            {
                "severity": "error",
                "check": "domain_prefix",
                "description": "Variable prefix does not match domain",
                "affected_variables": [["DM", "BRTHDTC"]],
                "suggested_fix": "Correct prefix to DM",
            }
        ]
        result = write_consistency_sheet(wb, issues)
        assert result is True
        assert "Consistency_Issues" in wb.sheetnames
        ws = wb["Consistency_Issues"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Severity", "Check", "Description", "Affected_Variables", "Suggested_Fix"]
        assert ws["A2"].value == "error"
        assert ws["B2"].value == "domain_prefix"
        assert ws["D2"].value == "DM.BRTHDTC"


class TestApplyCriticFlags:
    def test_apply_critic_flags_marks_affected_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["metadata_table", "metadata_variable", "Critic_Flag"])
        ws.append(["DM", "BRTHDTC", ""])
        ws.append(["AE", "AETERM", ""])
        issues = [
            {
                "severity": "error",
                "affected_variables": [["DM", "BRTHDTC"]],
            }
        ]
        flagged = apply_critic_flags(ws, "C", issues, table_col_letter="A", variable_col_letter="B")
        assert flagged == 1
        assert ws["C2"].value == "error"
        assert ws["C2"].border.left.color.rgb[-6:] == "C00000"
        assert ws["C3"].value is None or ws["C3"].value == ""

    def test_apply_critic_flags_prefers_error_over_warning(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["metadata_table", "metadata_variable", "Critic_Flag"])
        ws.append(["DM", "BRTHDTC", ""])
        issues = [
            {
                "severity": "warning",
                "affected_variables": [["DM", "BRTHDTC"]],
            },
            {
                "severity": "error",
                "affected_variables": [["DM", "BRTHDTC"]],
            },
        ]
        flagged = apply_critic_flags(ws, "C", issues, table_col_letter="A", variable_col_letter="B")
        assert flagged == 1
        assert ws["C2"].value == "error"
        assert ws["C2"].border.left.color.rgb[-6:] == "C00000"


class TestAddCriticSummaryComment:
    def test_add_critic_summary_comment_writes_counts(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        issues = [
            {"severity": "error", "description": "E1"},
            {"severity": "error", "description": "E2"},
            {"severity": "warning", "description": "W1"},
        ]
        add_critic_summary_comment(ws, issues)
        comment = ws["A1"].comment
        assert comment is not None
        assert "2 errors / 1 warnings" in comment.text
        assert comment.author == "MappingCritic"


class TestHighlightIssuesTab:
    def test_highlight_issues_tab_sets_red_on_errors(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        issues = [{"severity": "error", "description": "E1"}]
        write_consistency_sheet(wb, issues)
        highlight_issues_tab(wb, issues)
        ws = wb["Consistency_Issues"]
        assert ws.sheet_properties.tabColor.rgb[-6:] == CRITIC_TAB_ERROR

    def test_highlight_issues_tab_sets_yellow_when_only_warnings(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        issues = [{"severity": "warning", "description": "W1"}]
        write_consistency_sheet(wb, issues)
        highlight_issues_tab(wb, issues)
        ws = wb["Consistency_Issues"]
        assert ws.sheet_properties.tabColor.rgb[-6:] == CRITIC_TAB_WARNING
