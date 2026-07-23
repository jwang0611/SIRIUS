"""Unit tests for the Spec Mapper structured write-result observability.

Covers:
  * the :mod:`write_result` data models (counts, invariants, serialization,
    safety of the serialized payload);
  * :meth:`ExcelWriter.update_cells` per-cell success / failure accounting;
  * :meth:`ExcelWriter.write_codelist_records` insert / merge / skip accounting
    against the **real** IG 3.2 template (copied into tmp_path).
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.spec_mapper import SpecMapper
from src.spec_mapper.core.excel_writer import ExcelWriter
from src.spec_mapper.helpers import insert_supp_rows, process_conditional_mappings
from src.spec_mapper.models import CellUpdate, CodelistRecord, ConditionalRecord, SUPPRecord
from src.spec_mapper.models.write_result import (
    STAGE_CELL_UPDATES,
    StageWriteResult,
    WriteIssue,
    WriteResult,
)

V32_TEMPLATE = Path("data/knowledge_base/template_spec/SDTM_template_IG3.2.xlsx")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
class TestWriteResultModel:
    def test_stage_invariant_attempted_equals_written_skipped_errors(self) -> None:
        stage = StageWriteResult(stage="s")
        stage.record_written()
        stage.record_written()
        stage.record_skipped(WriteIssue(code="sheet_not_found", stage="s", operation="op"))
        stage.record_error(WriteIssue(code="boom", stage="s", operation="op", detail="ValueError"))

        assert stage.attempted == 4
        assert stage.written == 2
        assert stage.skipped == 1
        assert len(stage.errors) == 1
        # invariant: attempted == written + skipped + len(errors)
        assert stage.attempted == stage.written + stage.skipped + len(stage.errors)

    def test_add_warning_does_not_change_counts(self) -> None:
        stage = StageWriteResult(stage="s")
        stage.record_written()
        stage.add_warning(WriteIssue(code="supp_label_too_long", stage="s", operation="op"))
        assert stage.attempted == 1
        assert stage.written == 1
        assert len(stage.warnings) == 1

    def test_aggregate_properties_and_write_problems(self) -> None:
        result = WriteResult()
        good = StageWriteResult(stage="a")
        good.record_written()
        good.record_written()
        bad = StageWriteResult(stage="b")
        bad.record_error(WriteIssue(code="boom", stage="b", operation="op"))
        result.merge_stage(good)
        result.merge_stage(bad)

        assert result.attempted == 3
        assert result.written == 2
        assert result.error_count == 1
        assert result.has_errors is True
        assert result.has_write_problems is True

    def test_clean_result_has_no_write_problems(self) -> None:
        result = WriteResult()
        stage = StageWriteResult(stage="a")
        stage.record_written()
        result.merge_stage(stage)
        assert result.has_write_problems is False
        assert result.summary() == {
            "attempted": 1,
            "written": 1,
            "skipped": 0,
            "warnings": 0,
            "errors": 0,
        }

    def test_skip_only_counts_as_write_problem(self) -> None:
        """A safe skip (written < attempted) must be surfaced as a problem."""
        result = WriteResult()
        stage = StageWriteResult(stage="a")
        stage.record_skipped(WriteIssue(code="domain_not_found", stage="a", operation="op"))
        result.merge_stage(stage)
        assert result.written < result.attempted
        assert result.has_write_problems is True

    def test_merge_stage_accumulates_same_stage_name(self) -> None:
        result = WriteResult()
        first = StageWriteResult(stage="cell_updates")
        first.record_written()
        second = StageWriteResult(stage="cell_updates")
        second.record_written()
        result.merge_stage(first)
        result.merge_stage(second)
        assert set(result.stages) == {"cell_updates"}
        assert result.stages["cell_updates"].written == 2

    def test_to_dict_is_serializable_and_safe(self) -> None:
        """Serialized issues expose only safe fields — never traceback / message text."""
        result = WriteResult()
        stage = StageWriteResult(stage="cell_updates")
        stage.record_error(
            WriteIssue(
                code="cell_write_failed",
                stage="cell_updates",
                operation="update_cell",
                sheet="DM",
                row=15,
                column=8,
                detail="ValueError",
            )
        )
        result.merge_stage(stage)

        payload = result.to_dict()
        import json

        json.dumps(payload)  # must be JSON-serializable

        assert payload["summary"]["errors"] == 1
        issue = payload["errors"][0]
        assert set(issue) == {"code", "stage", "operation", "sheet", "row", "column", "detail"}
        # detail is only the exception class name, not a message/traceback/path.
        assert issue["detail"] == "ValueError"


# ---------------------------------------------------------------------------
# ExcelWriter.update_cells
# ---------------------------------------------------------------------------
def _domain_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "DM"
    ws.cell(row=15, column=8, value="")
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return path


class TestUpdateCellsAccounting:
    def test_written_only_counts_successful_mutations(self, tmp_path: Path) -> None:
        writer = ExcelWriter(_domain_workbook(tmp_path))
        updates = [
            CellUpdate(sheet_name="DM", row=15, col=8, value="[RAW]SUBJECT.STUDYID"),
            # Targets a sheet that does not exist -> recoverable per-cell error.
            CellUpdate(sheet_name="NOPE", row=15, col=8, value="x"),
        ]
        result = writer.update_cells(updates)

        assert result.stage == STAGE_CELL_UPDATES
        assert result.attempted == 2
        assert result.written == 1
        assert len(result.errors) == 1
        err = result.errors[0]
        assert err.code == "cell_write_failed"
        assert err.sheet == "NOPE"
        # The good cell really was written.
        assert writer.workbook["DM"].cell(row=15, column=8).value == "[RAW]SUBJECT.STUDYID"

    def test_empty_updates_yield_empty_result(self, tmp_path: Path) -> None:
        writer = ExcelWriter(_domain_workbook(tmp_path))
        result = writer.update_cells([])
        assert result.attempted == 0
        assert result.written == 0


class TestGuardedMethodsReturnMutationCounts:
    """Guarded write methods must report *actual* mutation counts so a no-op is
    never miscounted as a real write."""

    def _domain(self, tmp_path: Path, *, with_content: bool) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "DM"
        ws.cell(row=14, column=1, value="STUDYID")
        ws.cell(row=15, column=1, value="DOMAIN")
        if with_content:
            ws.cell(row=20, column=1, value="CONTENT")
        path = tmp_path / "wb.xlsx"
        wb.save(path)
        return path

    def test_add_content_link_returns_zero_without_content_cell(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._domain(tmp_path, with_content=False))
        assert writer.add_content_link_to_domain("DM") == 0  # no-op

    def test_add_content_link_returns_one_with_content_cell(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._domain(tmp_path, with_content=True))
        assert writer.add_content_link_to_domain("DM") == 1  # real mutation

    def test_set_active_sheet_reports_missing_as_zero(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._domain(tmp_path, with_content=False))
        assert writer.set_active_sheet("NOPE") == 0
        assert writer.set_active_sheet("DM") == 1


# ---------------------------------------------------------------------------
# ExcelWriter.write_codelist_records (real template)
# ---------------------------------------------------------------------------
def _codelist_row_count(path: Path) -> int:
    wb = load_workbook(path)
    ws = wb["CODELIST"]
    count = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=1).value)
    wb.close()
    return count


@pytest.mark.skipif(not V32_TEMPLATE.exists(), reason="IG 3.2 template not present")
class TestCodelistAccounting:
    def test_insert_new_record_counts_as_written(self, tmp_path: Path) -> None:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(V32_TEMPLATE, tpl)
        before = _codelist_row_count(tpl)

        writer = ExcelWriter(tpl)
        rec = CodelistRecord(
            domain="ZZ",
            testcd_var="ZZTESTCD",
            testcd_value="ZZVAL",
            source_variable="源变量",
            metadata_variable="ZZRAW",
        )
        result = writer.write_codelist_records([rec])
        out = tmp_path / "out.xlsx"
        writer.save(out)
        writer.close()

        assert result.written == 1
        assert result.attempted == 1
        assert not result.errors
        # A brand-new row was inserted for the new key.
        assert _codelist_row_count(out) == before + 1

    def test_merge_into_existing_row_fills_blank_and_inserts_nothing(self, tmp_path: Path) -> None:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(V32_TEMPLATE, tpl)

        # Discover an existing CODELIST row (domain, ID, Term) whose EDC VAR (J)
        # is still blank so the merge path fills it in place.
        wb = load_workbook(tpl)
        ws = wb["CODELIST"]
        target = None
        for r in range(3, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            f = ws.cell(row=r, column=6).value
            j = ws.cell(row=r, column=10).value
            if a and b and f and not j:
                target = (str(a), str(b), str(f), r)
                break
        wb.close()
        assert target is not None, "expected at least one mergeable CODELIST row in the real template"
        domain, testcd_var, testcd_value, row_idx = target
        before = _codelist_row_count(tpl)

        writer = ExcelWriter(tpl)
        rec = CodelistRecord(
            domain=domain,
            testcd_var=testcd_var,
            testcd_value=testcd_value,
            source_variable="",  # leave I alone
            metadata_variable="MERGEDVAR",
        )
        result = writer.write_codelist_records([rec])
        out = tmp_path / "out.xlsx"
        writer.save(out)
        writer.close()

        assert result.written == 1
        assert not result.errors

        wb2 = load_workbook(out)
        ws2 = wb2["CODELIST"]
        assert str(ws2.cell(row=row_idx, column=10).value) == "MERGEDVAR"
        wb2.close()
        # Merge must not append a duplicate row.
        assert _codelist_row_count(out) == before

    def test_fully_satisfied_record_is_not_written(self, tmp_path: Path) -> None:
        """A record whose row already exists with I and J populated is a genuine
        no-op: it must be recorded as skipped (codelist_unchanged), never as a
        phantom write, and must not append a duplicate row."""
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(V32_TEMPLATE, tpl)

        writer = ExcelWriter(tpl)
        rec = CodelistRecord(
            domain="ZZ",
            testcd_var="ZZTESTCD",
            testcd_value="ZZVAL",
            source_variable="SRC",
            metadata_variable="ZZRAW",
        )
        # First write inserts the row and fills I/J -> a real mutation.
        first = writer.write_codelist_records([rec])
        assert first.written == 1

        # Re-writing the identical record now finds the row with I and J already
        # populated: zero cells change -> not written, recorded as unchanged.
        second = writer.write_codelist_records([rec])
        assert second.written == 0
        assert second.skipped == 1
        assert second.warnings and second.warnings[0].code == "codelist_unchanged"

        out = tmp_path / "out.xlsx"
        writer.save(out)
        writer.close()

        wb = load_workbook(out)
        ws = wb["CODELIST"]
        hits = [
            r
            for r in range(3, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").upper() == "ZZ"
            and str(ws.cell(row=r, column=2).value or "").upper() == "ZZTESTCD"
            and str(ws.cell(row=r, column=6).value or "").upper() == "ZZVAL"
        ]
        wb.close()
        assert len(hits) == 1  # the no-op second write appended nothing

    def test_missing_codelist_sheet_marks_records_skipped(self, tmp_path: Path) -> None:
        wb = Workbook()
        wb.active.title = "DM"
        path = tmp_path / "no_codelist.xlsx"
        wb.save(path)

        writer = ExcelWriter(path)
        rec = CodelistRecord(domain="EG", testcd_var="EGTESTCD", testcd_value="QT", source_variable="s")
        result = writer.write_codelist_records([rec])

        assert result.written == 0
        assert result.skipped == 1
        assert result.warnings[0].code == "sheet_not_found"


# ---------------------------------------------------------------------------
# Idempotency of the other CONTENT / domain insert paths (re-run audit)
# ---------------------------------------------------------------------------
class TestInsertPathsIdempotent:
    """The non-standard-domain and external-coding insert paths must not append
    duplicate rows when the pipeline is re-run on a previously-generated
    workbook (companion to the end-to-end SUPP/CODELIST re-run gate)."""

    def _content_workbook(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "CONTENT"
        ws.cell(row=1, column=1, value="CONTENT")
        ws.cell(row=2, column=1, value="AE")  # an existing 2-char domain row
        path = tmp_path / "content.xlsx"
        wb.save(path)
        return path

    def _domain_workbook(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "AE"
        ws.cell(row=13, column=1, value="Variable")  # header
        ws.cell(row=14, column=1, value="AETERM")
        ws.cell(row=14, column=9, value="SDTM")
        path = tmp_path / "dom.xlsx"
        wb.save(path)
        return path

    def test_nonstandard_domain_added_only_once(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._content_workbook(tmp_path))
        # First insert mutates; the second is a no-op (already present).
        assert writer.add_nonstandard_domain_to_content("XX", "CONTENT", reference_row=2) == 1
        assert writer.add_nonstandard_domain_to_content("XX", "CONTENT", reference_row=2) == 0

        ws = writer.workbook["CONTENT"]
        xx = [r for r in range(1, ws.max_row + 1) if str(ws.cell(row=r, column=1).value or "").strip().upper() == "XX"]
        writer.close()
        assert len(xx) == 1

    def test_external_coding_variables_added_only_once(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._domain_workbook(tmp_path))
        variables = [{"name": "AEDVER", "label": "MedDRA Version", "type": "Char", "source": "指定"}]
        # First call inserts one variable; the second finds it and inserts none.
        assert writer.add_external_coding_variables("AE", variables, variable_type="SUPP") == 1
        assert writer.add_external_coding_variables("AE", variables, variable_type="SUPP") == 0

        ws = writer.workbook["AE"]
        hits = [
            r
            for r in range(14, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").strip().upper() == "AEDVER"
        ]
        writer.close()
        assert len(hits) == 1


# ---------------------------------------------------------------------------
# Per-item atomicity: pre-validation must run BEFORE any destructive mutation
# ---------------------------------------------------------------------------
def _dm_supp_state(ws) -> tuple[int, list[str]]:
    """(count, names) of SUPP-type rows in a domain sheet (column I == SUPP)."""
    rows = [r for r in range(14, ws.max_row + 1) if str(ws.cell(row=r, column=9).value or "").strip() == "SUPP"]
    return len(rows), [str(ws.cell(row=r, column=1).value or "") for r in rows]


@pytest.mark.skipif(not V32_TEMPLATE.exists(), reason="IG 3.2 template not present")
class TestSuppInsertAtomicity:
    """A SUPP record whose values openpyxl would reject must be caught by
    pre-validation BEFORE the existing SUPP block is deleted or any row is
    inserted — never leaving a half-written row or destroying old data."""

    def _writer(self, tmp_path: Path) -> ExcelWriter:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(V32_TEMPLATE, tpl)
        return ExcelWriter(tpl)

    def _record(self, name: str, label: str) -> SUPPRecord:
        return SUPPRecord(
            sheet_name="DM",
            variable_name=name,
            variable_label=label,
            transformation_def="[RAW]SUBJECT.X",
            insert_row=0,
        )

    def test_all_invalid_leaves_domain_untouched(self, tmp_path: Path) -> None:
        writer = self._writer(tmp_path)
        before_count, before_names = _dm_supp_state(writer.workbook["DM"])
        max_row_before = writer.workbook["DM"].max_row

        bad = self._record("BADVAR", "bad\x01label")  # illegal control character
        result = insert_supp_rows(writer, [bad], highlight=True)

        assert result.written == 0
        assert result.errors and result.errors[0].code == "illegal_characters"
        # The template's existing SUPP block was NOT deleted and no row was
        # inserted: the domain sheet is byte-for-byte untouched in shape.
        after_count, after_names = _dm_supp_state(writer.workbook["DM"])
        assert (after_count, after_names) == (before_count, before_names)
        assert writer.workbook["DM"].max_row == max_row_before
        writer.close()

    def test_mixed_records_write_valid_fully_and_skip_invalid(self, tmp_path: Path) -> None:
        writer = self._writer(tmp_path)
        good = self._record("GOODVAR", "正常标签")
        bad = self._record("BADVAR", "bad\x01label")

        result = insert_supp_rows(writer, [good, bad], highlight=True)

        assert result.written == 1
        assert len(result.errors) == 1 and result.errors[0].code == "illegal_characters"

        ws = writer.workbook["DM"]
        _, names = _dm_supp_state(ws)
        assert "GOODVAR" in names and "BADVAR" not in names
        # The valid row is COMPLETE (no half-product): name, label, type, I set.
        row = next(r for r in range(14, ws.max_row + 1) if ws.cell(row=r, column=1).value == "GOODVAR")
        assert ws.cell(row=row, column=2).value == "正常标签"
        assert ws.cell(row=row, column=9).value == "SUPP"
        writer.close()


class TestUpdateCellsPreValidation:
    def test_illegal_character_value_is_error_and_cell_untouched(self, tmp_path: Path) -> None:
        writer = ExcelWriter(_domain_workbook(tmp_path))
        writer.workbook["DM"].cell(row=15, column=8, value="old")
        updates = [CellUpdate(sheet_name="DM", row=15, col=8, value="bad\x02value")]

        result = writer.update_cells(updates)

        assert result.written == 0
        assert result.errors and result.errors[0].code == "illegal_characters"
        # Pre-validation means the cell was never touched at all.
        assert writer.workbook["DM"].cell(row=15, column=8).value == "old"


@pytest.mark.skipif(not V32_TEMPLATE.exists(), reason="IG 3.2 template not present")
class TestCodelistPreValidation:
    def test_illegal_character_record_is_error_and_nothing_inserted(self, tmp_path: Path) -> None:
        tpl = tmp_path / "tpl.xlsx"
        shutil.copy2(V32_TEMPLATE, tpl)
        before = _codelist_row_count(tpl)

        writer = ExcelWriter(tpl)
        rec = CodelistRecord(
            domain="ZZ",
            testcd_var="ZZTESTCD",
            testcd_value="bad\x03value",
            source_variable="s",
        )
        result = writer.write_codelist_records([rec])

        assert result.written == 0
        assert result.errors and result.errors[0].code == "illegal_characters"

        out = tmp_path / "out.xlsx"
        writer.save(out)
        writer.close()
        assert _codelist_row_count(out) == before  # no row inserted


# ---------------------------------------------------------------------------
# Conditional mapping idempotency (column reuse on repeated runs)
# ---------------------------------------------------------------------------
class TestConditionalMappingIdempotent:
    def _test_sheet_workbook(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "EGTEST"
        ws.cell(row=1, column=1, value="Test")
        ws.cell(row=2, column=1, value="QT")
        path = tmp_path / "egtest.xlsx"
        wb.save(path)
        return path

    def test_repeated_run_reuses_columns_and_clears_stale_rows(self, tmp_path: Path) -> None:
        writer = ExcelWriter(self._test_sheet_workbook(tmp_path))
        rec2rows = ConditionalRecord(
            sheet_name="EGTEST",
            column_names=["CRF_TESTCD", "CRF_ORRES"],
            mappings=[("QT", "[RAW]EG.QTCF"), ("HR", "[RAW]EG.HR")],
        )
        rec1row = ConditionalRecord(
            sheet_name="EGTEST",
            column_names=["CRF_TESTCD", "CRF_ORRES"],
            mappings=[("QT", "[RAW]EG.QTCF")],
        )

        assert process_conditional_mappings(writer, [rec2rows]).written == 1
        # Second run (fewer rows) must reuse the same columns, not append.
        assert process_conditional_mappings(writer, [rec1row]).written == 1

        ws = writer.workbook["EGTEST"]
        headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
        assert headers.count("CRF_TESTCD") == 1
        assert headers.count("CRF_ORRES") == 1
        col = headers.index("CRF_TESTCD") + 1
        # Fresh data written, stale second row from run 1 cleared.
        assert ws.cell(row=2, column=col).value == "QT"
        assert ws.cell(row=3, column=col).value is None
        writer.close()


# ---------------------------------------------------------------------------
# _guard records full batch mutation counts
# ---------------------------------------------------------------------------
@pytest.mark.skipif(not V32_TEMPLATE.exists(), reason="IG 3.2 template not present")
class TestGuardBatchCounting:
    def _mapper(self, tmp_path: Path) -> SpecMapper:
        als = tmp_path / "als.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(als)
        return SpecMapper(als_file=als, template_file=V32_TEMPLATE, als_sheet="Sheet1", log_level="CRITICAL")

    def test_batch_return_count_is_recorded_verbatim(self, tmp_path: Path) -> None:
        """A guarded method returning N > 1 contributes N to written/attempted,
        so actual.written reflects real mutations, not call sites."""
        mapper = self._mapper(tmp_path)
        result = WriteResult()
        ret = mapper._guard(result, "styles", "batch_op", lambda: 47)
        assert ret == 47
        stage = result.stages["styles"]
        assert stage.written == 47
        assert stage.attempted == 47

    def test_single_and_zero_returns_still_correct(self, tmp_path: Path) -> None:
        mapper = self._mapper(tmp_path)
        result = WriteResult()
        mapper._guard(result, "styles", "one_op", lambda: 1)
        mapper._guard(result, "styles", "no_op_call", lambda: 0)
        stage = result.stages["styles"]
        assert stage.written == 1
        assert stage.skipped == 1
        assert stage.attempted == 2
