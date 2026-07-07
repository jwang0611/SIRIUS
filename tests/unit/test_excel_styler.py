"""Unit tests for src.processors.excel_styler."""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.processors.excel_styler import (
    SCORE_GREEN,
    SCORE_RED,
    SCORE_YELLOW,
    SOURCE_CORRECTION,
    SOURCE_KB,
    SOURCE_LLM,
    SOURCE_PROJECT,
    SOURCE_RAG,
    apply_score_fill,
    apply_source_fill,
    style_results_workbook,
    write_legend_block,
)


def _assert_fill_color(cell, expected_hex: str) -> None:
    assert cell.fill.start_color.rgb[-6:] == expected_hex


class TestApplyScoreFill:
    def test_apply_score_fill_tiers(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Score"])
        ws.append([0.95])
        ws.append([0.85])
        ws.append([0.5])
        ws.append(["bad"])

        count = apply_score_fill(ws, "A", header_row=1)
        assert count == 3

        _assert_fill_color(ws["A2"], SCORE_GREEN)
        _assert_fill_color(ws["A3"], SCORE_YELLOW)
        _assert_fill_color(ws["A4"], SCORE_RED)
        # non-numeric ignored
        assert ws["A5"].fill.start_color.rgb in ("00000000", "FFFFFFFF") or ws["A5"].fill.fill_type is None


class TestApplySourceFill:
    def test_apply_source_fill_handles_all_sources(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Source"])
        ws.append(["KB"])
        ws.append(["eCRF_direct_match"])
        ws.append(["RAG"])
        ws.append(["LLM"])
        ws.append(["correction"])
        ws.append(["project:foo"])

        count = apply_source_fill(ws, "A", header_row=1)
        assert count == 6

        _assert_fill_color(ws["A2"], SOURCE_KB)
        _assert_fill_color(ws["A3"], SOURCE_KB)
        _assert_fill_color(ws["A4"], SOURCE_RAG)
        _assert_fill_color(ws["A5"], SOURCE_LLM)
        _assert_fill_color(ws["A6"], SOURCE_CORRECTION)
        _assert_fill_color(ws["A7"], SOURCE_PROJECT)

    def test_apply_source_fill_ignores_unknown_source(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Source"])
        ws.append(["UNKNOWN"])
        ws.append([None])
        ws.append([""])

        count = apply_source_fill(ws, "A", header_row=1)
        assert count == 0

        # No fill applied
        for row in (2, 3, 4):
            cell = ws[f"A{row}"]
            assert cell.fill.fill_type is None or cell.fill.start_color.rgb in ("00000000", "FFFFFFFF")


class TestWriteLegendBlock:
    def test_write_legend_block_produces_legend_sheet(self) -> None:
        wb = Workbook()
        # remove default sheet so we can check fresh
        wb.remove(wb.active)

        rows_written = write_legend_block(wb, sheet_name="Legend")
        assert rows_written > 0
        assert "Legend" in wb.sheetnames

        ws = wb["Legend"]
        headers = [cell.value for cell in ws[1]]
        assert "Tier" in headers
        assert "Color" in headers

        # Second header row (row 6) should be Source / Color
        row6_headers = [cell.value for cell in ws[6]]
        assert "Source" in row6_headers
        assert "Color" in row6_headers


class TestStyleResultsWorkbook:
    def test_style_results_workbook_round_trip(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["SDTM_Domain", "Score", "Source"])
        ws.append(["DM", 0.95, "KB"])
        ws.append(["VS", 0.5, "LLM"])
        ws.append(["AE", 0.8, "project:x"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            wb.save(path)

            style_results_workbook(path)

            wb2 = load_workbook(path)
            ws2 = wb2["Results"]
            _assert_fill_color(ws2["B2"], SCORE_GREEN)
            _assert_fill_color(ws2["B3"], SCORE_RED)
            _assert_fill_color(ws2["B4"], SCORE_YELLOW)
            _assert_fill_color(ws2["C2"], SOURCE_KB)
            _assert_fill_color(ws2["C3"], SOURCE_LLM)
            _assert_fill_color(ws2["C4"], SOURCE_PROJECT)

    def test_style_results_workbook_is_idempotent(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Score", "Source"])
        ws.append([0.95, "KB"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            wb.save(path)

            style_results_workbook(path)
            wb2 = load_workbook(path)
            assert wb2.sheetnames.count("Legend") == 1

            style_results_workbook(path)
            wb3 = load_workbook(path)
            assert wb3.sheetnames.count("Legend") == 1
