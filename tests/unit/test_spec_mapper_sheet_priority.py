"""Lock the documented ALS sheet-selection priority.

Two entry points with *different* effective precedence are covered here; the
per-entry-point table lives in ``src/spec_mapper/README.md``（「ALS sheet 解析优先级」）:

* ``SpecMapper(...)``           显式参数 > ``ALS_DEFAULT_SHEET`` > 配置 > ``"Sheet1"``
* ``scripts/generate_full_spec`` ``--als-sheet`` > ``ALS_DEFAULT_SHEET`` > ``"Sheet1"``
  （配置层不可达 —— 该 CLI 在构造 SpecMapper 前就把 env/默认值展开成显式参数）
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any, ClassVar

import pytest
from openpyxl import Workbook

from src.spec_mapper import SpecMapper


@pytest.fixture
def workbook_paths(tmp_path: Path) -> tuple[Path, Path]:
    als_path = tmp_path / "als.xlsx"
    template_path = tmp_path / "template.xlsx"

    als_workbook = Workbook()
    als_workbook.active.title = "Sheet1"
    als_workbook.create_sheet("eCRF")
    als_workbook.save(als_path)

    template_workbook = Workbook()
    template_workbook.active.title = "CONTENT"
    template_workbook["CONTENT"]["B4"] = "3.2"
    template_workbook.save(template_path)

    return als_path, template_path


def _config_file(tmp_path: Path, sheet_name: str) -> Path:
    config_path = tmp_path / f"config-{sheet_name}.yaml"
    config_path.write_text(f'als_defaults:\n  sheet_name: "{sheet_name}"\n', encoding="utf-8")
    return config_path


def test_explicit_sheet_overrides_environment_and_config(
    workbook_paths: tuple[Path, Path], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    als_path, template_path = workbook_paths
    monkeypatch.setenv("ALS_DEFAULT_SHEET", "EnvironmentSheet")

    mapper = SpecMapper(
        als_file=als_path,
        template_file=template_path,
        als_sheet="ExplicitSheet",
        config_file=_config_file(tmp_path, "ConfiguredSheet"),
    )

    assert mapper.als_sheet == "ExplicitSheet"


def test_environment_sheet_overrides_config(
    workbook_paths: tuple[Path, Path], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    als_path, template_path = workbook_paths
    monkeypatch.setenv("ALS_DEFAULT_SHEET", "EnvironmentSheet")

    mapper = SpecMapper(
        als_file=als_path,
        template_file=template_path,
        config_file=_config_file(tmp_path, "ConfiguredSheet"),
    )

    assert mapper.als_sheet == "EnvironmentSheet"


def test_config_sheet_is_used_without_argument_or_environment(
    workbook_paths: tuple[Path, Path], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    als_path, template_path = workbook_paths
    monkeypatch.delenv("ALS_DEFAULT_SHEET", raising=False)

    mapper = SpecMapper(
        als_file=als_path,
        template_file=template_path,
        config_file=_config_file(tmp_path, "ConfiguredSheet"),
    )

    assert mapper.als_sheet == "ConfiguredSheet"


def test_packaged_config_defaults_to_sheet1(workbook_paths: tuple[Path, Path], monkeypatch: pytest.MonkeyPatch) -> None:
    als_path, template_path = workbook_paths
    monkeypatch.delenv("ALS_DEFAULT_SHEET", raising=False)

    mapper = SpecMapper(als_file=als_path, template_file=template_path)

    assert mapper.als_sheet == "Sheet1"


# ---------------------------------------------------------------------------
# CLI layer: scripts/generate_full_spec.py
#
# NOTE: ``scripts.generate_full_spec.map_als_to_spec`` is a DIFFERENT function
# from ``src.spec_mapper.map_als_to_spec``. The script-level one resolves the
# env/default *before* constructing SpecMapper and hands the result over as an
# explicit argument, which short-circuits SpecMapper's own config lookup.
# ---------------------------------------------------------------------------


class _RecordingSpecMapper(SpecMapper):
    """Real ``SpecMapper.__init__`` (so resolution logic is exercised) with the
    workbook processing replaced by a stub — no Excel writes, no LLM, no network.

    ``config_override`` points the config layer at a file whose
    ``als_defaults.sheet_name`` is deliberately distinctive, so that any test
    reaching that layer would be visible in ``als_sheet``.
    """

    instances: ClassVar[list[_RecordingSpecMapper]] = []
    config_override: ClassVar[Path | None] = None

    @staticmethod
    def _select_config_path(template_file: Path) -> Path:
        override = _RecordingSpecMapper.config_override
        if override is not None:
            return override
        return SpecMapper._select_config_path(template_file)

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        _RecordingSpecMapper.instances.append(self)

    def process(self, *_args: Any, **_kwargs: Any) -> dict[str, int]:
        # Keys the CLI logs right after process() returns; a missing key would
        # raise KeyError before the assertion under test is reached.
        return {
            "als_records": 0,
            "template_records": 0,
            "updates": 0,
            "supp_records": 0,
            "conditional_records": 0,
        }


@pytest.fixture
def cli_recorder(
    workbook_paths: tuple[Path, Path], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> Iterator[tuple[Any, dict[str, Path]]]:
    """Patch ``scripts.generate_full_spec.SpecMapper`` with the recorder.

    The name must be patched in the *script* module, not in ``src.spec_mapper``,
    or the CLI would keep using the real class.
    """
    import scripts.generate_full_spec as cli

    als_path, template_path = workbook_paths

    _RecordingSpecMapper.instances = []
    _RecordingSpecMapper.config_override = _config_file(tmp_path, "ConfiguredSheet")
    monkeypatch.setattr(cli, "SpecMapper", _RecordingSpecMapper)
    # A developer .env must not decide what this test observes.
    monkeypatch.setattr(cli, "load_dotenv", lambda *a, **k: False)

    paths = {"als": als_path, "template": template_path, "output": tmp_path / "spec_out.xlsx"}
    yield cli, paths

    _RecordingSpecMapper.instances = []
    _RecordingSpecMapper.config_override = None


def _run_cli(cli: Any, paths: dict[str, Path], als_sheet: str | None) -> str | None:
    cli.map_als_to_spec(
        als_file=str(paths["als"]),
        template_file=str(paths["template"]),
        output_file=str(paths["output"]),
        als_sheet=als_sheet,
        highlight=False,
        dry_run=True,
    )
    assert len(_RecordingSpecMapper.instances) == 1
    return _RecordingSpecMapper.instances[0].als_sheet


def test_cli_explicit_als_sheet_wins_over_environment(
    cli_recorder: tuple[Any, dict[str, Path]], monkeypatch: pytest.MonkeyPatch
) -> None:
    cli, paths = cli_recorder
    monkeypatch.setenv("ALS_DEFAULT_SHEET", "EnvironmentSheet")

    assert _run_cli(cli, paths, als_sheet="ExplicitSheet") == "ExplicitSheet"


def test_cli_environment_used_when_als_sheet_omitted(
    cli_recorder: tuple[Any, dict[str, Path]], monkeypatch: pytest.MonkeyPatch
) -> None:
    cli, paths = cli_recorder
    monkeypatch.setenv("ALS_DEFAULT_SHEET", "EnvironmentSheet")

    assert _run_cli(cli, paths, als_sheet=None) == "EnvironmentSheet"


def test_cli_falls_back_to_sheet1_and_never_defers_to_config(
    cli_recorder: tuple[Any, dict[str, Path]], monkeypatch: pytest.MonkeyPatch
) -> None:
    """The documented CLI precedence is ``--als-sheet > ALS_DEFAULT_SHEET > "Sheet1"``.

    The config layer is structurally unreachable: the CLI collapses env+default
    into an explicit argument, which takes the ``als_sheet is not None`` branch
    in ``SpecMapper.__init__``. ``config_override`` would yield ``ConfiguredSheet``
    if that branch were ever skipped, so this asserts unreachability rather than
    a coincidental equality.
    """
    cli, paths = cli_recorder
    monkeypatch.delenv("ALS_DEFAULT_SHEET", raising=False)

    resolved = _run_cli(cli, paths, als_sheet=None)

    assert resolved != "ConfiguredSheet"
    assert resolved == "Sheet1"


def test_recorder_config_override_would_be_visible_without_the_cli(
    workbook_paths: tuple[Path, Path], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Control for the test above: constructing the same recorder directly with
    ``als_sheet=None`` DOES reach the config layer and yields ``ConfiguredSheet``.

    Without this control, ``!= "ConfiguredSheet"`` above could pass simply
    because the override never worked.
    """
    als_path, template_path = workbook_paths
    monkeypatch.delenv("ALS_DEFAULT_SHEET", raising=False)
    _RecordingSpecMapper.instances = []
    _RecordingSpecMapper.config_override = _config_file(tmp_path, "ConfiguredSheet")

    mapper = _RecordingSpecMapper(als_file=als_path, template_file=template_path)

    assert mapper.als_sheet == "ConfiguredSheet"
