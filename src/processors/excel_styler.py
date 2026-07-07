"""Excel styling utilities for SDTM result workbooks.

Applies conditional color fills to Score and Source columns, and writes a Legend sheet.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

SCORE_GREEN = "C6EFCE"
SCORE_YELLOW = "FFEB9C"
SCORE_RED = "FFC7CE"

SOURCE_KB = "BDD7EE"
SOURCE_RAG = "D9D2E9"
SOURCE_LLM = "FCE4D6"
SOURCE_CORRECTION = "C6EFCE"
SOURCE_PROJECT = "A2C4C9"

SOURCE_COLOR_MAP: dict[str, str] = {
    "KB": SOURCE_KB,
    "eCRF_direct_match": SOURCE_KB,
    "RAG": SOURCE_RAG,
    "LLM": SOURCE_LLM,
    "correction": SOURCE_CORRECTION,
}

DIFF_MATCH = "C6EFCE"
DIFF_VAR_DIFF = "FFEB9C"
DIFF_DOMAIN_DIFF = "FFC7CE"
DIFF_AI_ONLY = "D9D9D9"
DIFF_REF_ONLY = "D9D2E9"

DIFF_COLOR_MAP: dict[str, str] = {
    "match": DIFF_MATCH,
    "var_diff": DIFF_VAR_DIFF,
    "domain_diff": DIFF_DOMAIN_DIFF,
    "ai_only": DIFF_AI_ONLY,
    "ref_only": DIFF_REF_ONLY,
}

CRITIC_ERROR_BORDER = "C00000"
CRITIC_WARNING_BORDER = "E6B800"
CRITIC_TAB_ERROR = "FFC7CE"
CRITIC_TAB_WARNING = "FFEB9C"

# IG 3.4 check column colours
IG34_PASS = "C6EFCE"  # green — same tier as Score high
IG34_FAIL = "FFC7CE"  # red — same as Score low; cell text is bolded
IG34_SKIP = "D9D9D9"  # light gray — not applicable

SEVERITY_PRIORITY = {"error": 2, "warning": 1, "info": 0}

_SKIP_SHEETS = {"Legend", "Consistency_Issues"}


def _format_affected_variables(affected):
    if not affected:
        return ""
    parts = []
    for item in affected:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            parts.append(f"{item[0]}.{item[1]}")
        elif isinstance(item, dict):
            t = item.get("metadata_table") or item.get("table") or ""
            v = item.get("metadata_variable") or item.get("variable") or ""
            parts.append(f"{t}.{v}")
        else:
            parts.append(str(item))
    return ", ".join(parts)


def write_consistency_sheet(wb, consistency_issues: list[dict], sheet_name: str = "Consistency_Issues") -> bool:
    if not consistency_issues:
        return False
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Severity", "Check", "Description", "Affected_Variables", "Suggested_Fix"])
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for issue in consistency_issues:
        ws.append(
            [
                str(issue.get("severity", "")),
                str(issue.get("check", "")),
                str(issue.get("description", "")),
                _format_affected_variables(issue.get("affected_variables")),
                str(issue.get("suggested_fix", "")),
            ]
        )
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    return True


def _affected_pairs(issue: dict) -> list[tuple[str, str]]:
    out = []
    for item in issue.get("affected_variables") or []:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            out.append((str(item[0]), str(item[1])))
        elif isinstance(item, dict):
            out.append(
                (
                    str(item.get("metadata_table") or item.get("table") or ""),
                    str(item.get("metadata_variable") or item.get("variable") or ""),
                )
            )
    return out


def apply_critic_flags(
    ws,
    critic_col_letter: str,
    issues: list[dict],
    header_row: int = 1,
    table_col_letter: str = "A",
    variable_col_letter: str = "B",
) -> int:
    pair_severity: dict[tuple[str, str], str] = {}
    for issue in issues:
        severity = str(issue.get("severity", "")).lower()
        for pair in _affected_pairs(issue):
            current = pair_severity.get(pair)
            if current is None or SEVERITY_PRIORITY.get(severity, 0) > SEVERITY_PRIORITY.get(current, 0):
                pair_severity[pair] = severity

    flagged = 0
    for row in range(header_row + 1, ws.max_row + 1):
        t = ws[f"{table_col_letter}{row}"].value
        v = ws[f"{variable_col_letter}{row}"].value
        if t is None or v is None:
            continue
        key = (str(t), str(v))
        matched_severity: str | None = pair_severity.get(key)
        if matched_severity is None:
            continue
        cell = ws[f"{critic_col_letter}{row}"]
        cell.value = matched_severity
        border_color = CRITIC_ERROR_BORDER if matched_severity == "error" else CRITIC_WARNING_BORDER
        cell.border = Border(left=Side(style="thick", color=border_color))
        flagged += 1
    return flagged


def add_critic_summary_comment(ws, issues: list[dict]) -> None:
    if not issues:
        return
    error_count = sum(1 for i in issues if str(i.get("severity", "")).lower() == "error")
    warning_count = sum(1 for i in issues if str(i.get("severity", "")).lower() == "warning")
    text = f"MappingCritic detected {error_count} errors / {warning_count} warnings. See Consistency_Issues sheet."
    ws["A1"].comment = Comment(text, "MappingCritic")


def highlight_issues_tab(wb, issues: list[dict]) -> None:
    if "Consistency_Issues" not in wb.sheetnames or not issues:
        return
    has_error = any(str(i.get("severity", "")).lower() == "error" for i in issues)
    has_warning = any(str(i.get("severity", "")).lower() == "warning" for i in issues)
    ws = wb["Consistency_Issues"]
    if has_error:
        ws.sheet_properties.tabColor = CRITIC_TAB_ERROR
    elif has_warning:
        ws.sheet_properties.tabColor = CRITIC_TAB_WARNING


def _find_column_letter(ws: Worksheet, header_text: str, header_row: int = 1) -> str | None:
    """Return the column letter for *header_text* (case-insensitive) or None."""
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip().lower() == header_text.lower():
            return get_column_letter(cell.column)
    return None


def apply_score_fill(ws: Worksheet, score_col_letter: str, header_row: int = 1) -> int:
    """Color each cell in *score_col_letter* based on its numeric value.

    Returns the number of styled cells.
    """
    green_fill = PatternFill(start_color=SCORE_GREEN, end_color=SCORE_GREEN, fill_type="solid")
    yellow_fill = PatternFill(start_color=SCORE_YELLOW, end_color=SCORE_YELLOW, fill_type="solid")
    red_fill = PatternFill(start_color=SCORE_RED, end_color=SCORE_RED, fill_type="solid")

    styled = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws[f"{score_col_letter}{row}"]
        try:
            score = float(cell.value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            continue
        if score >= 0.9:
            cell.fill = green_fill
        elif score >= 0.7:
            cell.fill = yellow_fill
        else:
            cell.fill = red_fill
        styled += 1
    return styled


def apply_source_fill(ws: Worksheet, source_col_letter: str, header_row: int = 1) -> int:
    """Color each cell in *source_col_letter* based on its source badge.

    Returns the number of styled cells.
    """
    fills: dict[str, PatternFill] = {
        color: PatternFill(start_color=color, end_color=color, fill_type="solid")
        for color in {SOURCE_KB, SOURCE_RAG, SOURCE_LLM, SOURCE_CORRECTION, SOURCE_PROJECT}
    }

    styled = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws[f"{source_col_letter}{row}"]
        raw = str(cell.value) if cell.value is not None else ""
        raw_stripped = raw.strip()

        color: str | None = None
        if raw_stripped in SOURCE_COLOR_MAP:
            color = SOURCE_COLOR_MAP[raw_stripped]
        elif raw_stripped.startswith("project:"):
            color = SOURCE_PROJECT

        if color is not None:
            cell.fill = fills[color]
            styled += 1
    return styled


def apply_ig34_check_fill(ws: Worksheet, ig34_col_letter: str, header_row: int = 1) -> int:
    """Color-code IG34_Check column cells (Pass / Fail / Skip / empty).

    - Pass  -> green fill
    - Fail  -> red fill + bold font (most visible)
    - Skip  -> gray fill
    - empty -> no fill (domain not in IG 3.4, cannot validate)

    Returns the number of styled cells.
    """
    pass_fill = PatternFill(start_color=IG34_PASS, end_color=IG34_PASS, fill_type="solid")
    fail_fill = PatternFill(start_color=IG34_FAIL, end_color=IG34_FAIL, fill_type="solid")
    skip_fill = PatternFill(start_color=IG34_SKIP, end_color=IG34_SKIP, fill_type="solid")
    bold_font = Font(bold=True)

    styled = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws[f"{ig34_col_letter}{row}"]
        value = str(cell.value).strip() if cell.value is not None else ""
        if value == "Pass":
            cell.fill = pass_fill
            styled += 1
        elif value == "Fail":
            cell.fill = fail_fill
            cell.font = bold_font
            styled += 1
        elif value == "Skip":
            cell.fill = skip_fill
            styled += 1
    return styled


def apply_diff_fill(ws: Worksheet, diff_col_letter: str, header_row: int = 1) -> int:
    """Color-code Diff_Status values below header_row. Returns count filled."""
    count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws[f"{diff_col_letter}{row}"]
        raw = cell.value
        if raw is None:
            continue
        color = DIFF_COLOR_MAP.get(str(raw).strip())
        if color is None:
            continue
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        count += 1
    return count


def write_legend_block(wb: Workbook, sheet_name: str = "Legend") -> int:
    """Create or overwrite a Legend sheet with Score tier and Source badge color keys.

    Returns the number of rows written.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Score tiers
    ws.append(["Tier", "Threshold", "Color"])
    ws.append(["High", ">= 0.9", SCORE_GREEN])
    ws.append(["Medium", "0.7 <= score < 0.9", SCORE_YELLOW])
    ws.append(["Low", "< 0.7", SCORE_RED])
    ws.append([])

    # Source badges
    ws.append(["Source", "Color"])
    ws.append(["KB / eCRF_direct_match", SOURCE_KB])
    ws.append(["RAG", SOURCE_RAG])
    ws.append(["LLM", SOURCE_LLM])
    ws.append(["correction", SOURCE_CORRECTION])
    ws.append(["project:*", SOURCE_PROJECT])
    ws.append([])

    # IG 3.4 check
    ws.append(["IG34_Check", "Meaning", "Color"])
    ws.append(["Pass", "Variable is in the SDTM IG 3.4 standard list for the given domain", IG34_PASS])
    ws.append(["Fail", "Variable is NOT in the IG 3.4 standard list — requires manual review", IG34_FAIL])
    ws.append(["Skip", "Not applicable: NOT_SUBMITTED, SUPP-type, or no extractable variable token", IG34_SKIP])
    ws.append(["(empty)", "Domain is not covered by IG 3.4 data (e.g. APMH, XU) — cannot validate", ""])

    return ws.max_row


def style_results_workbook(xlsx_path: Path, consistency_issues: list[dict] | None = None) -> None:
    """Load *xlsx_path*, apply score/source fills to every data sheet, add Legend, and save.

    Calling this function twice on the same file is idempotent.
    """
    wb = load_workbook(xlsx_path)

    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        score_col = _find_column_letter(ws, "Score")
        if score_col:
            apply_score_fill(ws, score_col)
        source_col = _find_column_letter(ws, "Source")
        if source_col:
            apply_source_fill(ws, source_col)
        diff_col = _find_column_letter(ws, "Diff_Status")
        if diff_col:
            apply_diff_fill(ws, diff_col)
        ig34_col = _find_column_letter(ws, "IG34_Check")
        if ig34_col:
            apply_ig34_check_fill(ws, ig34_col)
        critic_col = _find_column_letter(ws, "Critic_Flag")
        if critic_col and consistency_issues:
            table_col = _find_column_letter(ws, "metadata_table")
            variable_col = _find_column_letter(ws, "metadata_variable")
            if table_col and variable_col:
                apply_critic_flags(
                    ws,
                    critic_col_letter=critic_col,
                    issues=consistency_issues,
                    table_col_letter=table_col,
                    variable_col_letter=variable_col,
                )
                add_critic_summary_comment(ws, consistency_issues)

    write_legend_block(wb)

    if consistency_issues and write_consistency_sheet(wb, consistency_issues):
        highlight_issues_tab(wb, consistency_issues)

    wb.save(xlsx_path)
