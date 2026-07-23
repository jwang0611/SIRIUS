"""Helper functions for SUPP and conditional mapping processing."""

import contextlib
import logging
from collections import defaultdict
from copy import copy as copy_obj
from typing import Any

from openpyxl.styles import PatternFill

from .core import ExcelWriter
from .models import ConditionalRecord, SUPPRecord
from .models.write_result import (
    STAGE_CONDITIONAL_MAPPINGS,
    STAGE_SUPP_ROWS,
    StageWriteResult,
    WriteIssue,
    contains_illegal_characters,
)

logger = logging.getLogger(__name__)


def calculate_char_length(text: str) -> int:
    """Calculate character length with Chinese characters counting as 3.

    This is used for SUPP variable label length validation.
    Chinese characters (CJK Unified Ideographs) count as 3 characters.
    All other characters count as 1.

    Args:
        text: The text to calculate length for

    Returns:
        The calculated character length

    Examples:
        >>> calculate_char_length("Hello")
        5
        >>> calculate_char_length("中文")
        6  # 2 Chinese chars * 3 = 6
        >>> calculate_char_length("中文ABC")
        9  # 2*3 + 3 = 9
    """
    if not text:
        return 0

    length = 0
    for char in text:
        # Check if character is in CJK Unified Ideographs range
        if (
            "\u4e00" <= char <= "\u9fff"
            or "\u3400" <= char <= "\u4dbf"
            or "\u20000" <= char <= "\u2a6df"
            or "\uf900" <= char <= "\ufaff"
        ):
            length += 3
        else:
            length += 1

    return length


def insert_supp_rows(writer: ExcelWriter, supp_records: list[SUPPRecord], highlight: bool) -> StageWriteResult:
    """Insert SUPP domain rows into the Excel file.

    Args:
        writer: ExcelWriter instance
        supp_records: List of SUPPRecord objects
        highlight: Whether to highlight the inserted rows

    Returns:
        A :class:`StageWriteResult`; each SUPP record counts as one attempt and is
        ``written`` only after its row was actually inserted and filled. Records
        whose target domain is missing from the template are ``skipped``.

    Atomicity: every record is pre-validated (illegal characters) BEFORE the
    domain's existing SUPP block is deleted and before any row is inserted, so a
    recoverable per-record problem never leaves a half-filled row behind and
    never destroys the template's existing SUPP rows. If no record of a domain
    survives validation the domain is left completely untouched. Any exception
    raised during the actual insert/fill is unknown/fatal and propagates (the
    job is marked ``failed`` rather than saving a partial workbook).
    """
    result = StageWriteResult(stage=STAGE_SUPP_ROWS)

    # Group SUPP records by sheet/domain
    by_domain: dict[str, list[SUPPRecord]] = defaultdict(list)
    for record in supp_records:
        by_domain[record.sheet_name].append(record)

    # Process each domain (sorted by domain name for consistent output)
    for domain, records in sorted(by_domain.items()):
        # Ensure records within each domain are in consistent order
        # Records should already be sorted from mapper, but maintain insert_row order if set
        records_sorted = sorted(
            records, key=lambda r: (r.insert_row if r.insert_row > 0 else float("inf"), r.source_order)
        )
        logger.info(f"Processing {len(records_sorted)} SUPP rows in domain '{domain}'...")

        # Check if domain exists in template
        if domain not in writer.workbook.sheetnames:
            logger.warning(f"Domain '{domain}' not found in template, skipping {len(records_sorted)} SUPP rows")
            for _record in records_sorted:
                result.record_skipped(
                    WriteIssue(
                        code="domain_not_found",
                        stage=STAGE_SUPP_ROWS,
                        operation="insert_supp_row",
                        sheet=domain,
                    )
                )
            continue

        # Step 0: Pre-validate every record BEFORE any destructive operation
        # (deleting the existing SUPP block, inserting rows). A record whose
        # values openpyxl would reject is recorded as an error with zero
        # mutation. If nothing survives, the domain is left untouched so the
        # template's existing SUPP rows are never lost to a failed rewrite.
        valid_records = []
        for record in records_sorted:
            if contains_illegal_characters(*record.to_row_data().values()):
                result.record_error(
                    WriteIssue(
                        code="illegal_characters",
                        stage=STAGE_SUPP_ROWS,
                        operation="insert_supp_row",
                        sheet=domain,
                    )
                )
                logger.warning(
                    "Skipped SUPP row '%s' in '%s': value contains illegal characters",
                    record.variable_name,
                    domain,
                )
            else:
                valid_records.append(record)
        if not valid_records:
            logger.warning(f"No writable SUPP rows for domain '{domain}'; existing SUPP block left untouched")
            continue
        records_sorted = valid_records

        ws = writer.get_worksheet(domain)

        # Step 1: Find and save CONTENT hyperlink information (before any modifications)
        content_row = None
        content_hyperlink = None
        for row_idx in range(14, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value and str(cell.value).upper() == "CONTENT":
                content_row = row_idx
                content_hyperlink = cell.hyperlink
                logger.debug(f"Found CONTENT at row {content_row} with hyperlink: {content_hyperlink}")
                break

        # Step 2: Delete existing SUPP rows from template
        rows_to_delete = []
        for row_idx in range(14, ws.max_row + 1):
            var_type = ws.cell(row=row_idx, column=9).value  # Variable Type column
            if var_type == "SUPP":
                rows_to_delete.append(row_idx)

        # Delete from bottom to top to avoid row index shifting issues
        if rows_to_delete:
            logger.info(f"  Removing {len(rows_to_delete)} existing SUPP rows from template")
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx, 1)
                logger.debug(f"  Deleted existing SUPP row at {row_idx}")
                # Adjust content_row if it was pushed up by deletion
                if content_row and row_idx < content_row:
                    content_row -= 1

        # Step 3: Find the last SDTM row (after deletion)
        last_sdtm_row = writer.find_last_row_with_type(
            sheet_name=domain,
            type_column=9,  # Variable Type column
            type_value="SDTM",
            start_row=14,  # Start from row 14 (after header)
        )

        if last_sdtm_row is None:
            # No SDTM rows, insert after header
            insert_row = 14
            logger.debug(f"No SDTM rows found, inserting at row {insert_row}")
        else:
            # Insert after the last SDTM row
            insert_row = last_sdtm_row + 1
            logger.debug(f"Inserting after last SDTM row {last_sdtm_row}")

        # Step 4: Insert and fill SUPP rows. Records were pre-validated above,
        # so any exception here is unknown/fatal and intentionally propagates —
        # continuing after a mid-fill failure would save a half-written row.
        num_inserted = 0
        for record in records_sorted:
            current_insert_row = insert_row + num_inserted

            # Insert new row, copying format from the row BEFORE the insertion point
            # (not the row after, which may have different formatting)
            copy_from_row = insert_row - 1 if insert_row > 14 else 14
            writer.insert_row(
                sheet_name=domain, row_index=current_insert_row, amount=1, copy_format_from_row=copy_from_row
            )

            # Fill the row with SUPP data
            record.insert_row = current_insert_row
            writer.fill_row(
                sheet_name=domain, row=current_insert_row, column_data=record.to_row_data(), highlight=highlight
            )

            # Overwrite length cell (column D) with a numeric value so that the
            # cell number format inherited from the SDTM row (e.g. "$"0) renders
            # correctly.  fill_row writes str(), which Excel treats as text.
            length_cell = ws.cell(row=current_insert_row, column=4)
            ref_cell = ws.cell(row=copy_from_row, column=4)
            with contextlib.suppress(ValueError, TypeError):
                length_cell.value = int(record.length)
            if ref_cell.number_format:
                length_cell.number_format = ref_cell.number_format

            # Set variable order formula (J column = column 10)
            ws.cell(row=current_insert_row, column=10).value = "=ROW()-13"

            num_inserted += 1
            result.record_written()

            # Check B column (variable_label) character length
            # Chinese characters count as 3, others as 1
            # If exceeds 40, highlight B column with red
            label_length = calculate_char_length(record.variable_label)

            # Also check if SUPP variable has multiple sources
            # Multiple sources means transformation_def has multiple [RAW] references
            has_multi_source = False
            if record.transformation_def:
                trans_def = str(record.transformation_def)
                # Count [RAW] references - this is more accurate than counting semicolons
                # because notes/hints also use semicolons
                raw_count = trans_def.count("[RAW]")
                if raw_count > 1:
                    has_multi_source = True

            # Highlight B column red if: label too long OR has multiple sources
            if label_length > 40 or has_multi_source:
                cell_b = ws.cell(row=current_insert_row, column=2)
                cell_b.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                if label_length > 40:
                    result.add_warning(
                        WriteIssue(
                            code="supp_label_too_long",
                            stage=STAGE_SUPP_ROWS,
                            operation="insert_supp_row",
                            sheet=domain,
                            row=current_insert_row,
                            column=2,
                        )
                    )
                    logger.warning(
                        f"  SUPP variable '{record.variable_name}' label exceeds 40 chars (length={label_length})"
                    )
                if has_multi_source:
                    result.add_warning(
                        WriteIssue(
                            code="supp_multi_source",
                            stage=STAGE_SUPP_ROWS,
                            operation="insert_supp_row",
                            sheet=domain,
                            row=current_insert_row,
                            column=2,
                        )
                    )
                    logger.info(
                        f"  SUPP variable '{record.variable_name}' has multiple sources, B column highlighted red"
                    )

            logger.info(f"  Inserted SUPP row: {domain}.{record.variable_name} at row {current_insert_row}")

        # Step 5: Fix CONTENT hyperlink if it was affected by insertion
        if content_row and content_hyperlink and num_inserted and insert_row <= content_row:
            # CONTENT was pushed down by insertion
            new_content_row = content_row + num_inserted

            # Remove old hyperlink from the original position
            old_cell = ws.cell(row=content_row, column=1)
            old_cell.hyperlink = None

            # Add hyperlink to the new CONTENT position
            new_cell = ws.cell(row=new_content_row, column=1)
            new_cell.hyperlink = content_hyperlink

            logger.info(f"  Fixed CONTENT hyperlink: moved from row {content_row} to row {new_content_row}")

    return result


def _find_conditional_column_group(ws: Any, column_names: list[str]) -> int | None:
    """Locate an existing conditional COLUMN GROUP in a TEST sheet.

    Returns the 1-based start column of a contiguous run of row-1 headers that
    matches ``column_names`` exactly (e.g. ``CRF_TESTCD`` immediately followed
    by ``CRF_ORRES``), or ``None`` when the group is not present. Matching the
    whole group — never a single header name — is what keeps the TESTCD and
    TEST groups of the same sheet distinct even though both end in
    ``CRF_ORRES``.
    """
    n = len(column_names)
    if n == 0:
        return None
    for start in range(1, ws.max_column - n + 2):
        if all(str(ws.cell(row=1, column=start + i).value or "").strip() == column_names[i] for i in range(n)):
            return start
    return None


def process_conditional_mappings(writer: ExcelWriter, conditional_records: list[ConditionalRecord]) -> StageWriteResult:
    """Process conditional mapping records by adding columns to target sheets.

    For each conditional record:
    1. Check if target sheet exists (warn if not)
    2. Find the last non-empty column
    3. Add new columns after it
    4. Write header row (row 1)
    5. Write data rows starting from row 2

    Args:
        writer: ExcelWriter instance
        conditional_records: List of ConditionalRecord objects

    Returns:
        A :class:`StageWriteResult`; each conditional record counts as one attempt.
        Records whose target sheet is missing are ``skipped``; records whose
        values openpyxl would reject are recorded as ``illegal_characters``
        errors BEFORE any cell is touched.

    Idempotent per COLUMN GROUP: when the record's ``column_names`` already
    exist as an exact contiguous header run in row 1 (e.g. from a previous run
    whose output is re-used as the template), that group is reused in place —
    its data rows are overwritten and stale rows below the new data cleared —
    instead of appending a duplicate group on every run. Groups are matched as
    a unit, so a sheet holding both a TESTCD group (``CRF_TESTCD, CRF_ORRES``)
    and a TEST group (``CRF_TEST, CRF_ORRES``) keeps both ``CRF_ORRES`` columns
    with their own data (same layout as the original append-only behaviour).

    Examples:
        Sheet FATEST:
        - Add columns CRF_TESTCD (condition values) and CRF_ORRES (raw references)
        - Row 1: CRF_TESTCD, CRF_ORRES
        - Row 2: THDIA, [RAW]TH.THDIA
        - Row 3: THFDTC, [RAW]TH.THFDAT
    """
    logger.info(f"=== Processing {len(conditional_records)} conditional mapping group(s) ===")
    result = StageWriteResult(stage=STAGE_CONDITIONAL_MAPPINGS)

    for record in conditional_records:
        sheet_name = record.sheet_name

        # Check if sheet exists
        if not writer.sheet_exists(sheet_name):
            logger.warning(
                f"Sheet '{sheet_name}' not found in template, skipping conditional mapping "
                f"for columns: {', '.join(record.column_names)}"
            )
            result.record_skipped(
                WriteIssue(
                    code="sheet_not_found",
                    stage=STAGE_CONDITIONAL_MAPPINGS,
                    operation="write_conditional_columns",
                    sheet=sheet_name,
                )
            )
            continue

        # Pre-validate all values BEFORE touching the sheet so a recoverable
        # problem never leaves partially-written columns behind.
        flat_values = list(record.column_names) + [v for row in record.mappings for v in row]
        if contains_illegal_characters(*flat_values):
            result.record_error(
                WriteIssue(
                    code="illegal_characters",
                    stage=STAGE_CONDITIONAL_MAPPINGS,
                    operation="write_conditional_columns",
                    sheet=sheet_name,
                )
            )
            logger.warning("Skipped conditional mapping for '%s': value contains illegal characters", sheet_name)
            continue

        logger.info(f"Processing conditional mapping for sheet '{sheet_name}': {len(record.mappings)} row(s)")

        # Get worksheet
        ws = writer.get_worksheet(sheet_name)

        # Get template font style from an existing cell (row 1, col 1)
        # This ensures new columns match the template's font
        template_cell = ws.cell(row=1, column=1)
        template_font: Any = template_cell.font

        # Idempotent reuse works on the whole COLUMN GROUP, never on a single
        # header name: a sheet can legitimately hold both a TESTCD group
        # (CRF_TESTCD + CRF_ORRES) and a TEST group (CRF_TEST + CRF_ORRES), so
        # CRF_ORRES is not globally unique. Reusing by individual header would
        # make the second group overwrite the first group's ORRES data. A group
        # is reused only when the record's column_names appear as an exact
        # contiguous run of row-1 headers (which is how every run writes them);
        # otherwise the whole group is appended after the last header, exactly
        # like the pre-idempotency behaviour.
        last_col = 1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value:
                last_col = col

        group_start = _find_conditional_column_group(ws, record.column_names)
        if group_start is not None:
            col_indices = list(range(group_start, group_start + len(record.column_names)))
            logger.info(
                f"  Reusing existing column group ({', '.join(record.column_names)}) at position "
                f"{group_start} in '{sheet_name}' (repeated run); no duplicates added"
            )
        else:
            col_indices = list(range(last_col + 1, last_col + 1 + len(record.column_names)))
            logger.info(f"  Adding {len(record.column_names)} column(s) starting at position {last_col + 1}")

        # Write headers for all columns
        for col_name, col_idx in zip(record.column_names, col_indices, strict=True):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.value = col_name
            if template_font:
                header_cell.font = copy_obj(template_font)

        # Write data rows
        for row_offset, row_values in enumerate(record.mappings, start=2):
            # row_values is a tuple with len(column_names) values
            for i, value in enumerate(row_values):
                cell = ws.cell(row=row_offset, column=col_indices[i])
                cell.value = value
                if template_font:
                    cell.font = copy_obj(template_font)

            logger.debug(f"    Row {row_offset}: {row_values}")

        # Clear stale data below the freshly-written block (a previous run may
        # have written more rows than this one).
        first_stale_row = 2 + len(record.mappings)
        for col_idx in col_indices:
            for row_idx in range(first_stale_row, ws.max_row + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    ws.cell(row=row_idx, column=col_idx).value = None

        result.record_written()
        logger.info(f"  ✓ Added {len(record.mappings)} row(s) to '{sheet_name}' ({', '.join(record.column_names)})")

    logger.info("=== Completed conditional mappings ===")
    return result
