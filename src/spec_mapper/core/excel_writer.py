"""Excel file writer that preserves formatting."""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models.template_record import CellUpdate
from ..models.write_result import (
    STAGE_CELL_UPDATES,
    STAGE_CODELIST_RECORDS,
    StageWriteResult,
    WriteIssue,
    contains_illegal_characters,
)
from .excel import format_utils as _fmt

logger = logging.getLogger(__name__)

# CONTENT "class" column labels that mark a findings-class domain (needs SDTM.SV).
# Chinese labels appear in the v3.2 (v1.1) template; English ones in the IG 3.4 template.
FINDINGS_CLASS_LABELS = frozenset({"发现", "Findings"})


class ExcelWriter:
    """Write to Excel files while preserving formatting.

    This writer uses openpyxl to load existing Excel files and only
    modifies cell values, preserving all formatting, styles, and structure.

    Examples:
        >>> writer = ExcelWriter("template.xlsx")
        >>> writer.update_cell_value("DM", 15, 5, "[RAW]DM.SUBJID")
        >>> writer.save("output.xlsx")
    """

    def __init__(self, file_path: str | Path):
        """Initialize Excel writer by loading existing workbook.

        Args:
            file_path: Path to the Excel file to modify

        Raises:
            FileNotFoundError: If the file does not exist
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        logger.info(f"Loading workbook '{self.file_path.name}'")
        # Load with formatting preserved (data_only=False to keep formulas)
        # keep_vba=False to avoid compatibility issues
        self.workbook: Workbook = load_workbook(
            self.file_path,
            data_only=False,
            keep_vba=False,  # Don't preserve VBA macros to avoid issues
        )
        logger.info(f"Workbook loaded with {len(self.workbook.sheetnames)} sheets")

        # Track sheets that have been modified (for tab highlighting)
        self.modified_sheets: set[str] = set()

        # Per-item skips made *inside* a batch write method whose public return
        # value is only a mutation count. Without this queue such an item is
        # invisible in the job result (the batch reports attempted == written
        # and no issue). The caller drains it right after each call — see
        # :meth:`drain_pending_skips` and ``SpecMapper._guard``.
        self.pending_skips: list[WriteIssue] = []

        # Optional {source_label: translated_label} map applied whenever a Source
        # (F) column value is written. Used to render the IG 3.4 (English)
        # template's Source column in English; empty = no translation (v3.2).
        self.source_label_overrides: dict[str, str] = {}

    def _record_pending_skip(
        self,
        code: str,
        operation: str,
        sheet: str | None = None,
        variable: str | None = None,
    ) -> None:
        """Queue one per-item skip performed inside a batch write method.

        ``stage`` is left empty on purpose: the writer knows the operation,
        the sheet and the variable, but only the caller knows which write
        stage the call belongs to. :meth:`drain_pending_skips` stamps it.

        Contract: only methods invoked through ``SpecMapper._guard`` may queue
        skips here, because the guard is what drains them. A method that
        returns its own :class:`StageWriteResult` must record skips on that
        result instead.
        """
        self.pending_skips.append(
            WriteIssue(code=code, stage="", operation=operation, sheet=sheet, variable=variable or None)
        )

    def drain_pending_skips(self, stage: str) -> list[WriteIssue]:
        """Return and clear the per-item skips queued since the last drain.

        Each returned issue is stamped with *stage* so it can be recorded on
        the matching :class:`StageWriteResult`.
        """
        issues = self.pending_skips
        self.pending_skips = []
        for issue in issues:
            issue.stage = stage
        return issues

    def _src_label(self, value: str | None) -> str | None:
        """Translate a Source (F column) label via ``source_label_overrides``."""
        if value is None:
            return None
        return self.source_label_overrides.get(str(value).strip(), value)

    def _preserve_source_labels(self) -> set[str]:
        """Source values that ``update_domain_source_column`` must not overwrite.

        Always includes the Chinese assigned/derived labels; when a translation
        is active, also includes their translated forms so an English template's
        existing "Assigned"/"Derived" values are preserved too.
        """
        return {"指定", "衍生", *self.source_label_overrides.values()}

    def update_cell_value(
        self, sheet_name: str, row: int, col: int, value: str, highlight: bool = False, clear_existing: bool = False
    ) -> None:
        """Update a single cell value while preserving formatting.

        Args:
            sheet_name: Name of the worksheet
            row: Row number (1-based, Excel row number)
            col: Column number (1-based, Excel column number)
            value: New value to write to the cell
            highlight: If True, apply yellow background and italic font
            clear_existing: If True, clear existing value before writing new one

        Raises:
            ValueError: If sheet name doesn't exist or row/col are invalid
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        if row <= 0 or col <= 0:
            raise ValueError(f"Row and column must be positive: row={row}, col={col}")

        ws: Worksheet = self.workbook[sheet_name]

        # Get the cell
        cell = ws.cell(row=row, column=col)
        old_value = cell.value

        # Clear hyperlink if exists (hyperlinks may interfere with cell value display)
        if cell.hyperlink:
            logger.debug(f"Clearing hyperlink from {sheet_name}!{cell.coordinate}")
            cell.hyperlink = None

        # Clear existing value if requested
        if clear_existing and old_value:
            logger.debug(f"Clearing existing value: '{old_value}'")

        # A value that starts with "=" but contains a newline is not a valid
        # Excel formula (e.g. an aggregated HYPERLINK + description).  Strip
        # the leading "=" so Excel treats it as plain text.
        if isinstance(value, str) and value.startswith("=") and "\n" in value:
            value = value[1:]

        # Set new value
        cell.value = value

        # Track that this sheet has been modified
        self.modified_sheets.add(sheet_name)

        # Apply formatting if highlight is True
        if highlight:
            # Yellow fill (RGB: FFFF00)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Italic font, preserve other font properties
            if cell.font:
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=True,  # Make italic
                    color=cell.font.color,
                )
            else:
                cell.font = Font(italic=True)

            logger.debug(f"Applied highlight formatting (yellow fill, italic) to {cell.coordinate}")

        logger.debug(f"Updated {sheet_name}!{cell.coordinate}: '{old_value}' -> '{value}'")

    def update_cells(
        self, updates: list[CellUpdate], highlight: bool = False, clear_existing: bool = False
    ) -> StageWriteResult:
        """Update multiple cells at once.

        Each cell is attempted independently. Recognized recoverable
        preconditions (missing sheet, values openpyxl would reject) are checked
        BEFORE the cell is touched and recorded as structured errors, so a
        recoverable outcome never leaves the cell partially modified (e.g. a
        cleared hyperlink without the new value). Any exception raised during
        the actual write is unknown/fatal and propagates.

        Args:
            updates: List of CellUpdate objects describing the changes
            highlight: If True, apply yellow background and italic font to all updates
            clear_existing: If True, clear existing values before writing new ones

        Returns:
            A :class:`StageWriteResult` where ``written`` counts only cells whose
            workbook mutation actually succeeded.
        """
        logger.info(f"Applying {len(updates)} cell updates")
        result = StageWriteResult(stage=STAGE_CELL_UPDATES)

        for update in updates:
            value = update.value
            # Keep Source column labels language-aware even for generic
            # mapper updates (e.g. assignment mappings that set F="指定").
            if update.col == 6 and isinstance(value, str):
                mapped = self._src_label(value)
                if mapped is not None:
                    value = mapped
            # Recognized recoverable precondition: the target sheet is absent.
            # Record it explicitly rather than relying on a broad ValueError
            # catch (which would also mask genuine bugs).
            if update.sheet_name not in self.workbook.sheetnames:
                result.record_error(
                    WriteIssue(
                        code="cell_write_failed",
                        stage=STAGE_CELL_UPDATES,
                        operation="update_cell",
                        sheet=update.sheet_name,
                        row=update.row,
                        column=update.col,
                        detail="sheet_not_found",
                    )
                )
                logger.warning(
                    "Skipped cell update at %s!R%sC%s (sheet not found)",
                    update.sheet_name,
                    update.row,
                    update.col,
                )
                continue
            # Recognized recoverable precondition: openpyxl would reject the
            # value. Checked BEFORE update_cell_value so the cell is never
            # partially modified (its hyperlink is cleared before the value is
            # assigned in there).
            if contains_illegal_characters(value):
                result.record_error(
                    WriteIssue(
                        code="illegal_characters",
                        stage=STAGE_CELL_UPDATES,
                        operation="update_cell",
                        sheet=update.sheet_name,
                        row=update.row,
                        column=update.col,
                    )
                )
                logger.warning(
                    "Skipped cell update at %s!R%sC%s (illegal characters)",
                    update.sheet_name,
                    update.row,
                    update.col,
                )
                continue
            self.update_cell_value(
                update.sheet_name,
                update.row,
                update.col,
                value,
                highlight=highlight,
                clear_existing=clear_existing,
            )
            result.record_written()
            if update.reason:
                logger.debug(f"  Reason: {update.reason}")

        logger.info("Cell updates: %s/%s written, %s error(s)", result.written, result.attempted, len(result.errors))
        return result

    def insert_row(
        self, sheet_name: str, row_index: int, amount: int = 1, copy_format_from_row: int | None = None
    ) -> None:
        """Insert one or more rows at the specified position.

        Note: This will shift existing rows down. If copy_format_from_row is specified,
        the formatting will be copied from that row to the newly inserted rows.
        Merged cells that overlap the inserted rows will be unmerged to avoid issues.

        Args:
            sheet_name: Name of the worksheet
            row_index: Position to insert row (1-based)
            amount: Number of rows to insert
            copy_format_from_row: If specified, copy formatting from this row (1-based)

        Raises:
            ValueError: If sheet name doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Track that this sheet has been modified
        self.modified_sheets.add(sheet_name)

        # Store formatting if needed
        format_data = None
        if copy_format_from_row is not None:
            format_data = self._copy_row_format(ws, copy_format_from_row)

        # Insert rows (this pushes existing rows down, including their values)
        ws.insert_rows(row_index, amount)

        # **Critical: Clear all cell values in the newly inserted rows**
        # This prevents leftover data from being retained after insertion
        for row_offset in range(amount):
            current_row = row_index + row_offset
            for col_idx in range(1, 15):  # Clear first 14 columns (typical SDTM range)
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = None

        # Unmerge any cells in the newly inserted rows to avoid unwanted merges
        # openpyxl's insert_rows() may automatically extend merged cells
        self._unmerge_inserted_rows(ws, row_index, amount)

        # Apply formatting to new rows
        if format_data:
            for i in range(amount):
                self._apply_row_format(ws, row_index + i, format_data)

        logger.info(f"Inserted {amount} row(s) at position {row_index} in '{sheet_name}'")

    def save(self, output_path: str | Path | None = None) -> None:
        """Save the workbook to a file.

        Args:
            output_path: Path to save the file. If None, overwrites original file.

        Raises:
            IOError: If file cannot be saved
        """
        if output_path is None:
            output_path = self.file_path
        else:
            output_path = Path(output_path)

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Saving workbook '{output_path.name}'")
        try:
            self.workbook.save(output_path)
            logger.info(f"Successfully saved workbook '{output_path.name}'")
        except Exception as e:
            logger.error("Failed to save workbook '%s' (%s)", output_path.name, type(e).__name__)
            raise OSError(f"Failed to save workbook '{output_path.name}'") from e

    def highlight_modified_sheet_tabs(self, color: str = "FFFF00") -> int:
        """Highlight tabs of sheets that have been modified.

        Returns the number of tabs highlighted (0 = nothing to do).
        """
        highlighted = _fmt.highlight_sheet_tabs(
            self.workbook,
            modified_sheets=self.modified_sheets,
            color=color,
            clear_existing=True,
        )
        if highlighted:
            logger.info(f"Highlighted {highlighted} sheet tab(s) with color #{color}")
        else:
            logger.debug("No modified sheets to highlight")
        return highlighted

    def set_active_sheet(self, sheet_name: str) -> int:
        """Set the active sheet (the one shown when opening the file).

        Args:
            sheet_name: Name of the sheet to set as active

        Returns:
            1 if the active sheet was set, 0 if the sheet does not exist.
        """
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(
                f"Cannot set active sheet '{sheet_name}': sheet not found. "
                f"Available sheets: {', '.join(self.workbook.sheetnames)}"
            )
            return 0

        # Get the sheet index
        sheet_index = self.workbook.sheetnames.index(sheet_name)

        # Set the active sheet
        self.workbook.active = sheet_index

        logger.info(f"Set active sheet to '{sheet_name}' (index: {sheet_index})")
        return 1

    def close(self) -> None:
        """Close the workbook.

        Note: After closing, the writer cannot be used anymore.
        """
        self.workbook.close()
        logger.info("Workbook closed")

    def sheet_exists(self, sheet_name: str) -> bool:
        """Check if a sheet exists in the workbook.

        Args:
            sheet_name: Name of the worksheet to check

        Returns:
            True if sheet exists, False otherwise
        """
        return sheet_name in self.workbook.sheetnames

    def get_worksheet(self, sheet_name: str) -> Worksheet:
        """Get a worksheet by name.

        Args:
            sheet_name: Name of the worksheet

        Returns:
            Worksheet object

        Raises:
            ValueError: If sheet name doesn't exist
        """
        if not self.sheet_exists(sheet_name):
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        return self.workbook[sheet_name]

    def write_cell(self, sheet_name: str, row: int, col: int, value: str) -> None:
        """Write a value to a cell without any formatting changes.

        Args:
            sheet_name: Name of the worksheet
            row: Row number (1-based)
            col: Column number (1-based)
            value: Value to write

        Raises:
            ValueError: If sheet doesn't exist or row/col are invalid
        """
        ws = self.get_worksheet(sheet_name)

        if row <= 0 or col <= 0:
            raise ValueError(f"Row and column must be positive: row={row}, col={col}")

        ws.cell(row=row, column=col).value = value
        logger.debug(f"Wrote '{value}' to {sheet_name}!{row}:{col}")

    def __enter__(self) -> "ExcelWriter":
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:  # type: ignore
        """Context manager exit - automatically save and close."""
        if exc_type is None:
            # No exception occurred, save the file
            self.save()
        self.close()

    def get_cell_value(self, sheet_name: str, row: int, col: int) -> str:
        """Get current value of a cell.

        Args:
            sheet_name: Name of the worksheet
            row: Row number (1-based)
            col: Column number (1-based)

        Returns:
            Current cell value as string

        Raises:
            ValueError: If sheet name doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]
        cell = ws.cell(row=row, column=col)
        value = cell.value

        if value is None:
            return ""
        return str(value)

    def fill_row(
        self, sheet_name: str, row: int, column_data: dict, highlight: bool = False, clear_first: bool = True
    ) -> None:
        """Fill a row with data from a dictionary.

        Args:
            sheet_name: Name of the worksheet
            row: Row number (1-based)
            column_data: Dictionary mapping column indices (1-based) to values
            highlight: If True, apply yellow background and italic font
            clear_first: If True, clear all cell values in relevant columns before filling

        Raises:
            ValueError: If sheet name doesn't exist

        Examples:
            >>> writer.fill_row("DM", 50, {1: "SITENAM", 2: "Site Name", 8: "[RAW]DM.SITE"})
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Clear relevant columns first if requested (to avoid leftover data)
        if clear_first:
            # Clear columns that should be filled (1-11 typically for SDTM)
            for col_idx in range(1, 12):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = None

        # Fill with new data
        for col_idx, value in column_data.items():
            if value:  # Only fill non-empty values
                self.update_cell_value(
                    sheet_name=sheet_name,
                    row=row,
                    col=col_idx,
                    value=str(value),
                    highlight=highlight,
                    clear_existing=False,
                )

        logger.info(f"Filled row {row} in '{sheet_name}' with {len(column_data)} values")

    def _copy_row_format(self, ws: Worksheet, row_num: int) -> dict:
        """Copy formatting from a row (delegates to :mod:`excel.format_utils`)."""
        return _fmt.copy_row_format(ws, row_num)

    def _apply_row_format(self, ws: Worksheet, row_num: int, format_data: dict) -> None:
        """Apply captured formatting to a row (delegates to :mod:`excel.format_utils`)."""
        _fmt.apply_row_format(ws, row_num, format_data)

    def _unmerge_inserted_rows(self, ws: Worksheet, start_row: int, amount: int) -> None:
        """Unmerge cells touching ``[start_row, start_row+amount-1]``."""
        unmerged = _fmt.unmerge_rows_in_range(ws, start_row, amount)
        if unmerged:
            end_row = start_row + amount - 1
            for range_str in unmerged:
                logger.debug(f"Unmerged cells {range_str} that overlapped with inserted rows {start_row}-{end_row}")

    def find_last_row_with_type(
        self, sheet_name: str, type_column: int, type_value: str, start_row: int = 1
    ) -> int | None:
        """Find the last row where a specific column has a specific value.

        Args:
            sheet_name: Name of the worksheet
            type_column: Column index to search in (1-based)
            type_value: Value to search for
            start_row: Starting row for search (1-based, default: 1)

        Returns:
            Row number of the last matching row, or None if not found

        Examples:
            >>> writer.find_last_row_with_type("DM", 9, "SUPP", 14)
            45  # Last row in DM where column 9 = "SUPP"
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]
        last_row = _fmt.find_last_row_with_value(ws, column=type_column, value=type_value, start_row=start_row)

        if last_row:
            logger.debug(f"Found last row with {type_value} in column {type_column}: row {last_row}")
        else:
            logger.debug(f"No row found with {type_value} in column {type_column}")

        return last_row

    def add_supp_to_content_sheet(
        self, domain: str, content_sheet_name: str = "CONTENT", reference_row: int = 52, highlight: bool = True
    ) -> int:
        """Add SUPP information to the CONTENT sheet for a given domain.

        This method:
        1. Finds the row where column A equals the domain name
        2. Inserts a new row below it
        3. Copies formatting from the reference row (default: row 52)
        4. Sets column A to SUPPXX with a hyperlink to the domain sheet
        5. Sets column B to "补充修饰语数据集XX"
        6. Sets sorting variables based on domain (DM vs others)
        7. Applies yellow highlight if highlight=True

        Args:
            domain: Domain name (e.g., "AE", "DM")
            content_sheet_name: Name of the CONTENT sheet (default: "CONTENT")
            reference_row: Row number to copy formatting from (default: 52)
            highlight: Whether to apply yellow highlight to filled cells (default: True)

        Idempotent: if ``SUPP{domain}`` already exists in the CONTENT sheet the
        existing row is updated in place, so re-running the pipeline on a
        previously-generated workbook never appends a duplicate CONTENT row.

        Returns:
            1 if the SUPP row was inserted or refreshed, 0 if the domain row was
            not found in the CONTENT sheet.

        Raises:
            ValueError: If the CONTENT sheet does not exist.

        Examples:
            >>> writer.add_supp_to_content_sheet("AE")  # Adds SUPPAE below AE row
        """
        if content_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{content_sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[content_sheet_name]

        # Find the row where column A equals the domain
        domain_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip().upper() == domain.upper():
                domain_row = row_idx
                break

        if domain_row is None:
            logger.warning(
                f"Domain '{domain}' not found in column A of '{content_sheet_name}'. Cannot add SUPP information."
            )
            return 0

        logger.info(f"Found domain '{domain}' at row {domain_row} in '{content_sheet_name}'")

        # Idempotency: if SUPP{domain} is already present (e.g. re-running the
        # pipeline on a previously-generated workbook), update that row in place
        # instead of appending a duplicate CONTENT row.
        supp_key = f"SUPP{domain}".upper()
        existing_supp_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip().upper() == supp_key:
                existing_supp_row = row_idx
                break

        if existing_supp_row is not None:
            insert_row_idx = existing_supp_row
            logger.info(
                f"'{supp_key}' already present in '{content_sheet_name}' at row {insert_row_idx}; "
                "updating in place (no duplicate row)"
            )
        else:
            # Find the last domain/SUPP row to insert at the bottom (avoid disturbing header)
            last_domain_row = None
            for row_idx in range(ws.max_row, 0, -1):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # Check if it's a valid domain name (2-char or SUPP*)
                    if (len(cell_str) == 2 and cell_str.isalpha()) or cell_str.startswith("SUPP"):
                        last_domain_row = row_idx
                        break

            # Insert new row at the bottom (after the last domain/SUPP)
            insert_row_idx = (last_domain_row or domain_row) + 1
            self.insert_row(
                sheet_name=content_sheet_name, row_index=insert_row_idx, amount=1, copy_format_from_row=reference_row
            )

        # Prepare SUPP data
        supp_domain = f"SUPP{domain}"
        supp_description = f"补充修饰语数据集{domain}"

        # Fixed text for C, D, E, G, H columns
        col_c_text = "每个受试者一个域一个修饰变量一条记录"
        col_d_sas = f"{domain.lower()}.sas"
        col_e_text = "特殊用途"
        col_g_key_vars = "STUDYID, RDOMAIN, USUBJID, IDVAR, IDVARVAL, QNAM"
        col_h_sort_vars = "STUDYID, RDOMAIN, USUBJID, QNAM"

        # Fill the new row with SUPP information
        # Column A: SUPPXX (with hyperlink)
        cell_a = ws.cell(row=insert_row_idx, column=1)
        cell_a.value = supp_domain

        # Add hyperlink to the domain sheet (if it exists)
        if domain in self.workbook.sheetnames:
            cell_a.hyperlink = f"#{domain}!A1"
            # Apply hyperlink style (blue, underlined)
            cell_a.font = Font(
                name=cell_a.font.name if cell_a.font else None,
                size=cell_a.font.size if cell_a.font else None,
                bold=cell_a.font.bold if cell_a.font else None,
                italic=cell_a.font.italic if cell_a.font else None,
                color="0563C1",  # Blue color for hyperlinks
                underline="single",
            )
            logger.debug(f"Added hyperlink from {supp_domain} to {domain} sheet")

        # Column B: Description
        ws.cell(row=insert_row_idx, column=2).value = supp_description

        # Column C: Fixed text
        ws.cell(row=insert_row_idx, column=3).value = col_c_text

        # Column D: Source SAS file (e.g. "tu.sas")
        ws.cell(row=insert_row_idx, column=4).value = col_d_sas

        # Column E: Fixed text
        ws.cell(row=insert_row_idx, column=5).value = col_e_text

        # Column G: Key variables (fixed for all SUPP)
        ws.cell(row=insert_row_idx, column=7).value = col_g_key_vars

        # Column H: Sorting variables (depends on domain)
        ws.cell(row=insert_row_idx, column=8).value = col_h_sort_vars

        # Apply yellow highlight to all filled cells if highlight=True
        if highlight:
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Columns to highlight: A, B, C, D, E, G, H (1, 2, 3, 4, 5, 7, 8)
            for col_idx in [1, 2, 3, 4, 5, 7, 8]:
                ws.cell(row=insert_row_idx, column=col_idx).fill = yellow_fill

        # Note: Hyperlink fixing is done separately after all SUPP rows are inserted
        # See fix_content_sheet_hyperlinks() method

        logger.info(
            f"Added {supp_domain} information to '{content_sheet_name}' at row {insert_row_idx}"
            f"{' with yellow highlight' if highlight else ''}"
        )
        return 1

    def fix_content_sheet_hyperlinks(self, content_sheet_name: str = "CONTENT") -> int:
        """Fix all hyperlinks in CONTENT sheet.

        After inserting SUPP rows, hyperlinks in CONTENT sheet may be misaligned.
        This method rebuilds all hyperlinks based on the cell values (domain names).

        Args:
            content_sheet_name: Name of the CONTENT sheet (default: "CONTENT")

        Raises:
            ValueError: If CONTENT sheet doesn't exist
        """
        from openpyxl.worksheet.hyperlink import Hyperlink

        if content_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{content_sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[content_sheet_name]
        fixed_count = 0

        # Iterate through all rows in CONTENT sheet
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            cell_value = cell.value

            if not cell_value:
                continue

            cell_value_str = str(cell_value).strip()

            # Skip SUPP-- (reference row)
            if cell_value_str == "SUPP--":
                continue

            # Determine the correct hyperlink location
            if cell_value_str.startswith("SUPP"):
                # SUPP entries link to their parent domain
                parent_domain = cell_value_str.replace("SUPP", "")
                if parent_domain and parent_domain in self.workbook.sheetnames:
                    expected_location = f"{parent_domain}!A1"
                    display_text = cell_value_str
                else:
                    continue
            elif cell_value_str in self.workbook.sheetnames:
                # Regular domain entries link to themselves
                expected_location = f"{cell_value_str}!A1"
                display_text = cell_value_str
            else:
                # Not a domain, skip
                continue

            # Always recreate the hyperlink to ensure it's correct
            # (openpyxl may adjust hyperlinks during save, so we force recreate)
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=expected_location, display=display_text)
            fixed_count += 1
            logger.debug(f"Set hyperlink at row {row_idx} ({cell_value_str}): -> {expected_location}")

        if fixed_count > 0:
            logger.info(f"Fixed {fixed_count} hyperlink(s) in '{content_sheet_name}' sheet")
        return fixed_count

    def apply_fixed_variable_rules(
        self,
        sheet_name: str,
        fixed_variables: dict,
        start_row: int = 14,
        variable_name_column: int = 1,
        source_column: int = 6,
        transformation_column: int = 8,
        highlight: bool = True,
    ) -> int:
        """Apply fixed variable rules from configuration.

        This method applies predefined transformation rules for specific SDTM variables
        like USUBJID, STUDYID, xxSEQ, VISIT, VISITNUM, and DOMAIN.

        Args:
            sheet_name: Name of the worksheet (domain sheet like "AE", "CM", "DM")
            fixed_variables: Dictionary of fixed variable rules from config
            start_row: Starting row for updates (default: 14, first data row)
            variable_name_column: Column index for variable name (default: 1, column A)
            source_column: Column index for source (default: 6, column F)
            transformation_column: Column index for transformation (default: 8, column H)
            highlight: If True, apply yellow background to updated cells (default: True)

        Raises:
            ValueError: If sheet name doesn't exist

        Examples:
            >>> writer.apply_fixed_variable_rules("AE", fixed_vars_config)
        """
        import re

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]
        domain = sheet_name.upper()
        is_dm = domain == "DM"

        updated_count = 0

        # Iterate through rows
        for row_idx in range(start_row, ws.max_row + 1):
            variable_name = ws.cell(row=row_idx, column=variable_name_column).value

            if not variable_name:
                continue

            var_name_str = str(variable_name).strip().upper()

            # Check each fixed variable rule
            for rule_name, rule_config in fixed_variables.items():
                matched = False

                # Skip rule if current domain is in the exclude list
                exclude_domains = rule_config.get("exclude_domains", [])
                if domain in [d.upper() for d in exclude_domains]:
                    continue

                # Check if this is a pattern-based rule
                if "pattern" in rule_config:
                    pattern = rule_config["pattern"]
                    if re.match(pattern, var_name_str):
                        # Check exclude_pattern if present
                        exclude_pattern = rule_config.get("exclude_pattern")
                        if exclude_pattern and re.match(exclude_pattern, var_name_str):
                            continue  # Skip this rule, try next
                        matched = True
                        rule_to_apply = rule_config.get("all", {})
                else:
                    # Exact match
                    if var_name_str == rule_name.upper():
                        matched = True
                        # Determine which rule to apply based on domain
                        if is_dm and "dm" in rule_config:
                            rule_to_apply = rule_config["dm"]
                        elif not is_dm and "other" in rule_config:
                            rule_to_apply = rule_config["other"]
                        elif "all" in rule_config:
                            rule_to_apply = rule_config["all"]
                        else:
                            continue

                if matched:
                    # Apply source column if specified
                    source_value = rule_to_apply.get("source")
                    if source_value is not None:
                        cell_f = ws.cell(row=row_idx, column=source_column)
                        cell_f.value = self._src_label(source_value)
                        if highlight:
                            cell_f.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    # Helper function to replace variable placeholders
                    def replace_var_placeholders(
                        text: str,
                        vn: str = var_name_str,
                        dom: str = domain,
                    ) -> str:
                        """Replace variable placeholders like {var}, {dtc_var}, {orres_var}."""
                        result = text.replace("{domain}", dom)
                        result = result.replace("{var}", vn)

                        # For xxDY, derive the corresponding xxDTC variable
                        # e.g., AESTDY -> AESTDTC, AEENDY -> AEENDTC
                        if vn.endswith("DY"):
                            dtc_var = vn[:-2] + "DTC"  # Replace DY with DTC
                            result = result.replace("{dtc_var}", dtc_var)

                        # For xxSTAT, derive the corresponding xxORRES variable
                        # e.g., AESTAT -> AEORRES, CMSTAT -> CMORRES
                        if vn.endswith("STAT"):
                            # Get the domain prefix (first 2 chars)
                            domain_prefix = vn[:2] if len(vn) > 4 else dom
                            orres_var = domain_prefix + "ORRES"
                            result = result.replace("{orres_var}", orres_var)

                        return result

                    # Apply transformation if specified (replace existing)
                    transformation_value = rule_to_apply.get("transformation")
                    if transformation_value is not None:
                        transformation_value = replace_var_placeholders(transformation_value)

                        # Skip hyperlinks whose target sheet doesn't exist, e.g.
                        # CODELIST may be absent, and the IG 3.4 template has no
                        # VISIT sheet (only SV). Leaves the template cell untouched.
                        if "HYPERLINK" in transformation_value.upper():
                            missing_ref = next(
                                (
                                    ref
                                    for ref in re.findall(r"#([^!]+)!", transformation_value)
                                    if ref not in self.workbook.sheetnames
                                ),
                                None,
                            )
                            if missing_ref:
                                logger.debug(f"Skipping hyperlink for {var_name_str}: sheet '{missing_ref}' not found")
                                continue

                        cell_h = ws.cell(row=row_idx, column=transformation_column)
                        cell_h.value = transformation_value
                        if highlight:
                            cell_h.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        # Add underline for HYPERLINK formulas
                        if "HYPERLINK" in transformation_value.upper():
                            # Get current font properties and add underline
                            current_font = cell_h.font
                            cell_h.font = Font(
                                name=current_font.name if current_font else None,
                                size=current_font.size if current_font else None,
                                bold=current_font.bold if current_font else None,
                                italic=current_font.italic if current_font else None,
                                color=current_font.color if current_font else None,
                                underline="single",
                            )

                    # Apply append_note if specified (append to existing)
                    append_note = rule_to_apply.get("append_note")
                    if append_note is not None:
                        append_note = replace_var_placeholders(append_note)
                        cell_h = ws.cell(row=row_idx, column=transformation_column)
                        existing_value = cell_h.value or ""
                        # Append note with semicolon + newline separator
                        if existing_value:
                            new_value = f"{existing_value};\n{append_note}"
                        else:
                            new_value = append_note
                        cell_h.value = new_value
                        if highlight:
                            cell_h.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    updated_count += 1
                    logger.debug(f"Applied fixed rule '{rule_name}' to {sheet_name}.{var_name_str}")
                    break  # Only apply first matching rule

        if updated_count > 0:
            self.modified_sheets.add(sheet_name)
            logger.info(f"Applied {updated_count} fixed variable rule(s) in '{sheet_name}'")
        return updated_count

    def update_variable_order_column(
        self, sheet_name: str, start_row: int = 14, order_column: int = 10, type_column: int = 9
    ) -> None:
        """Update the variable order column (J column) with formula =ROW()-13 for SUPP rows only.

        This method updates only rows where column I (type_column) contains "SUPP".

        Args:
            sheet_name: Name of the worksheet
            start_row: Starting row for updates (default: 14, first data row)
            order_column: Column index for variable order (default: 10, column J)
            type_column: Column index for variable type (default: 9, column I)

        Raises:
            ValueError: If sheet name doesn't exist

        Examples:
            >>> writer.update_variable_order_column("AE")  # Updates J column for SUPP rows in AE sheet
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Track that this sheet has been modified
        self.modified_sheets.add(sheet_name)

        updated_count = 0

        # Iterate through rows starting from start_row
        for row_idx in range(start_row, ws.max_row + 1):
            # Check if column I (type_column) contains "SUPP"
            type_value = ws.cell(row=row_idx, column=type_column).value
            if type_value and str(type_value).strip().upper() == "SUPP":
                # Set the formula for SUPP rows only
                cell = ws.cell(row=row_idx, column=order_column)
                cell.value = "=ROW()-13"
                updated_count += 1

        logger.info(
            f"Updated {updated_count} SUPP variable order formulas in '{sheet_name}' "
            f"column {order_column} starting from row {start_row}"
        )

    def update_content_f_column(
        self,
        content_sheet_name: str = "CONTENT",
        domain_sources: dict | None = None,
        f_column: int = 6,
        e_column: int = 5,
        domain_column: int = 1,
        start_row: int = 2,
        highlight: bool = True,
    ) -> int:
        """Update F column (source references) in CONTENT sheet.

        Only domains present in *domain_sources* (derived from the ALS file)
        are touched; template-original rows are left unchanged.

        For each domain row the F value is built as:
          1. Source table names from ALS (``metadata_table`` values)
          2. ``SDTM.DM`` (for every domain except DM itself)
          3. ``SDTM.SV`` (for findings domains, where column E is a findings
             label — "发现" in v3.2 or "Findings" in IG 3.4)

        After building F for main domain rows, every ``SUPPxx`` row copies its
        parent domain's F value so they stay in sync.

        Args:
            content_sheet_name: Name of the CONTENT sheet (default: "CONTENT")
            domain_sources: ``{domain_upper: {table1, table2, ...}}`` built from
                ALS records.  If *None*, no updates are performed.
            f_column: Column index for F column (default: 6)
            e_column: Column index for E column / class (default: 5)
            domain_column: Column index for domain name (default: 1, column A)
            start_row: Starting row for updates (default: 2, skip header)
            highlight: If True, apply yellow background to updated cells
        """
        if not domain_sources:
            logger.info("No domain_sources provided, skipping CONTENT F column update")
            return 0

        if content_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{content_sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[content_sheet_name]
        self.modified_sheets.add(content_sheet_name)

        # Pass 1: update main domain rows, remember computed F per domain
        domain_f_values: dict[str, str] = {}
        updated_count = 0

        for row_idx in range(start_row, ws.max_row + 1):
            domain_value = ws.cell(row=row_idx, column=domain_column).value
            if not domain_value:
                continue
            domain_str = str(domain_value).strip().upper()

            if not (len(domain_str) == 2 and domain_str.isalpha()):
                continue
            if domain_str not in domain_sources:
                continue

            cell = ws.cell(row=row_idx, column=f_column)
            current_value = str(cell.value or "").strip()

            # Replace template value with ALS-derived sources
            parts: list[str] = []
            seen: set[str] = set()

            for tbl in sorted(domain_sources[domain_str]):
                if tbl not in seen:
                    parts.append(tbl)
                    seen.add(tbl)

            # SDTM.DM for every domain except DM
            if domain_str != "DM" and "SDTM.DM" not in seen:
                parts.append("SDTM.DM")
                seen.add("SDTM.DM")

            # SDTM.SV for findings domains
            class_value = ws.cell(row=row_idx, column=e_column).value
            if class_value and str(class_value).strip() in FINDINGS_CLASS_LABELS and "SDTM.SV" not in seen:
                parts.append("SDTM.SV")
                seen.add("SDTM.SV")

            new_value = ", ".join(parts)
            domain_f_values[domain_str] = new_value

            if new_value != current_value:
                cell.value = new_value
                if highlight:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                updated_count += 1
                logger.debug(f"Updated F column for '{domain_str}' at row {row_idx}: {new_value}")

        # Pass 2: copy F value to SUPPxx rows
        for row_idx in range(start_row, ws.max_row + 1):
            domain_value = ws.cell(row=row_idx, column=domain_column).value
            if not domain_value:
                continue
            domain_str = str(domain_value).strip().upper()
            if not domain_str.startswith("SUPP"):
                continue
            parent = domain_str[4:]  # "SUPPAE" -> "AE"
            if parent not in domain_f_values:
                continue
            cell = ws.cell(row=row_idx, column=f_column)
            cell.value = domain_f_values[parent]
            if highlight:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            updated_count += 1
            logger.debug(f"Copied F column from '{parent}' to '{domain_str}' at row {row_idx}")

        logger.info(f"Updated {updated_count} F column value(s) in '{content_sheet_name}'")
        return updated_count

    def update_domain_source_column(
        self,
        sheet_name: str,
        start_row: int = 14,
        source_column: int = 6,
        transformation_column: int = 8,
        variable_name_column: int = 1,
        highlight: bool = True,
    ) -> int:
        """Update F column (Source) based on transformation definition.

        Rules (applied only to rows that have a non-empty transformation):
        - xxTESTCD variables -> F = "指定"
        - Variables whose F is already "指定" or "衍生" -> skip (preserve
          template / fixed-variable-rule values)
        - All other variables with a transformation -> F = "CRF"

        Args:
            sheet_name: Name of the worksheet (domain sheet like "AE", "CM")
            start_row: Starting row for updates (default: 14, first data row)
            source_column: Column index for Source (default: 6, column F)
            transformation_column: Column index for transformation definition (default: 8, column H)
            variable_name_column: Column index for variable name (default: 1, column A)
            highlight: If True, apply yellow background to updated cells (default: True)

        Raises:
            ValueError: If sheet name doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        updated_count = 0

        for row_idx in range(start_row, ws.max_row + 1):
            transformation_value = ws.cell(row=row_idx, column=transformation_column).value
            variable_name_value = ws.cell(row=row_idx, column=variable_name_column).value

            if not transformation_value:
                continue

            var_name_str = str(variable_name_value).strip().upper() if variable_name_value else ""

            cell = ws.cell(row=row_idx, column=source_column)
            current_source = str(cell.value).strip() if cell.value else ""

            if var_name_str.endswith("TESTCD"):
                cell.value = self._src_label("指定")
                if highlight:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                updated_count += 1
                logger.debug(f"Set F column to assigned-label for TESTCD variable at row {row_idx}")
            elif current_source in self._preserve_source_labels():
                continue
            else:
                cell.value = "CRF"
                if highlight:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                updated_count += 1
                logger.debug(f"Set F column to 'CRF' at row {row_idx}")

        if updated_count > 0:
            self.modified_sheets.add(sheet_name)
            logger.info(f"Updated {updated_count} F column value(s) in '{sheet_name}'")
        return updated_count

    def add_nonstandard_domain_to_content(
        self, domain: str, content_sheet_name: str = "CONTENT", reference_row: int = 52
    ) -> int:
        """Add a non-standard domain to the CONTENT sheet with yellow highlight.

        This method adds a new row at the end of the domain list for domains
        that exist in ALS but not in the SDTM template.

        Args:
            domain: Domain name to add (e.g., "XX")
            content_sheet_name: Name of the CONTENT sheet (default: "CONTENT")
            reference_row: Row number to copy formatting from (default: 52)

        Raises:
            ValueError: If CONTENT sheet doesn't exist

        Examples:
            >>> writer.add_nonstandard_domain_to_content("XX")
        """
        if content_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{content_sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[content_sheet_name]

        # Idempotency: skip if this non-standard domain is already present in the
        # CONTENT sheet (e.g. on a re-run), so no duplicate row is appended.
        target_domain = domain.upper()
        for row_idx in range(1, ws.max_row + 1):
            existing = ws.cell(row=row_idx, column=1).value
            if existing and str(existing).strip().upper() == target_domain:
                logger.info(
                    f"Non-standard domain '{target_domain}' already in '{content_sheet_name}' "
                    f"at row {row_idx}; skipping duplicate insert"
                )
                return 0

        # Find the last domain row (before any footer/notes)
        last_domain_row = None
        for row_idx in range(ws.max_row, 0, -1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # Check if it's a valid domain name (2-char or SUPP*)
                if (len(cell_str) == 2 and cell_str.isalpha()) or cell_str.startswith("SUPP"):
                    last_domain_row = row_idx
                    break

        if last_domain_row is None:
            logger.warning(f"Could not find last domain row in '{content_sheet_name}'")
            return 0

        # Insert new row after the last domain
        insert_row_idx = last_domain_row + 1
        self.insert_row(
            sheet_name=content_sheet_name, row_index=insert_row_idx, amount=1, copy_format_from_row=reference_row
        )

        # Fill the row with domain info
        cell_a = ws.cell(row=insert_row_idx, column=1)
        cell_a.value = domain.upper()

        # Apply yellow highlight to indicate non-standard domain
        cell_a.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Add description in column B
        ws.cell(row=insert_row_idx, column=2).value = f"非标准域 {domain.upper()} (请确认)"
        ws.cell(row=insert_row_idx, column=2).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        # Add SDTM.DM to F column
        ws.cell(row=insert_row_idx, column=6).value = "SDTM.DM"

        logger.info(
            f"Added non-standard domain '{domain.upper()}' to '{content_sheet_name}' "
            f"at row {insert_row_idx} with yellow highlight"
        )
        return 1

    def add_content_link_to_domain(self, sheet_name: str, highlight: bool = True) -> int:
        """Replace existing CONTENT hyperlink with formula in a domain sheet.

        This finds the existing "CONTENT" cell and replaces it with:
        =HYPERLINK("#CONTENT!$A"&MATCH(LEFT(A1,IF(A15="DOMAIN",2,6)),CONTENT!A:A,0),"CONTENT")

        Args:
            sheet_name: Name of the domain worksheet
            highlight: If True, apply yellow background (default: True)

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Find the existing CONTENT cell (search from bottom up)
        content_row = None
        for row_idx in range(ws.max_row, 0, -1):
            cell = ws.cell(row=row_idx, column=1)
            cell_value = cell.value
            if cell_value:
                cell_str = str(cell_value).strip().upper()
                # Check if this cell contains "CONTENT" (text or hyperlink)
                if cell_str == "CONTENT":
                    content_row = row_idx
                    break

        if content_row is None:
            logger.warning(f"No existing CONTENT link found in '{sheet_name}', skipping")
            return 0

        # Formula: =IFERROR(HYPERLINK("#CONTENT!$A"&MATCH(LEFT(A1,IF(A15="DOMAIN",2,6)),CONTENT!A:A,0),"CONTENT"),"CONTENT")
        # Use IFERROR to handle cases where MATCH doesn't find a result
        formula = (
            '=IFERROR(HYPERLINK("#CONTENT!$A"&MATCH(LEFT(A1,IF(A15="DOMAIN",2,6)),CONTENT!A:A,0),"CONTENT"),"CONTENT")'
        )

        cell = ws.cell(row=content_row, column=1)
        # Clear existing hyperlink if any
        cell.hyperlink = None
        cell.value = formula

        if highlight:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        self.modified_sheets.add(sheet_name)
        logger.info(f"Replaced CONTENT link with formula in '{sheet_name}' at row {content_row}")
        return 1

    def update_domain_sort_key_formula(
        self,
        sheet_name: str,
        content_sheet_name: str = "CONTENT",
        sort_key_row: int = 10,
        sort_key_col: int = 4,
        highlight: bool = True,
    ) -> int:
        """Set cell D10 in a domain sheet to a formula that looks up the
        sorting variables from the CONTENT sheet's H column.

        Formula produced::

            =IFERROR(INDEX(CONTENT!$H:$H, MATCH("{domain}", CONTENT!$A:$A, 0)), "")

        Args:
            sheet_name: Domain sheet name (e.g. "AE")
            content_sheet_name: Name of the CONTENT sheet
            sort_key_row: Row containing the sort-key cell (default 10)
            sort_key_col: Column of the sort-key cell (default 4 = D)
            highlight: Apply yellow background
        """
        if sheet_name not in self.workbook.sheetnames:
            return 0
        ws = self.workbook[sheet_name]
        domain = sheet_name.upper()
        formula = f'=IFERROR(INDEX({content_sheet_name}!$H:$H,MATCH("{domain}",{content_sheet_name}!$A:$A,0)),"")'
        cell = ws.cell(row=sort_key_row, column=sort_key_col)
        cell.value = formula
        if highlight:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.modified_sheets.add(sheet_name)
        logger.debug(f"Set D{sort_key_row} sort-key formula in '{sheet_name}'")
        return 1

    def add_external_coding_variables(
        self, sheet_name: str, variables: list, variable_type: str = "SUPP", highlight: bool = True
    ) -> int:
        """Add external coding file related variables to a domain sheet.

        Each variable dict should have keys:
        - name: Variable name (A column)
        - label: Variable label (B column)
        - type: Data type, "Char" or "Num" (C column)
        - length: Length like "$200" (D column, optional)
        - controlled_terms: Controlled terms like "MedDRA" or "WHODRUG" (E column, optional)
        - source: Source like "指定" (F column)
        - core: Core status like "期望", "必需" (G column, optional)
        - transformation: Transformation definition (H column)

        Args:
            sheet_name: Name of the domain worksheet
            variables: List of variable dictionaries
            variable_type: Type for I column ("SUPP" or "SDTM", default: "SUPP")
            highlight: If True, apply yellow background (default: True)

        Idempotent: variables whose name already exists in the sheet are skipped,
        so re-running the pipeline never appends duplicate rows. Each such skip
        is queued on :attr:`pending_skips` (``variable_already_present``) so it
        stays visible in the job result instead of hiding behind
        ``attempted == written``.

        Returns:
            The number of variables actually inserted (0 if all already existed).

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Find the last row with type matching variable_type or last SDTM row
        last_row = 14  # Default start row
        for row_idx in range(ws.max_row, 13, -1):
            cell_type = ws.cell(row=row_idx, column=9).value
            if cell_type:
                type_str = str(cell_type).strip().upper()
                if type_str == variable_type.upper() or type_str == "SDTM":
                    last_row = row_idx
                    break

        # Idempotency: don't re-insert a variable that already exists in the
        # sheet (e.g. on a re-run), which would otherwise append duplicate rows.
        existing_names: set[str] = set()
        for row_idx in range(14, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=1).value
            if name_val:
                existing_names.add(str(name_val).strip().upper())

        # Insert rows for each variable
        insert_row = last_row + 1
        inserted = 0

        for var in variables:
            var_name = str(var.get("name", "")).strip().upper()
            if var_name and var_name in existing_names:
                logger.debug(f"External coding variable '{var_name}' already present in '{sheet_name}', skipping")
                self._record_pending_skip(
                    code="variable_already_present",
                    operation="add_external_coding_variables",
                    sheet=sheet_name,
                    variable=var_name,
                )
                continue
            # Copy format from previous row
            copy_from_row = insert_row - 1 if insert_row > 14 else 14
            self.insert_row(sheet_name=sheet_name, row_index=insert_row, amount=1, copy_format_from_row=copy_from_row)

            # Fill column data
            # A: Variable Name
            cell_a = ws.cell(row=insert_row, column=1)
            cell_a.value = var.get("name", "")
            if highlight:
                cell_a.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # B: Variable Label
            cell_b = ws.cell(row=insert_row, column=2)
            cell_b.value = var.get("label", "")
            if highlight:
                cell_b.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # C: Type (Char/Num)
            cell_c = ws.cell(row=insert_row, column=3)
            cell_c.value = var.get("type", "")
            if highlight:
                cell_c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # D: Length — write as numeric so the cell number format renders correctly
            length_val = var.get("length", "")
            cell_d = ws.cell(row=insert_row, column=4)
            try:
                cell_d.value = int(length_val) if length_val else ""
            except (ValueError, TypeError):
                cell_d.value = length_val
            ref_cell_d = ws.cell(row=copy_from_row, column=4)
            if ref_cell_d.number_format:
                cell_d.number_format = ref_cell_d.number_format
            if highlight and length_val:
                cell_d.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # E: Controlled Terms
            cell_e = ws.cell(row=insert_row, column=5)
            cell_e.value = var.get("controlled_terms", "")
            if highlight and var.get("controlled_terms"):
                cell_e.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # F: Source
            cell_f = ws.cell(row=insert_row, column=6)
            cell_f.value = self._src_label(var.get("source", ""))
            if highlight:
                cell_f.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # G: Core
            cell_g = ws.cell(row=insert_row, column=7)
            cell_g.value = var.get("core", "")
            if highlight and var.get("core"):
                cell_g.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # H: Transformation
            cell_h = ws.cell(row=insert_row, column=8)
            transformation_val = var.get("transformation", "")
            cell_h.value = transformation_val
            if highlight:
                cell_h.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Add underline for HYPERLINK formulas
            if transformation_val and "HYPERLINK" in str(transformation_val).upper():
                current_font = cell_h.font
                cell_h.font = Font(
                    name=current_font.name if current_font else None,
                    size=current_font.size if current_font else None,
                    bold=current_font.bold if current_font else None,
                    italic=current_font.italic if current_font else None,
                    color=current_font.color if current_font else None,
                    underline="single",
                )

            # I: Variable Type
            cell_i = ws.cell(row=insert_row, column=9)
            cell_i.value = variable_type
            if highlight:
                cell_i.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # J: Variable Order (formula)
            ws.cell(row=insert_row, column=10).value = "=ROW()-13"

            existing_names.add(var_name)
            insert_row += 1
            inserted += 1

        if inserted:
            self.modified_sheets.add(sheet_name)
        logger.info(f"Added {inserted}/{len(variables)} external coding variables to '{sheet_name}'")
        return inserted

    def update_existing_variables(
        self, sheet_name: str, variables: list, start_row: int = 14, highlight: bool = True
    ) -> int:
        """Update existing variables in a domain sheet with external coding info.

        Each variable dict should have keys:
        - name: Variable name to find (A column match)
        - label: Variable label (B column)
        - type: Data type (C column)
        - length: Length (D column, optional)
        - controlled_terms: Controlled terms (E column, optional)
        - source: Source (F column)
        - core: Core status (G column, optional)
        - transformation: Transformation definition (H column)

        Args:
            sheet_name: Name of the domain worksheet
            variables: List of variable dictionaries
            start_row: Starting row for search (default: 14)
            highlight: If True, apply yellow background (default: True)

        A configured variable that is not present in the sheet cannot be
        updated; that skip is queued on :attr:`pending_skips`
        (``variable_not_found``) so it stays visible in the job result instead
        of hiding behind ``attempted == written``.

        Returns:
            The number of variables actually updated.

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]

        # Build lookup of variable name -> row index
        var_name_to_row = {}
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value:
                var_name_to_row[str(cell_value).strip().upper()] = row_idx

        updated_count = 0
        for var in variables:
            var_name = var.get("name", "").upper()
            if var_name not in var_name_to_row:
                logger.debug(f"Variable '{var_name}' not found in '{sheet_name}', skipping")
                self._record_pending_skip(
                    code="variable_not_found",
                    operation="update_existing_variables",
                    sheet=sheet_name,
                    variable=var_name,
                )
                continue

            row_idx = var_name_to_row[var_name]

            # Update B: Variable Label
            if var.get("label"):
                cell_b = ws.cell(row=row_idx, column=2)
                cell_b.value = var["label"]
                if highlight:
                    cell_b.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update C: Type
            if var.get("type"):
                cell_c = ws.cell(row=row_idx, column=3)
                cell_c.value = var["type"]
                if highlight:
                    cell_c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update D: Length — write as numeric so the cell number format renders correctly
            if var.get("length"):
                cell_d = ws.cell(row=row_idx, column=4)
                try:
                    cell_d.value = int(var["length"])
                except (ValueError, TypeError):
                    cell_d.value = var["length"]
                if highlight:
                    cell_d.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update E: Controlled Terms
            if var.get("controlled_terms"):
                cell_e = ws.cell(row=row_idx, column=5)
                cell_e.value = var["controlled_terms"]
                if highlight:
                    cell_e.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update F: Source
            if var.get("source"):
                cell_f = ws.cell(row=row_idx, column=6)
                cell_f.value = self._src_label(var["source"])
                if highlight:
                    cell_f.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update G: Core
            if var.get("core"):
                cell_g = ws.cell(row=row_idx, column=7)
                cell_g.value = var["core"]
                if highlight:
                    cell_g.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Update H: Transformation
            if var.get("transformation"):
                cell_h = ws.cell(row=row_idx, column=8)
                transformation_val = var["transformation"]
                cell_h.value = transformation_val
                if highlight:
                    cell_h.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                # Add underline for HYPERLINK formulas
                if "HYPERLINK" in str(transformation_val).upper():
                    current_font = cell_h.font
                    cell_h.font = Font(
                        name=current_font.name if current_font else None,
                        size=current_font.size if current_font else None,
                        bold=current_font.bold if current_font else None,
                        italic=current_font.italic if current_font else None,
                        color=current_font.color if current_font else None,
                        underline="single",
                    )

            updated_count += 1
            logger.debug(f"Updated variable '{var_name}' in '{sheet_name}' at row {row_idx}")

        if updated_count > 0:
            self.modified_sheets.add(sheet_name)
            logger.info(f"Updated {updated_count} variables in '{sheet_name}'")
        return updated_count

    def set_column_wrap_text(self, sheet_name: str, column: int = 8, start_row: int = 14) -> int:
        """Set wrap_text=True for a column in a domain sheet.

        This enables automatic line wrapping for cells in the specified column,
        so multi-line content is displayed properly in Excel.

        Args:
            sheet_name: Name of the worksheet
            column: Column index to set wrap_text (default: 8, column H)
            start_row: Starting row for updates (default: 14, first data row)

        Raises:
            ValueError: If sheet doesn't exist

        Examples:
            >>> writer.set_column_wrap_text("AE")  # Set H column wrap_text in AE sheet
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws: Worksheet = self.workbook[sheet_name]
        updated_count = _fmt.set_column_wrap_text(ws, column=column, start_row=start_row)

        logger.info(f"Set wrap_text for {updated_count} cells in column {column} of '{sheet_name}'")
        return updated_count

    def write_codelist_records(
        self, codelist_records: list, sheet_name: str = "CODELIST", start_row: int = 3, highlight: bool = True
    ) -> StageWriteResult:
        """Write CODELIST records to CODELIST sheet.

        Merges with existing template content:
          - If a (domain, testcd_var, testcd_value) row already exists, only
            fill in columns I and J when they are empty.
          - New entries for a domain that already has rows are appended right
            after the last existing row of that domain (keeps domains grouped).
          - Completely new domains are appended after the last occupied row.

        Column mapping:
            A (1): Domain
            B (2): TESTCD variable name (e.g., EGTESTCD)
            E (5): Data type (Char)
            F (6): TESTCD value (e.g., QT)
            I (9): Source variable name
            J (10): EDC variable name (metadata_variable)

        Args:
            codelist_records: List of CodelistRecord objects
            sheet_name: Name of CODELIST sheet (default: "CODELIST")
            start_row: First data row (default: 3, after two header rows)
            highlight: If True, apply yellow background (default: True)

        Returns:
            A :class:`StageWriteResult`; each input record counts as one attempt
            and is ``written`` only when it actually mutated the workbook (a new
            row was inserted, or a previously-blank I/J cell was filled). A
            record whose row already exists with I/J populated is a no-op and is
            recorded as ``skipped`` (``codelist_unchanged``) — never as a write.
        """
        result = StageWriteResult(stage=STAGE_CODELIST_RECORDS)

        if not codelist_records:
            logger.info("No CODELIST records to write")
            return result

        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook, skipping CODELIST write")
            for _rec in codelist_records:
                result.record_skipped(
                    WriteIssue(
                        code="sheet_not_found",
                        stage=STAGE_CODELIST_RECORDS,
                        operation="write_codelist",
                        sheet=sheet_name,
                    )
                )
            return result

        ws: Worksheet = self.workbook[sheet_name]

        # Set column J header
        ws.cell(row=2, column=10).value = "EDC VAR"

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # --- Scan existing rows ------------------------------------------
        # existing_keys: {(domain, testcd_var, testcd_value)} -> row_idx
        # domain_last_row: {domain} -> last row_idx for that domain block
        existing_keys: dict[tuple[str, str, str], int] = {}
        domain_last_row: dict[str, int] = {}
        global_last_row = start_row - 1

        for row_idx in range(start_row, ws.max_row + 1):
            col_a = ws.cell(row=row_idx, column=1).value
            if not col_a:
                continue
            domain_val = str(col_a).strip().upper()
            testcd_var_val = str(ws.cell(row=row_idx, column=2).value or "").strip().upper()
            testcd_val = str(ws.cell(row=row_idx, column=6).value or "").strip().upper()

            existing_keys[(domain_val, testcd_var_val, testcd_val)] = row_idx
            domain_last_row[domain_val] = row_idx
            if row_idx > global_last_row:
                global_last_row = row_idx

        updated_count = 0
        inserted_count = 0

        # --- Process records ----------------------------------------------
        # Group by domain to preserve domain ordering
        from collections import defaultdict

        by_domain: dict[str, list] = defaultdict(list)
        for rec in codelist_records:
            by_domain[rec.domain.upper()].append(rec)

        for domain, recs in by_domain.items():
            for rec in recs:
                # Pre-validate BEFORE mutating: an invalid value must not leave
                # a half-filled inserted row (or a lone I without J) behind.
                if contains_illegal_characters(*rec.to_row_data().values()):
                    result.record_error(
                        WriteIssue(
                            code="illegal_characters",
                            stage=STAGE_CODELIST_RECORDS,
                            operation="write_codelist",
                            sheet=sheet_name,
                        )
                    )
                    logger.warning("Skipped CODELIST record for domain '%s': value contains illegal characters", domain)
                else:
                    key = (rec.domain.upper(), rec.testcd_var.upper(), rec.testcd_value.upper())
                    rec_mutations = 0

                    if key in existing_keys:
                        # Row already exists — only fill I/J if currently empty
                        exist_row = existing_keys[key]
                        for col, value in [(9, rec.source_variable), (10, rec.metadata_variable)]:
                            if not value:
                                continue
                            cell = ws.cell(row=exist_row, column=col)
                            if not cell.value:
                                cell.value = value
                                if highlight:
                                    cell.fill = yellow
                                updated_count += 1
                                rec_mutations += 1
                    else:
                        # New entry — insert a row after the domain's last row
                        if domain in domain_last_row:
                            target_row = domain_last_row[domain] + 1
                        else:
                            target_row = global_last_row + 1

                        # Insert a blank row so existing content below is not overwritten
                        ws.insert_rows(target_row)

                        # Shift all tracked row indices that are >= target_row
                        existing_keys = {k: (v + 1 if v >= target_row else v) for k, v in existing_keys.items()}
                        domain_last_row = {d: (r + 1 if r >= target_row else r) for d, r in domain_last_row.items()}
                        if global_last_row >= target_row:
                            global_last_row += 1

                        row_data = rec.to_row_data()
                        for col, value in row_data.items():
                            if value:
                                cell = ws.cell(row=target_row, column=col)
                                cell.value = value
                                if highlight:
                                    cell.fill = yellow

                        # Update bookkeeping
                        existing_keys[key] = target_row
                        domain_last_row[domain] = target_row
                        if target_row > global_last_row:
                            global_last_row = target_row
                        inserted_count += 1
                        rec_mutations += 1

                    # Count as written only when this record actually mutated the
                    # workbook (inserted a new row or filled a previously-blank
                    # cell). An entry whose row already existed with I/J populated
                    # is a genuine no-op and must not inflate ``written``.
                    if rec_mutations > 0:
                        result.record_written()
                    else:
                        result.record_skipped(
                            WriteIssue(
                                code="codelist_unchanged",
                                stage=STAGE_CODELIST_RECORDS,
                                operation="write_codelist",
                                sheet=sheet_name,
                            )
                        )

        self.modified_sheets.add(sheet_name)
        logger.info(
            f"CODELIST: updated {updated_count} existing cell(s), "
            f"inserted {inserted_count} new row(s) in '{sheet_name}'"
        )
        return result
