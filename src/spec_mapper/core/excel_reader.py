"""Excel file reader with support for ALS and template files."""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..models.als_record import ALSRecord
from ..models.template_record import TemplateRecord
from ..utils.config_loader import ConfigLoader

logger = logging.getLogger(__name__)


class ExcelReader:
    """Read and parse Excel files for ALS and template data.

    This reader supports:
    - ALS files with default sheet "Sheet1"
    - Template files with 2-character sheet names
    - Template files start reading from row 13 (index 12)

    Examples:
        >>> reader = ExcelReader("alsfile/data.xlsx")
        >>> als_records = reader.read_als_records()
        >>> len(als_records)
        150
    """

    def __init__(self, file_path: str | Path, config: ConfigLoader | None = None):
        """Initialize Excel reader.

        Args:
            file_path: Path to the Excel file
            config: Configuration loader instance. If None, creates new instance.

        Raises:
            FileNotFoundError: If the file does not exist
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self.config = config or ConfigLoader()
        logger.info(f"Initialized ExcelReader for {self.file_path}")

    def read_als_records(self, sheet_name: str | None = None) -> list[ALSRecord]:
        """Read ALS records from the Excel file.

        Args:
            sheet_name: Sheet name to read. If None, uses default from config ("Sheet1")

        Returns:
            List of ALSRecord objects

        Raises:
            ValueError: If required columns are missing
        """
        if sheet_name is None:
            sheet_name = self.config.get("als_defaults.sheet_name", "Sheet1")

        # Get column mappings
        columns = self.config.get_als_columns()
        header_row = self.config.get("als_defaults.header_row", 0)

        logger.info(f"Reading ALS data from sheet '{sheet_name}'")

        try:
            # Read Excel file with pandas
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=header_row,
            )

            # Build column mapping with fallback support (Chinese -> English -> extra)
            def find_column(primary_key: str, *fallback_keys: str) -> str | None:
                """Find column by primary key, trying each fallback key in order."""
                primary = columns.get(primary_key)
                if primary and primary in df.columns:
                    return primary
                for fk in fallback_keys:
                    fallback = columns.get(fk)
                    if fallback and fallback in df.columns:
                        return fallback
                return None

            # Resolve actual column names with fallbacks
            table_col = find_column("table", "table_en", "table_en2")
            variable_col = find_column("variable", "variable_en", "variable_en2", "variable_en3")
            variable_label_col = find_column(
                "variable_label", "variable_label_en", "variable_label_en2", "variable_label_en3"
            )
            sdtm_domain_col = find_column("sdtm_domain", "sdtm_domain_en")
            sdtm_variable_col = find_column("sdtm_variable", "sdtm_variable_en")

            # Optional column for [RAW] reference table name
            metadata_table_col = find_column("metadata_table")

            # Optional columns for codelist handling
            component_type_col = find_column("component_type")
            codelist_name_col = find_column("codelist_name")

            # Check for missing required columns
            missing_cols = []
            if not table_col:
                missing_cols.append(f"{columns.get('table')} or {columns.get('table_en')}")
            if not variable_col:
                missing_cols.append(f"{columns.get('variable')} or {columns.get('variable_en')}")
            if not variable_label_col:
                missing_cols.append(f"{columns.get('variable_label')} or {columns.get('variable_label_en')}")
            if not sdtm_domain_col:
                missing_cols.append(f"{columns.get('sdtm_domain')} or {columns.get('sdtm_domain_en')}")
            if not sdtm_variable_col:
                missing_cols.append(f"{columns.get('sdtm_variable')} or {columns.get('sdtm_variable_en')}")

            if missing_cols:
                raise ValueError(
                    f"Missing required columns in sheet '{sheet_name}': {missing_cols}. "
                    f"Available columns: {list(df.columns)}"
                )

            # Log optional columns status
            optional_cols_info = []
            if metadata_table_col:
                optional_cols_info.append(f"metadata_table={metadata_table_col}")
            if component_type_col:
                optional_cols_info.append(f"component_type={component_type_col}")
            if codelist_name_col:
                optional_cols_info.append(f"codelist_name={codelist_name_col}")

            logger.info(
                f"Using columns: table={table_col}, variable={variable_col}, "
                f"label={variable_label_col}, domain={sdtm_domain_col}, sdtm_var={sdtm_variable_col}"
                + (f", {', '.join(optional_cols_info)}" if optional_cols_info else "")
            )

            records: list[ALSRecord] = []
            format_col = columns.get("format")
            for idx, row in enumerate(df.to_dict("records")):
                try:
                    format_value = self._safe_str(row.get(format_col)) if format_col else None
                    metadata_table_value = self._safe_str(row.get(metadata_table_col)) if metadata_table_col else None
                    component_type_value = self._safe_str(row.get(component_type_col)) if component_type_col else None
                    codelist_name_value = self._safe_str(row.get(codelist_name_col)) if codelist_name_col else None

                    record = ALSRecord(
                        table=self._safe_str(row.get(table_col)),
                        variable=self._safe_str(row.get(variable_col)),
                        variable_label=self._safe_str(row.get(variable_label_col)),
                        format=format_value,
                        sdtm_domain=self._safe_str(row.get(sdtm_domain_col)),
                        sdtm_variable=self._safe_str(row.get(sdtm_variable_col)),
                        row_index=idx,
                        metadata_table=metadata_table_value if metadata_table_value else None,
                        component_type=component_type_value if component_type_value else None,
                        codelist_name=codelist_name_value if codelist_name_value else None,
                    )

                    if record.is_valid():
                        records.append(record)
                    else:
                        logger.debug(f"Skipping invalid record at row {idx}")

                except Exception as e:
                    logger.warning(f"Failed to parse row {idx}: {e}")
                    continue

            logger.info(f"Successfully read {len(records)} ALS records")
            return records

        except Exception as e:
            logger.error(f"Failed to read ALS file: {e}")
            raise

    def read_template_records(self, sheet_names: list[str] | None = None) -> list[TemplateRecord]:
        """Read template records from the Excel file.

        Only processes sheets with 2-character names (e.g., "DM", "AE").
        Header row is at index 12 (row 13 in Excel).

        **Optimized**: Opens workbook once and reuses it for all sheets.

        Args:
            sheet_names: Specific sheet names to read. If None, reads all 2-char sheets.

        Returns:
            List of TemplateRecord objects from all processed sheets

        Raises:
            ValueError: If required columns are missing
        """
        # Get column mappings
        columns = self.config.get_template_columns()
        header_row = self.config.get("template_defaults.header_row", 12)
        min_len = self.config.get("template_defaults.min_sheet_name_length", 2)
        max_len = self.config.get("template_defaults.max_sheet_name_length", 2)

        logger.debug(f"Loading workbook: {self.file_path}")
        wb = load_workbook(self.file_path, read_only=True, data_only=True)

        try:
            if sheet_names is None:
                sheet_names = [name for name in wb.sheetnames if min_len <= len(name) <= max_len]

            logger.info(f"Reading template data from {len(sheet_names)} sheets: {sheet_names}")

            all_records: list[TemplateRecord] = []

            for sheet_name in sheet_names:
                try:
                    records = self._read_single_template_sheet_from_workbook(wb, sheet_name, columns, header_row)
                    all_records.extend(records)
                    logger.debug(f"Read {len(records)} records from sheet '{sheet_name}'")
                except Exception as e:
                    logger.error(f"Failed to read sheet '{sheet_name}': {e}")
                    continue
        finally:
            wb.close()

        logger.info(f"Successfully read {len(all_records)} template records from {len(sheet_names)} sheets")
        return all_records

    def _read_single_template_sheet_from_workbook(
        self, workbook, sheet_name: str, columns: dict[str, str], header_row: int
    ) -> list[TemplateRecord]:
        """Read template records from a single sheet using an existing workbook.

        **Optimized**: Reads directly from workbook object without reopening file.

        Args:
            workbook: openpyxl Workbook object (already opened)
            sheet_name: Name of the sheet to read
            columns: Column name mappings
            header_row: Header row index (0-based, row 13 in Excel = index 12)

        Returns:
            List of TemplateRecord objects from this sheet
        """
        ws = workbook[sheet_name]

        # Read header row (row_index is 1-based in openpyxl)
        header_row_idx = header_row + 1  # Convert 0-based to 1-based
        header_cells = next(iter(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)))

        # Build column index mapping
        col_indices = {}
        for col_idx, cell_value in enumerate(header_cells):
            if cell_value:
                col_indices[str(cell_value).strip()] = col_idx

        # Auto-detect column names (Chinese or English)
        var_name_candidates = [columns.get("variable_name"), columns.get("variable_name_en")]
        trans_def_candidates = [columns.get("transformation_def"), columns.get("transformation_def_en")]

        var_name_col_idx = None
        trans_def_col_idx = None

        # Find variable name column
        for candidate in var_name_candidates:
            if candidate and candidate in col_indices:
                var_name_col_idx = col_indices[candidate]
                break

        # Find transformation definition column
        for candidate in trans_def_candidates:
            if candidate and candidate in col_indices:
                trans_def_col_idx = col_indices[candidate]
                break

        if var_name_col_idx is None:
            raise ValueError(f"Missing required column (tried: {var_name_candidates}) in sheet '{sheet_name}'")

        if trans_def_col_idx is None:
            logger.warning(f"Transformation definition column not found in sheet '{sheet_name}'")
            return []  # Return empty if no transformation column

        # Read data rows (starting from header_row + 1)
        records: list[TemplateRecord] = []
        data_start_row = header_row_idx + 1

        for row_idx, row_cells in enumerate(
            ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
        ):
            try:
                # Get variable name and transformation definition
                var_name = self._safe_str(row_cells[var_name_col_idx]) if var_name_col_idx < len(row_cells) else ""
                trans_def = self._safe_str(row_cells[trans_def_col_idx]) if trans_def_col_idx < len(row_cells) else ""

                # Skip empty variable names
                if not var_name:
                    continue

                record = TemplateRecord(
                    variable_name=var_name,
                    transformation_def=trans_def,
                    row_index=row_idx,  # Already 1-based from openpyxl
                    col_index=trans_def_col_idx + 1,  # Convert to 1-based
                    sheet_name=sheet_name,
                )

                if record.is_valid():
                    records.append(record)

            except Exception as e:
                logger.warning(f"Failed to parse row {row_idx} in sheet '{sheet_name}': {e}")
                continue

        return records

    def get_sheet_names(self) -> list[str]:
        """Get all sheet names from the Excel file.

        Returns:
            List of sheet names
        """
        wb = load_workbook(self.file_path, read_only=True)
        try:
            sheet_names = wb.sheetnames
        finally:
            wb.close()
        return sheet_names

    @staticmethod
    def _safe_str(value: Any) -> str:
        """Safely convert value to string.

        Args:
            value: Any value to convert

        Returns:
            String representation, empty string if None or NaN
        """
        if value is None:
            return ""
        if pd.isna(value):
            return ""
        return str(value).strip()
