"""Excel format utility helpers (openpyxl).

All helpers take an ``openpyxl`` worksheet or workbook object and operate
on it in-place. Logging is the caller's responsibility.
"""

from __future__ import annotations

from collections.abc import Iterable
from copy import copy as copy_obj
from typing import Any

from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def copy_row_format(ws: Worksheet, row_num: int) -> dict[int, dict[str, Any]]:
    """Return a ``{column_index: format_dict}`` snapshot for ``row_num``.

    The snapshot intentionally excludes merged-cell information so
    subsequent :func:`apply_row_format` calls don't propagate merges.
    """
    format_data: dict[int, dict[str, Any]] = {}
    for cell in ws[row_num]:
        if not cell.column:
            continue
        format_data[cell.column] = {
            "font": copy_obj(cell.font) if cell.font else None,
            "fill": copy_obj(cell.fill) if cell.fill else None,
            "border": copy_obj(cell.border) if cell.border else None,
            "alignment": copy_obj(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }
    return format_data


def apply_row_format(ws: Worksheet, row_num: int, format_data: dict[int, dict[str, Any]]) -> None:
    """Apply a previously captured ``format_data`` snapshot to ``row_num``."""
    for col_idx, fmt in format_data.items():
        cell = ws.cell(row=row_num, column=col_idx)
        if fmt.get("font"):
            cell.font = fmt["font"]
        if fmt.get("fill"):
            cell.fill = fmt["fill"]
        if fmt.get("border"):
            cell.border = fmt["border"]
        if fmt.get("alignment"):
            cell.alignment = fmt["alignment"]
        if fmt.get("number_format"):
            cell.number_format = fmt["number_format"]


def find_merged_ranges_overlapping(ws: Worksheet, start_row: int, end_row: int) -> list[Any]:
    """Return merged cell ranges that overlap rows ``[start_row, end_row]``."""
    overlapping: list[Any] = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= end_row and merged_range.max_row >= start_row:
            overlapping.append(merged_range)
    return overlapping


def unmerge_rows_in_range(ws: Worksheet, start_row: int, amount: int) -> list[str]:
    """Unmerge any merged ranges touching ``[start_row, start_row+amount-1]``.

    Returns the string representations of unmerged ranges (useful for logging).
    """
    if amount <= 0:
        return []
    end_row = start_row + amount - 1
    overlapping = find_merged_ranges_overlapping(ws, start_row, end_row)
    unmerged: list[str] = []
    for merged_range in overlapping:
        range_str = str(merged_range)
        ws.unmerge_cells(range_str)
        unmerged.append(range_str)
    return unmerged


def find_last_row_with_value(
    ws: Worksheet,
    *,
    column: int,
    value: str,
    start_row: int = 1,
    case_insensitive: bool = True,
) -> int | None:
    """Return the last row index where ``column`` equals ``value`` (stripped)."""
    target = value.upper() if case_insensitive else value
    last_row: int | None = None
    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=column).value
        if cell_value is None:
            continue
        candidate = str(cell_value).strip()
        if case_insensitive:
            candidate = candidate.upper()
        if candidate == target:
            last_row = row_idx
    return last_row


def set_column_wrap_text(ws: Worksheet, *, column: int, start_row: int = 1) -> int:
    """Enable ``wrap_text`` on ``column`` from ``start_row`` downward.

    Returns the number of cells updated. Existing alignment options
    (horizontal/vertical/indent/shrink) are preserved.
    """
    updated = 0
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=column)
        current = cell.alignment
        indent_val = 0
        if current and current.indent is not None:
            try:
                indent_val = int(current.indent)
            except (TypeError, ValueError):
                indent_val = 0
        cell.alignment = Alignment(
            horizontal=current.horizontal if current else None,
            vertical=current.vertical if current else None,
            wrap_text=True,
            shrink_to_fit=current.shrink_to_fit if current else None,
            indent=indent_val,
        )
        updated += 1
    return updated


def highlight_sheet_tabs(
    workbook: Workbook,
    *,
    modified_sheets: Iterable[str],
    color: str = "FFFF00",
    clear_existing: bool = True,
) -> int:
    """Highlight ``modified_sheets`` tabs with ``color``.

    When ``clear_existing`` is True, all tab colors are cleared first so
    stale highlights are removed between runs. Returns the number of
    sheets newly highlighted.
    """
    if clear_existing:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.sheet_properties.tabColor:
                sheet.sheet_properties.tabColor = None

    highlighted = 0
    for sheet_name in modified_sheets:
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].sheet_properties.tabColor = color
            highlighted += 1
    return highlighted


__all__ = [
    "apply_row_format",
    "copy_row_format",
    "find_last_row_with_value",
    "find_merged_ranges_overlapping",
    "highlight_sheet_tabs",
    "set_column_wrap_text",
    "unmerge_rows_in_range",
]
