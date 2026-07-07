"""
One-time script: parse SDTMIG_v3.4.xlsx -> data/knowledge_base/standards/sdtmig_v3.4_variables.json

Reads:
  data/knowledge_base/standards/SDTMIG_v3.4.xlsx
    Sheet "Datasets"  -> domain name / class metadata
    Sheet "Variables" -> variable name, label, type, core

Writes:
  data/knowledge_base/standards/sdtmig_v3.4_variables.json
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]

import openpyxl  # noqa: E402

XLSX_PATH = Path("data/knowledge_base/standards/SDTMIG_v3.4.xlsx")
OUT_PATH = Path("data/knowledge_base/standards/sdtmig_v3.4_variables.json")


def main() -> None:
    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)

    # --- Datasets sheet: domain name + class ---
    ws_ds = wb["Datasets"]
    domain_meta: dict[str, dict] = {}
    for row in ws_ds.iter_rows(min_row=2, values_only=True):
        _version, cls, dataset, label = row[0], row[1], row[2], row[3]
        if dataset:
            domain_meta[str(dataset)] = {
                "name": str(label) if label else str(dataset),
                "class": str(cls) if cls else "",
            }

    # --- Variables sheet: variable details ---
    ws_var = wb["Variables"]
    # Header: Version | Variable Order | Class | Dataset Name | Variable Name |
    #         Variable Label | Type | CDISC CT Codelist Code(s) | ... | Core
    domain_vars: dict[str, list[dict]] = defaultdict(list)
    for row in ws_var.iter_rows(min_row=2, values_only=True):
        dataset = row[3]
        var_name = row[4]
        var_label = row[5]
        var_type = row[6]
        # Core is the last non-None value in the row
        core = None
        for v in reversed(row):
            if v is not None:
                core = v
                break
        if dataset and var_name:
            domain_vars[str(dataset)].append(
                {
                    "variable": str(var_name).upper().strip(),
                    "label": str(var_label).strip() if var_label else "",
                    "type": str(var_type).strip() if var_type else "",
                    "core": str(core).strip() if core else "",
                }
            )

    wb.close()

    # --- Build combined result ---
    result: dict[str, dict] = {}

    # Start from Datasets order (canonical)
    for domain in sorted(domain_meta.keys()):
        meta = domain_meta[domain]
        result[domain] = {
            "name": meta["name"],
            "class": meta["class"],
            "variables": domain_vars.get(domain, []),
        }

    # Add any domains that appear in Variables but not Datasets
    for domain in sorted(domain_vars.keys()):
        if domain not in result:
            result[domain] = {
                "name": domain,
                "class": "",
                "variables": domain_vars[domain],
            }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    total_vars = sum(len(v["variables"]) for v in result.values())
    print(f"Written: {OUT_PATH}")
    print(f"Domains: {len(result)}, Variables: {total_vars}")

    # Spot checks
    for dom in ["AE", "LB", "GF", "OE", "MK", "CP", "XU", "APMH"]:
        entry = result.get(dom)
        if entry:
            sample = [v["variable"] for v in entry["variables"][:5]]
            print(f"  {dom} ({len(entry['variables'])} vars): {sample}")
        else:
            print(f"  {dom}: NOT FOUND")


if __name__ == "__main__":
    main()
