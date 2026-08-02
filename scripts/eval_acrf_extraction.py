#!/usr/bin/env python
"""Score aCRF/eCRF PDF extraction against an ALS ground-truth workbook.

Answers one question: of the (form, field) pairs the ALS says exist, how many
did the deterministic extractor recover, and what did it invent?

    # 1. discover the sheet and column names in the ground-truth workbook
    python scripts/eval_acrf_extraction.py --truth mapping.xlsx --list-sheets
    python scripts/eval_acrf_extraction.py --truth mapping.xlsx --sheet ALS2SDTM --list-columns

    # 2. score one project
    python scripts/eval_acrf_extraction.py \
        --pdf your_acrf.pdf --truth mapping.xlsx \
        --sheet ALS2SDTM --form-col 表名 --field-col 变量名 --header-row 2

    # 3. drill into one form's misses
    ... --detail 生命体征

Scoring notes:

* Forms are matched by name. A trailing bookmark OID ("Adverse Events (AE1)")
  is ignored, since the ALS stores the bare form name.
* Comparison is punctuation- and case-insensitive: an ALS export routinely
  loses opening curly quotes, so 若选择"否"… and 若选择否"… are one label.
* Only fields of *matched* forms are scored, so a form-name mismatch shows up
  in the form counts instead of silently wrecking precision.
* Recall is bounded by the ground truth. An ALS carries hidden/system fields a
  CRF never prints, and "check all that apply" groups often become one ALS
  field per choice, which the extractor deliberately drops as answer text.
  Read a miss list before treating a gap as a defect.

Nothing is written to disk; the PDF is only read.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path

_DEFAULT_REPO = Path(__file__).resolve().parents[1]

# Punctuation, spacing and quotes carry no meaning for label identity.
_NOISE_RE = re.compile(r"[\s　()（）\[\]【】:：,，.。;；/、\-—_*#'\"“”‘’`]+")
_OID_SUFFIX_RE = re.compile(r"\s*\([A-Za-z0-9_\-]+\)\s*$")


def key(text: object) -> str:
    """Comparison key for a field label."""
    if text is None:
        return ""
    return _NOISE_RE.sub("", unicodedata.normalize("NFKC", str(text))).strip().casefold()


def form_key(text: object) -> str:
    """Comparison key for a form name, ignoring a trailing bookmark OID."""
    return key(_OID_SUFFIX_RE.sub("", unicodedata.normalize("NFKC", str(text or ""))))


# --------------------------------------------------------------------------
# ground truth
# --------------------------------------------------------------------------
def _sheet_rows(xlsx: Path, sheet: str) -> list[tuple[object, ...]]:
    import openpyxl

    book = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise SystemExit(f"sheet {sheet!r} not found. Available: {book.sheetnames}")
        worksheet = book[sheet]
        # Some exporters write a bogus <dimension> ("A1:A1"), which read-only
        # mode trusts and then yields a single row. Re-scan instead.
        worksheet.reset_dimensions()
        return list(worksheet.iter_rows(values_only=True))
    finally:
        book.close()


def load_truth(xlsx: Path, sheet: str, form_col: str, field_col: str, header_row: int) -> dict[str, list[str]]:
    """``{form_name: [field_label, ...]}`` in sheet order."""
    rows = _sheet_rows(xlsx, sheet)
    if len(rows) <= header_row:
        raise SystemExit(f"sheet {sheet!r} has no data below header row {header_row}")

    header = [str(c).strip() if c is not None else "" for c in rows[header_row - 1]]
    index = {name: i for i, name in enumerate(header) if name}
    for wanted in (form_col, field_col):
        if wanted not in index:
            raise SystemExit(f"column {wanted!r} not in row {header_row} of {sheet!r}. Found: {header}")

    f_col, v_col = index[form_col], index[field_col]
    truth: dict[str, list[str]] = {}
    for row in rows[header_row:]:
        form = str(row[f_col]).strip() if f_col < len(row) and row[f_col] is not None else ""
        field = str(row[v_col]).strip() if v_col < len(row) and row[v_col] is not None else ""
        if not form:
            continue
        truth.setdefault(form, [])
        if field:
            truth[form].append(field)
    return truth


# --------------------------------------------------------------------------
# extraction
# --------------------------------------------------------------------------
def run_extract(repo: Path, pdf: Path) -> tuple[dict[str, list[str]], dict[str, object]]:
    sys.path.insert(0, str(repo))
    from src.processors.acrf import extract_acrf

    result = extract_acrf(str(pdf))
    forms: dict[str, list[str]] = {}
    for record in result.records:
        forms.setdefault(record.annotation_table, []).append(record.annotation_variable)
    return forms, {"stats": result.stats, "warnings": result.warnings, "skipped": result.skipped_forms}


# --------------------------------------------------------------------------
# scoring
# --------------------------------------------------------------------------
def score(pred: dict[str, list[str]], truth: dict[str, list[str]]) -> dict[str, object]:
    truth_by_key = {form_key(name): name for name in truth}
    matched = {name: truth_by_key[form_key(name)] for name in pred if form_key(name) in truth_by_key}

    tp = fp = fn = 0
    per_form = []
    for predicted_name, truth_name in matched.items():
        got = {key(f) for f in pred[predicted_name] if key(f)}
        want = {key(f) for f in truth[truth_name] if key(f)}
        tp, fp, fn = tp + len(got & want), fp + len(got - want), fn + len(want - got)
        per_form.append(
            {
                "form": truth_name,
                "truth_n": len(want),
                "pred_n": len(got),
                "hit": len(got & want),
                "missed": sorted(want - got),
                "extra": sorted(got - want),
            }
        )

    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    return {
        "forms_pred": len(pred),
        "forms_truth": len(truth),
        "forms_matched": len(matched),
        "forms_unmatched": sorted(n for n in pred if form_key(n) not in truth_by_key),
        "forms_missing": sorted(set(truth) - set(matched.values())),
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "precision": precision,
        "recall": recall,
        "f1": (2 * precision * recall / (precision + recall)) if precision + recall else 0.0,
        "per_form": per_form,
    }


def report(pdf: Path, result: dict[str, object], meta: dict[str, object], detail: str | None, limit: int) -> None:
    print(f"\n{'=' * 78}\n{pdf.name}\n{'=' * 78}")
    print(f"stats  : {meta['stats']}")
    print(f"forms  : extracted={result['forms_pred']} truth={result['forms_truth']} matched={result['forms_matched']}")
    print(
        f"fields : TP={result['tp']} FP={result['fp']} FN={result['fn']}  "
        f"P={result['precision']:.3f} R={result['recall']:.3f} F1={result['f1']:.3f}"
    )

    for title, names in (
        ("extracted forms with no ground-truth match", result["forms_unmatched"]),
        ("ground-truth forms never produced", result["forms_missing"]),
    ):
        if names:
            print(f"\n-- {title} ({len(names)}):")
            for name in names[:limit]:
                print(f"   {name}")
            if len(names) > limit:
                print(f"   … {len(names) - limit} more (raise --limit)")

    worst = sorted(result["per_form"], key=lambda r: (r["hit"] / r["truth_n"]) if r["truth_n"] else 1.0)
    print(f"\n-- lowest-recall matched forms (showing {min(limit, len(worst))}):")
    for row in worst[:limit]:
        print(f"   {row['form']}: truth={row['truth_n']} extracted={row['pred_n']} hit={row['hit']}")

    if detail:
        hit = [r for r in result["per_form"] if form_key(r["form"]) == form_key(detail)]
        if not hit:
            print(f"\n-- no matched form named {detail!r}")
        for row in hit:
            print(f"\n-- DETAIL {row['form']}: truth={row['truth_n']} extracted={row['pred_n']} hit={row['hit']}")
            print(f"   MISSED ({len(row['missed'])}): {row['missed']}")
            print(f"   EXTRA  ({len(row['extra'])}): {row['extra']}")

    warnings = meta["warnings"]
    if warnings:
        print(f"\n-- extractor warnings ({len(warnings)}):")
        for warning in warnings[:limit]:
            print(f"   {warning}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--pdf", type=Path, help="aCRF/eCRF PDF to extract (must have bookmarks).")
    parser.add_argument("--truth", type=Path, required=True, help="Ground-truth ALS workbook.")
    parser.add_argument("--sheet", help="Ground-truth sheet name.")
    parser.add_argument("--form-col", default="表名", help="Header of the form-name column (default: 表名).")
    parser.add_argument("--field-col", default="变量名", help="Header of the field-name column (default: 变量名).")
    parser.add_argument("--header-row", type=int, default=1, help="1-based row holding the headers (default: 1).")
    parser.add_argument("--detail", help="Print the full miss/extra list for this form.")
    parser.add_argument("--limit", type=int, default=15, help="Max lines per report section (default: 15).")
    parser.add_argument("--repo", type=Path, default=_DEFAULT_REPO, help="Project root to import src from.")
    parser.add_argument("--list-sheets", action="store_true", help="List sheets in --truth and exit.")
    parser.add_argument("--list-columns", action="store_true", help="List headers of --sheet and exit.")
    args = parser.parse_args()

    if args.list_sheets:
        import openpyxl

        book = openpyxl.load_workbook(args.truth, read_only=True)
        print("\n".join(book.sheetnames))
        book.close()
        return

    if args.list_columns:
        if not args.sheet:
            raise SystemExit("--list-columns needs --sheet")
        row = _sheet_rows(args.truth, args.sheet)[args.header_row - 1]
        for i, cell in enumerate(row):
            if cell is not None and str(cell).strip():
                print(f"[{i}] {cell}")
        return

    if not args.pdf or not args.sheet:
        raise SystemExit("--pdf and --sheet are required for scoring (use --list-sheets to explore)")

    truth = load_truth(args.truth, args.sheet, args.form_col, args.field_col, args.header_row)
    pred, meta = run_extract(args.repo.resolve(), args.pdf)
    report(args.pdf, score(pred, truth), meta, args.detail, args.limit)


if __name__ == "__main__":
    main()
